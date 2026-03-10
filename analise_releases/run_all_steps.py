# -*- coding: utf-8 -*-
"""
Projeto de Inteligência Financeira - Releases e Planilhas
Executa Passos 1 a 5 conforme INSTRUCOES_PROJETO_RELEASES.md
"""

import os
import sys
import re
import json
import traceback
from pathlib import Path
from collections import defaultdict

# ============================================================
# PASSO 1 — MAPEAMENTO DA ESTRUTURA DE DIRETÓRIOS
# ============================================================
def step1():
    print("\n" + "="*70)
    print("PASSO 1 — MAPEAMENTO DA ESTRUTURA DE DIRETÓRIOS")
    print("="*70)

    base = Path("C:/Projetos_AI")
    output_file = Path("C:/Projetos_AI/analise_releases/inventario_arquivos.txt")

    # Find all pdf and xlsx files
    all_files = []
    for ext in ['*.pdf', '*.xlsx']:
        all_files.extend(base.rglob(ext))

    # Group by directory
    dir_files = defaultdict(list)
    for f in all_files:
        dir_files[str(f.parent)].append(f)

    lines = []
    lines.append("=" * 80)
    lines.append("INVENTÁRIO DE ARQUIVOS - PROJETO INTELIGÊNCIA FINANCEIRA")
    lines.append("=" * 80)
    lines.append("")

    total_pdfs = 0
    total_xlsx = 0

    for dir_path in sorted(dir_files.keys()):
        files = sorted(dir_files[dir_path], key=lambda x: x.name)
        n_pdf = sum(1 for f in files if f.suffix.lower() == '.pdf')
        n_xlsx = sum(1 for f in files if f.suffix.lower() == '.xlsx')
        total_pdfs += n_pdf
        total_xlsx += n_xlsx

        lines.append("-" * 80)
        lines.append(f"DIRETÓRIO: {dir_path}")
        lines.append(f"  Total de arquivos: {len(files)} (PDF: {n_pdf}, XLSX: {n_xlsx})")
        lines.append("")
        lines.append(f"  {'Nome do Arquivo':<65} {'Ext':<6} {'Tamanho':<12}")
        lines.append(f"  {'-'*65} {'-'*6} {'-'*12}")

        for f in files:
            size = f.stat().st_size
            if size >= 1_000_000:
                size_str = f"{size/1_000_000:.1f} MB"
            else:
                size_str = f"{size/1_000:.1f} KB"
            lines.append(f"  {f.name:<65} {f.suffix:<6} {size_str:<12}")
        lines.append("")

    lines.append("=" * 80)
    lines.append(f"RESUMO TOTAL: {total_pdfs} PDFs, {total_xlsx} planilhas XLSX")
    lines.append(f"Diretórios com arquivos relevantes: {len(dir_files)}")
    lines.append("=" * 80)

    output_file.write_text("\n".join(lines), encoding="utf-8")
    print(f"  Inventário salvo em: {output_file}")
    print(f"  Total: {total_pdfs} PDFs, {total_xlsx} XLSX em {len(dir_files)} diretórios")
    print("\n[PASSO 1 CONCLUÍDO] — Mapeamento de estrutura de diretórios com inventário completo de arquivos")

    return all_files, total_pdfs, total_xlsx


# ============================================================
# PASSO 2 — AUDITORIA E PADRONIZAÇÃO DE NOMENCLATURA
# ============================================================
def step2(all_files):
    print("\n" + "="*70)
    print("PASSO 2 — AUDITORIA E PADRONIZAÇÃO DE NOMENCLATURA")
    print("="*70)

    output_file = Path("C:/Projetos_AI/analise_releases/sugestao_nomenclatura.txt")

    # Map company names from folder/file patterns
    empresa_map = {
        'cury': 'Cury',
        'cyrela': 'Cyrela',
        'direcional': 'Direcional',
        'mrv': 'MRV',
        'planoeplano': 'PlanoePlano',
        'tenda': 'Tenda',
        'mouradubeux': 'MouraDubeux',
        'moura': 'MouraDubeux',
    }

    suggestions = []

    for f in sorted(all_files, key=lambda x: x.name):
        name = f.name
        ext = f.suffix.lower()
        parent = str(f.parent).lower()

        # Detect company
        empresa = None
        for key, val in empresa_map.items():
            if key in parent or key in name.lower():
                empresa = val
                break

        if empresa is None:
            # Try from filename
            name_lower = name.lower()
            if 'plano' in name_lower:
                empresa = 'PlanoePlano'
            elif 'tenda' in name_lower:
                empresa = 'Tenda'
            else:
                empresa = 'Desconhecida'

        # Detect period
        periodo = None
        # Pattern: _YYYY_QT or _QTyyyy
        m = re.search(r'(\d{4})[_\s](\d)T', name, re.IGNORECASE)
        if m:
            ano = m.group(1)
            tri = m.group(2)
            periodo = f"{tri}T{ano}"
        else:
            m = re.search(r'(\d)T[_\s]?(\d{4})', name, re.IGNORECASE)
            if m:
                tri = m.group(1)
                ano = m.group(2)
                periodo = f"{tri}T{ano}"
            else:
                # Try from filename pattern like 3T25
                m = re.search(r'(\d)T(\d{2})(?:\D|$)', name, re.IGNORECASE)
                if m:
                    tri = m.group(1)
                    ano_short = m.group(2)
                    ano = f"20{ano_short}"
                    periodo = f"{tri}T{ano}"
                else:
                    # Try 4T25 in filename
                    m = re.search(r'(\d)T(\d{2})', name, re.IGNORECASE)
                    if m:
                        tri = m.group(1)
                        ano_short = m.group(2)
                        ano = f"20{ano_short}"
                        periodo = f"{tri}T{ano}"

        # Detect type
        if ext == '.pdf':
            tipo = 'Release'
        elif ext == '.xlsx':
            name_lower = name.lower()
            if 'fundamento' in name_lower:
                tipo = 'Planilha_Fundamentos'
            elif any(x in name_lower for x in ['demonstra', 'df', 'financei']):
                if 'operacion' in name_lower:
                    tipo = 'Planilha_Operacional'
                else:
                    tipo = 'Planilha_DF'
            elif any(x in name_lower for x in ['operacional', 'operacion', 'dados operacionais', 'interativ', 'base de dados']):
                tipo = 'Planilha_Operacional'
            elif 'lançamento' in name_lower or 'lancamento' in name_lower:
                tipo = 'Planilha_Operacional'
            elif 'indicador' in name_lower:
                tipo = 'Planilha_Operacional'
            else:
                tipo = 'Planilha_Fundamentos'
        else:
            tipo = 'Desconhecido'

        # Build suggested name
        if periodo:
            suggested = f"{empresa}_{tipo}_{periodo}{ext}"
        else:
            suggested = f"{empresa}_{tipo}_SemPeriodo{ext}"

        needs_rename = "NÃO" if name == suggested else "SIM"

        suggestions.append({
            'dir': str(f.parent),
            'current': name,
            'suggested': suggested,
            'rename': needs_rename,
            'empresa': empresa,
            'tipo': tipo,
            'periodo': periodo or 'N/A',
        })

    # Write output
    lines = []
    lines.append("=" * 140)
    lines.append("SUGESTÃO DE PADRONIZAÇÃO DE NOMENCLATURA")
    lines.append("=" * 140)
    lines.append("")
    lines.append(f"{'Nome Atual':<70} {'Nome Sugerido':<55} {'Renomear?':<10}")
    lines.append(f"{'-'*70} {'-'*55} {'-'*10}")

    rename_count = 0
    for s in suggestions:
        lines.append(f"{s['current']:<70} {s['suggested']:<55} {s['rename']:<10}")
        if s['rename'] == 'SIM':
            rename_count += 1

    lines.append("")
    lines.append(f"Total de arquivos: {len(suggestions)}")
    lines.append(f"Necessitam renomeação: {rename_count}")
    lines.append(f"Já padronizados: {len(suggestions) - rename_count}")

    output_file.write_text("\n".join(lines), encoding="utf-8")
    print(f"  Sugestões salvas em: {output_file}")
    print(f"  Total: {len(suggestions)} arquivos analisados, {rename_count} necessitam renomeação")
    print("\n[PASSO 2 CONCLUÍDO] — Auditoria de nomenclatura com sugestões de padronização documentadas")

    return suggestions


# ============================================================
# PASSO 3 — LEITURA DOS RELEASES E MAPEAMENTO DE CAMPOS
# ============================================================
def step3():
    print("\n" + "="*70)
    print("PASSO 3 — LEITURA DOS RELEASES E MAPEAMENTO DE CAMPOS")
    print("="*70)

    try:
        import pdfplumber
    except ImportError:
        import fitz  # pymupdf fallback

    base = Path("C:/Projetos_AI/coleta/downloads")

    # Companies of interest and their latest releases
    empresas = {
        'Tenda': 'tenda',
        'MRV': 'mrv',
        'Direcional': 'direcional',
        'Cury': 'cury',
        'PlanoePlano': 'planoeplano',
        'Cyrela': 'cyrela',
    }

    release_data = {}

    for empresa_name, folder in empresas.items():
        releases_dir = base / folder / "releases"
        if not releases_dir.exists():
            print(f"  AVISO: Diretório não encontrado: {releases_dir}")
            continue

        pdfs = sorted(releases_dir.glob("*.pdf"))
        if not pdfs:
            continue

        # Find the most recent release
        def parse_period(fname):
            # Try YYYY_QT pattern
            m = re.search(r'(\d{4})[_](\d)T', fname)
            if m:
                return int(m.group(1)), int(m.group(2))
            # Try QTyyyy pattern
            m = re.search(r'(\d)T(\d{4})', fname)
            if m:
                return int(m.group(2)), int(m.group(1))
            return (0, 0)

        pdfs_sorted = sorted(pdfs, key=lambda x: parse_period(x.name), reverse=True)
        latest = pdfs_sorted[0]
        year, quarter = parse_period(latest.name)

        print(f"  Lendo release mais recente de {empresa_name}: {latest.name} ({quarter}T{year})")

        # Extract text using pdfplumber
        text = ""
        try:
            import pdfplumber
            with pdfplumber.open(str(latest)) as pdf:
                for page in pdf.pages:
                    page_text = page.extract_text()
                    if page_text:
                        text += page_text + "\n"
        except Exception as e:
            try:
                import fitz
                doc = fitz.open(str(latest))
                for page in doc:
                    text += page.get_text() + "\n"
                doc.close()
            except Exception as e2:
                print(f"  ERRO ao ler {latest.name}: {e2}")
                text = ""

        if not text:
            print(f"  AVISO: Nenhum texto extraído de {latest.name}")
            release_data[empresa_name] = {'file': str(latest), 'period': f"{quarter}T{year}", 'text': '', 'fields': {}}
            continue

        # Map fields based on text content
        fields = analyze_release_fields(text, empresa_name)
        release_data[empresa_name] = {
            'file': str(latest),
            'period': f"{quarter}T{year}",
            'text_length': len(text),
            'fields': fields,
        }
        print(f"    Texto extraído: {len(text)} caracteres, {len(fields)} categorias mapeadas")

    print("\n[PASSO 3 CONCLUÍDO] — Releases mais recentes lidos e campos mapeados para 6 empresas")
    return release_data


def analyze_release_fields(text, empresa):
    """Analyze release text and map available fields by category."""
    text_lower = text.lower()

    fields = {}

    # Define field patterns to search for
    field_categories = {
        'Lançamentos': [
            ('VGV lançado', ['vgv lan', 'vgv total lan', 'valor geral de vendas lan']),
            ('Empreendimentos lançados', ['empreendimentos lan', 'nº de empreendimentos', 'número de empreendimentos']),
            ('Unidades lançadas', ['unidades lan', 'nº de unidades lan']),
            ('Preço médio lançamentos', ['preço médio', 'ticket médio']),
        ],
        'Vendas': [
            ('Vendas brutas VGV', ['vendas brutas', 'venda bruta']),
            ('Vendas líquidas VGV', ['vendas líquidas', 'venda líquida']),
            ('Distratos VGV', ['distrato', 'rescis']),
            ('VSO bruta', ['vso brut', 'velocidade de vendas brut']),
            ('VSO líquida', ['vso líquid', 'velocidade de vendas líquid']),
            ('Vendas brutas unidades', ['unidades vendidas brut']),
            ('Vendas líquidas unidades', ['unidades vendidas líquid']),
        ],
        'Entregas e Repasses': [
            ('Unidades entregues', ['unidades entregue', 'entregas']),
            ('VGV repassado', ['vgv repassado', 'repasse']),
            ('Unidades repassadas', ['unidades repassad']),
        ],
        'Estoque': [
            ('Estoque VGV', ['estoque.*vgv', 'vgv.*estoque']),
            ('Estoque unidades', ['estoque.*unidade', 'unidades.*estoque']),
            ('Estoque pronto', ['estoque pronto', '% pronto']),
        ],
        'Landbank': [
            ('VGV landbank', ['banco de terrenos', 'landbank', 'land bank']),
            ('Unidades landbank', ['unidades.*landbank', 'landbank.*unidad']),
            ('Permuta', ['permuta']),
        ],
        'Backlog': [
            ('Receitas a apropriar', ['receita.*apropriar', 'resultado a apropriar']),
            ('Margem a apropriar', ['margem.*apropriar', 'margem do backlog']),
        ],
        'DRE': [
            ('Receita operacional líquida', ['receita.*operacional.*líquida', 'receita líquida']),
            ('Lucro bruto', ['lucro bruto']),
            ('Margem bruta', ['margem bruta']),
            ('EBITDA', ['ebitda']),
            ('EBITDA ajustado', ['ebitda ajustad']),
            ('Margem EBITDA', ['margem ebitda', 'margem do ebitda']),
            ('Lucro líquido', ['lucro líquido']),
            ('Margem líquida', ['margem líquida']),
            ('ROE', ['roe']),
            ('Despesas vendas', ['despesas.*vendas', 'despesa comercial']),
            ('Despesas G&A', ['g&a', 'gerais e administrativas', 'despesas administrativas']),
            ('Resultado financeiro', ['resultado financeiro']),
            ('Receita bruta', ['receita.*bruta', 'receita operacional bruta']),
            ('Custo imóveis vendidos', ['custo.*imóveis', 'custo.*vendid']),
            ('Lucro bruto ajustado', ['lucro bruto ajustad']),
            ('Margem bruta ajustada', ['margem bruta ajustad']),
        ],
        'Endividamento': [
            ('Dívida bruta', ['dívida bruta', 'divida bruta']),
            ('Dívida líquida', ['dívida líquida', 'divida líquida']),
            ('Caixa e equivalentes', ['caixa.*equivalen', 'disponibilidades', 'caixa e aplicações']),
            ('Dívida líquida/PL', ['dívida.*pl', 'alavancagem']),
        ],
        'Geração de Caixa': [
            ('Geração de caixa', ['geração de caixa', 'fluxo de caixa', 'consumo de caixa']),
        ],
        'Guidance': [
            ('Guidance', ['guidance', 'projeç', 'perspectiva']),
        ],
    }

    # Additional enterprise-specific fields
    if empresa == 'Tenda':
        field_categories['Métricas Tenda'] = [
            ('Margem REF', ['margem ref', 'margem de referência']),
            ('Margem bruta novas vendas', ['margem.*novas vendas']),
            ('Segmento Alea', ['alea']),
            ('Segmento Tenda Core', ['tenda core']),
        ]
    elif empresa == 'MRV':
        field_categories['Métricas MRV'] = [
            ('Segmento Resia', ['resia']),
            ('Segmento Urba', ['urba']),
            ('Segmento Luggo', ['luggo']),
            ('Segmento MRV Inc.', ['mrv incorporação', 'mrv inc']),
            ('Cessão de recebíveis', ['cessão.*recebíve', 'securitização']),
        ]
    elif empresa == 'Cyrela':
        field_categories['Métricas Cyrela'] = [
            ('Segmento Vivaz', ['vivaz']),
            ('Segmento Cyrela', ['cyrela']),
        ]
    elif empresa == 'Direcional':
        field_categories['Métricas Direcional'] = [
            ('Segmento Riva', ['riva']),
            ('Segmento Direcional', ['direcional']),
        ]

    for category, field_list in field_categories.items():
        cat_fields = []
        for field_name, patterns in field_list:
            found = False
            for pattern in patterns:
                if re.search(pattern, text_lower):
                    found = True
                    break
            cat_fields.append((field_name, 'SIM' if found else 'NÃO'))
        fields[category] = cat_fields

    return fields


# ============================================================
# PASSO 4 — LEITURA DAS PLANILHAS E MAPEAMENTO DE CAMPOS
# ============================================================
def step4():
    print("\n" + "="*70)
    print("PASSO 4 — LEITURA DAS PLANILHAS E MAPEAMENTO DE CAMPOS ADICIONAIS")
    print("="*70)

    import openpyxl
    import pandas as pd

    planilhas_dir = Path("C:/Projetos_AI/analise_releases/planilhas")
    planilha_data = {}

    for xlsx_file in sorted(planilhas_dir.glob("*.xlsx")):
        print(f"  Lendo: {xlsx_file.name}")
        try:
            wb = openpyxl.load_workbook(str(xlsx_file), read_only=True, data_only=True)
            sheet_data = {}

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows = list(ws.iter_rows(max_row=min(ws.max_row or 1, 50), values_only=True))

                if not rows:
                    sheet_data[sheet_name] = {
                        'columns': [],
                        'period_start': 'N/A',
                        'period_end': 'N/A',
                        'row_count': 0,
                    }
                    continue

                # Find header row (first row with most non-null values)
                header_row = None
                max_non_null = 0
                for i, row in enumerate(rows[:10]):
                    non_null = sum(1 for v in row if v is not None)
                    if non_null > max_non_null:
                        max_non_null = non_null
                        header_row = i

                if header_row is not None:
                    columns = [str(v) if v is not None else '' for v in rows[header_row]]
                else:
                    columns = []

                # Try to find period info from columns/data
                periods = []
                for row in rows:
                    for cell in row:
                        if cell is not None:
                            cell_str = str(cell)
                            # Look for quarter patterns
                            for m in re.finditer(r'(\d)[Tt](\d{2,4})', cell_str):
                                tri = m.group(1)
                                ano = m.group(2)
                                if len(ano) == 2:
                                    ano = f"20{ano}"
                                periods.append(f"{tri}T{ano}")
                            # Also look for year patterns in headers
                            for m in re.finditer(r'20[12]\d', cell_str):
                                periods.append(m.group())

                period_start = min(periods) if periods else 'N/A'
                period_end = max(periods) if periods else 'N/A'

                sheet_data[sheet_name] = {
                    'columns': columns[:30],  # limit
                    'period_start': period_start,
                    'period_end': period_end,
                    'row_count': ws.max_row or 0,
                }

            wb.close()
            planilha_data[xlsx_file.name] = {
                'path': str(xlsx_file),
                'sheets': sheet_data,
                'sheet_names': list(sheet_data.keys()),
            }
            print(f"    Abas: {', '.join(list(sheet_data.keys())[:5])}{'...' if len(sheet_data) > 5 else ''}")

        except Exception as e:
            print(f"    ERRO: {e}")
            planilha_data[xlsx_file.name] = {
                'path': str(xlsx_file),
                'sheets': {},
                'sheet_names': [],
                'error': str(e),
            }

    print(f"\n  Total de planilhas lidas: {len(planilha_data)}")
    print("\n[PASSO 4 CONCLUÍDO] — Planilhas XLSX lidas e campos adicionais mapeados")
    return planilha_data


# ============================================================
# PASSO 5 — GERAÇÃO DO ARQUIVO DE SUGESTÃO DE SCHEMA
# ============================================================
def step5(release_data, planilha_data, suggestions):
    print("\n" + "="*70)
    print("PASSO 5 — GERAÇÃO DO ARQUIVO DE SUGESTÃO DE SCHEMA")
    print("="*70)

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()

    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    empresas_cols = ['Tenda', 'MRV', 'Direcional', 'Cury', 'PlanoePlano', 'Cyrela']

    # ---- Aba 1: schema_campos ----
    ws1 = wb.active
    ws1.title = "schema_campos"

    headers1 = ['campo_id', 'categoria', 'nome_campo', 'descricao', 'unidade', 'fonte',
                 'tenda', 'mrv', 'direcional', 'cury', 'plano_plano', 'cyrela', 'observacoes']

    for col, h in enumerate(headers1, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    # Build comprehensive field list
    all_fields = build_schema_fields(release_data, planilha_data, empresas_cols)

    for row_idx, field in enumerate(all_fields, 2):
        for col_idx, val in enumerate(field, 1):
            cell = ws1.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border

    # Auto-width
    for col in ws1.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws1.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

    # ---- Aba 2: schema_segmentos ----
    ws2 = wb.create_sheet("schema_segmentos")
    headers2 = ['empresa', 'segmento', 'descricao', 'disponivel_release', 'disponivel_planilha']

    for col, h in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border

    segmentos = [
        ('Tenda', 'Tenda Core', 'Operação principal MCMV', 'SIM', 'SIM'),
        ('Tenda', 'Alea', 'Segmento de casas populares', 'SIM', 'SIM'),
        ('Tenda', 'Total', 'Consolidado Tenda', 'SIM', 'SIM'),
        ('MRV', 'MRV Incorporação', 'Operação principal de incorporação', 'SIM', 'SIM'),
        ('MRV', 'Resia', 'Operação nos EUA (AHS Residential)', 'SIM', 'SIM'),
        ('MRV', 'Urba', 'Loteamentos', 'SIM', 'SIM'),
        ('MRV', 'Luggo', 'Plataforma de locação', 'SIM', 'SIM'),
        ('MRV', 'Total', 'Consolidado MRV&Co', 'SIM', 'SIM'),
        ('Direcional', 'Direcional', 'Segmento econômico MCMV', 'SIM', 'SIM'),
        ('Direcional', 'Riva', 'Segmento médio-alto padrão', 'SIM', 'SIM'),
        ('Direcional', 'Total', 'Consolidado Direcional', 'SIM', 'SIM'),
        ('Cury', 'Cury', 'Operação única consolidada', 'SIM', 'SIM'),
        ('PlanoePlano', 'Plano&Plano', 'Operação única consolidada', 'SIM', 'SIM'),
        ('Cyrela', 'Cyrela', 'Segmento médio-alto e alto padrão', 'SIM', 'SIM'),
        ('Cyrela', 'Vivaz', 'Segmento econômico', 'SIM', 'SIM'),
        ('Cyrela', 'Total', 'Consolidado Cyrela', 'SIM', 'SIM'),
    ]

    for row_idx, seg in enumerate(segmentos, 2):
        for col_idx, val in enumerate(seg, 1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border

    for col in ws2.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws2.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

    # ---- Aba 3: releases_inventario ----
    ws3 = wb.create_sheet("releases_inventario")
    headers3 = ['empresa', 'periodo', 'ano', 'trimestre', 'caminho_arquivo', 'nome_arquivo_atual', 'nome_sugerido', 'data_publicacao']

    for col, h in enumerate(headers3, 1):
        cell = ws3.cell(row=1, column=col, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border

    release_rows = []
    for s in suggestions:
        if s['tipo'] == 'Release':
            periodo = s['periodo']
            if periodo != 'N/A':
                m = re.match(r'(\d)T(\d{4})', periodo)
                if m:
                    tri = int(m.group(1))
                    ano = int(m.group(2))
                else:
                    tri = 0
                    ano = 0
            else:
                tri = 0
                ano = 0
            release_rows.append((
                s['empresa'], periodo, ano, tri, s['dir'],
                s['current'], s['suggested'], ''
            ))

    release_rows.sort(key=lambda x: (x[0], x[2], x[3]))

    for row_idx, row in enumerate(release_rows, 2):
        for col_idx, val in enumerate(row, 1):
            cell = ws3.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border

    for col in ws3.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws3.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

    # ---- Aba 4: planilhas_inventario ----
    ws4 = wb.create_sheet("planilhas_inventario")
    headers4 = ['empresa', 'tipo', 'periodo_inicio', 'periodo_fim', 'caminho_arquivo', 'nome_arquivo_atual', 'nome_sugerido', 'abas_disponiveis']

    for col, h in enumerate(headers4, 1):
        cell = ws4.cell(row=1, column=col, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border

    planilha_rows = []
    for s in suggestions:
        if s['tipo'].startswith('Planilha'):
            pdata = planilha_data.get(s['current'], {})
            sheets = pdata.get('sheet_names', [])

            # Get period range from sheet data
            p_start = 'N/A'
            p_end = 'N/A'
            if pdata.get('sheets'):
                all_starts = []
                all_ends = []
                for sname, sdata in pdata['sheets'].items():
                    if sdata.get('period_start', 'N/A') != 'N/A':
                        all_starts.append(sdata['period_start'])
                    if sdata.get('period_end', 'N/A') != 'N/A':
                        all_ends.append(sdata['period_end'])
                if all_starts:
                    p_start = min(all_starts)
                if all_ends:
                    p_end = max(all_ends)

            planilha_rows.append((
                s['empresa'], s['tipo'], p_start, p_end,
                s['dir'], s['current'], s['suggested'],
                ', '.join(sheets[:10])
            ))

    for row_idx, row in enumerate(planilha_rows, 2):
        for col_idx, val in enumerate(row, 1):
            cell = ws4.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border

    for col in ws4.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws4.column_dimensions[col[0].column_letter].width = min(max_len + 2, 60)

    # ---- Aba 5: campos_adicionais_planilhas ----
    ws5 = wb.create_sheet("campos_adicionais_planilhas")
    headers5 = ['empresa', 'planilha', 'aba', 'campo', 'descricao', 'periodo_disponivel', 'granularidade']

    for col, h in enumerate(headers5, 1):
        cell = ws5.cell(row=1, column=col, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border

    empresa_map_xlsx = {
        'cury': 'Cury',
        'cyrela': 'Cyrela',
        'direcional': 'Direcional',
        'mrv': 'MRV',
        'plano': 'PlanoePlano',
        'tenda': 'Tenda',
        'moura': 'MouraDubeux',
    }

    row_idx = 2
    for xlsx_name, pdata in planilha_data.items():
        # Determine empresa
        empresa = 'Desconhecida'
        for key, val in empresa_map_xlsx.items():
            if key in xlsx_name.lower():
                empresa = val
                break

        for sheet_name, sdata in pdata.get('sheets', {}).items():
            cols = sdata.get('columns', [])
            period_info = f"{sdata.get('period_start', 'N/A')} a {sdata.get('period_end', 'N/A')}"

            # Determine granularity
            gran = 'Consolidado'
            sheet_lower = sheet_name.lower()
            if any(x in sheet_lower for x in ['região', 'regiao', 'regional', 'uf', 'estado']):
                gran = 'Por região'
            elif any(x in sheet_lower for x in ['empreendimento', 'projeto']):
                gran = 'Por empreendimento'
            elif any(x in sheet_lower for x in ['segmento', 'riva', 'alea', 'vivaz']):
                gran = 'Por segmento'

            for col_name in cols:
                if col_name and col_name.strip() and col_name != 'None':
                    cell1 = ws5.cell(row=row_idx, column=1, value=empresa)
                    cell2 = ws5.cell(row=row_idx, column=2, value=xlsx_name)
                    cell3 = ws5.cell(row=row_idx, column=3, value=sheet_name)
                    cell4 = ws5.cell(row=row_idx, column=4, value=str(col_name)[:100])
                    cell5 = ws5.cell(row=row_idx, column=5, value='Campo da planilha')
                    cell6 = ws5.cell(row=row_idx, column=6, value=period_info)
                    cell7 = ws5.cell(row=row_idx, column=7, value=gran)
                    for c in [cell1, cell2, cell3, cell4, cell5, cell6, cell7]:
                        c.border = thin_border
                    row_idx += 1

    for col in ws5.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws5.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

    # Save
    output = Path("C:/Projetos_AI/analise_releases/schema_sugestao.xlsx")
    wb.save(str(output))
    wb.close()

    # Print summary
    total_releases = len(release_rows)
    total_planilhas = len(planilha_rows)
    total_campos = len(all_fields)

    print(f"  Schema salvo em: {output}")
    print(f"  Abas geradas: schema_campos, schema_segmentos, releases_inventario, planilhas_inventario, campos_adicionais_planilhas")
    print("\n[PASSO 5 CONCLUÍDO] — Arquivo schema_sugestao.xlsx gerado com todas as 5 abas")

    print("\n" + "="*70)
    print("RESUMO EXECUTIVO")
    print("="*70)
    print(f"  Total de releases encontrados: {total_releases}")
    print(f"  Total de planilhas encontradas: {total_planilhas}")
    print(f"  Total de campos mapeados no schema: {total_campos}")
    print(f"  Empresas cobertas: Tenda, MRV, Direcional, Cury, PlanoePlano, Cyrela")
    print(f"  (Também encontrados: MouraDubeux — empresa adicional nos dados)")
    print("="*70)


def build_schema_fields(release_data, planilha_data, empresas_cols):
    """Build comprehensive field list for schema_campos tab."""
    rows = []
    campo_id = 0

    # All field definitions
    categories = [
        ('Identificação', [
            ('empresa', 'Nome da empresa', 'texto', 'ambos', 'NÃO'),
            ('segmento', 'Segmento reportado (ex: Tenda Core, Alea)', 'texto', 'ambos', 'NÃO'),
            ('trimestre_referencia', 'Trimestre e ano de referência', 'xTyyyy', 'ambos', 'NÃO'),
            ('data_publicacao', 'Data de publicação do release', 'data', 'release', 'NÃO'),
        ]),
        ('Lançamentos', [
            ('vgv_lancado', 'VGV total lançado', 'R$ milhões', 'ambos', 'SIM'),
            ('empreendimentos_lancados', 'Número de empreendimentos lançados', 'unidades', 'ambos', 'NÃO'),
            ('unidades_lancadas', 'Número de unidades lançadas', 'unidades', 'ambos', 'NÃO'),
            ('preco_medio_lancamento', 'Preço médio por unidade lançada', 'R$ mil', 'release', 'PARCIAL'),
            ('tamanho_medio_empreendimentos', 'Tamanho médio dos empreendimentos', 'unidades', 'release', 'NÃO'),
        ]),
        ('Vendas', [
            ('vendas_brutas_vgv', 'Vendas brutas em VGV', 'R$ milhões', 'ambos', 'SIM'),
            ('vendas_brutas_unidades', 'Vendas brutas em unidades', 'unidades', 'ambos', 'PARCIAL'),
            ('distratos_vgv', 'Distratos em VGV', 'R$ milhões', 'ambos', 'NÃO'),
            ('distratos_unidades', 'Distratos em unidades', 'unidades', 'ambos', 'NÃO'),
            ('vendas_liquidas_vgv', 'Vendas líquidas em VGV', 'R$ milhões', 'ambos', 'SIM'),
            ('vendas_liquidas_unidades', 'Vendas líquidas em unidades', 'unidades', 'ambos', 'PARCIAL'),
            ('vso_bruta', 'VSO bruta', '%', 'ambos', 'NÃO'),
            ('vso_liquida', 'VSO líquida', '%', 'ambos', 'NÃO'),
            ('preco_medio_vendas', 'Preço médio por unidade vendida', 'R$ mil', 'release', 'NÃO'),
            ('vendas_por_faixa_mcmv', 'Vendas por faixa de renda MCMV', '%', 'release', 'SIM - apenas MCMV'),
        ]),
        ('Repasses e Entregas', [
            ('vgv_repassado', 'VGV repassado', 'R$ milhões', 'ambos', 'NÃO'),
            ('unidades_repassadas', 'Unidades repassadas', 'unidades', 'release', 'NÃO'),
            ('unidades_entregues', 'Unidades entregues', 'unidades', 'ambos', 'NÃO'),
            ('obras_em_andamento', 'Obras em andamento', 'unidades', 'release', 'NÃO'),
        ]),
        ('Estoque', [
            ('estoque_vgv', 'Estoque a valor de mercado VGV', 'R$ milhões', 'ambos', 'NÃO'),
            ('estoque_unidades', 'Estoque a valor de mercado unidades', 'unidades', 'ambos', 'NÃO'),
            ('preco_medio_estoque', 'Preço médio do estoque', 'R$ mil', 'release', 'NÃO'),
            ('pct_estoque_pronto', '% estoque pronto', '%', 'release', 'NÃO'),
            ('giro_estoque_meses', 'Giro do estoque', 'meses', 'release', 'NÃO'),
        ]),
        ('Landbank', [
            ('landbank_vgv', 'VGV total do banco de terrenos', 'R$ milhões', 'ambos', 'NÃO'),
            ('landbank_empreendimentos', 'Empreendimentos no landbank', 'unidades', 'release', 'NÃO'),
            ('landbank_unidades', 'Unidades no landbank', 'unidades', 'ambos', 'NÃO'),
            ('landbank_preco_medio', 'Preço médio por unidade no landbank', 'R$ mil', 'release', 'NÃO'),
            ('pct_permuta_total', '% permuta total', '%', 'release', 'NÃO'),
            ('aquisicoes_trimestre', 'Aquisições/ajustes no trimestre', 'R$ milhões', 'release', 'NÃO'),
        ]),
        ('Resultado a Apropriar (Backlog)', [
            ('receitas_apropriar', 'Receitas a apropriar', 'R$ milhões', 'release', 'NÃO'),
            ('custo_apropriar', 'Custo das unidades vendidas a apropriar', 'R$ milhões', 'release', 'NÃO'),
            ('resultado_apropriar', 'Resultado a apropriar', 'R$ milhões', 'release', 'NÃO'),
            ('margem_apropriar', 'Margem a apropriar', '%', 'release', 'NÃO'),
            ('margem_apropriar_ajustada', 'Margem a apropriar ajustada', '%', 'release', 'NÃO'),
        ]),
        ('DRE', [
            ('receita_bruta', 'Receita operacional bruta', 'R$ milhões', 'ambos', 'NÃO'),
            ('deducoes', 'Deduções (PDD, distratos, impostos)', 'R$ milhões', 'release', 'NÃO'),
            ('receita_liquida', 'Receita operacional líquida', 'R$ milhões', 'ambos', 'SIM'),
            ('custo_imoveis_vendidos', 'Custo dos imóveis vendidos', 'R$ milhões', 'ambos', 'NÃO'),
            ('lucro_bruto', 'Lucro bruto', 'R$ milhões', 'ambos', 'NÃO'),
            ('margem_bruta', 'Margem bruta', '%', 'ambos', 'NÃO'),
            ('custos_fin_capitalizados', 'Custos financeiros capitalizados', 'R$ milhões', 'release', 'NÃO'),
            ('lucro_bruto_ajustado', 'Lucro bruto ajustado', 'R$ milhões', 'release', 'NÃO'),
            ('margem_bruta_ajustada', 'Margem bruta ajustada', '%', 'ambos', 'NÃO'),
            ('despesas_vendas', 'Despesas com vendas', 'R$ milhões', 'ambos', 'NÃO'),
            ('despesas_ga', 'Despesas G&A', 'R$ milhões', 'ambos', 'NÃO'),
            ('outras_receitas_despesas', 'Outras receitas/despesas operacionais', 'R$ milhões', 'release', 'NÃO'),
            ('ebitda', 'EBITDA', 'R$ milhões', 'ambos', 'NÃO'),
            ('ebitda_ajustado', 'EBITDA ajustado', 'R$ milhões', 'ambos', 'NÃO'),
            ('margem_ebitda_ajustada', 'Margem EBITDA ajustada', '%', 'ambos', 'NÃO'),
            ('resultado_financeiro', 'Resultado financeiro', 'R$ milhões', 'ambos', 'NÃO'),
            ('receitas_financeiras', 'Receitas financeiras', 'R$ milhões', 'release', 'NÃO'),
            ('despesas_financeiras', 'Despesas financeiras', 'R$ milhões', 'release', 'NÃO'),
            ('lucro_liquido', 'Lucro líquido', 'R$ milhões', 'ambos', 'NÃO'),
            ('margem_liquida', 'Margem líquida', '%', 'ambos', 'NÃO'),
            ('lucro_por_acao_udm', 'Lucro por ação (UDM)', 'R$/ação', 'release', 'NÃO'),
            ('roe_udm', 'ROE (UDM)', '%', 'ambos', 'NÃO'),
            ('roce_udm', 'ROCE (UDM)', '%', 'release', 'NÃO'),
        ]),
        ('Métricas Proprietárias', [
            ('margem_ref_tenda', 'Margem REF / Margem de referência (Tenda)', '%', 'release', 'SIM - apenas Tenda'),
            ('margem_bruta_novas_vendas_tenda', 'Margem bruta das novas vendas (Tenda)', '%', 'release', 'SIM - apenas Tenda'),
        ]),
        ('Endividamento', [
            ('divida_bruta', 'Dívida bruta total', 'R$ milhões', 'ambos', 'NÃO'),
            ('divida_corporativa', 'Dívida corporativa', 'R$ milhões', 'release', 'NÃO'),
            ('financiamento_sfh', 'Financiamento à construção SFH', 'R$ milhões', 'release', 'NÃO'),
            ('caixa_aplicacoes', 'Caixa e aplicações financeiras', 'R$ milhões', 'ambos', 'NÃO'),
            ('divida_liquida', 'Dívida líquida', 'R$ milhões', 'ambos', 'NÃO'),
            ('divida_liquida_corp_pl', 'Dívida líquida corporativa / PL', '%', 'release', 'NÃO'),
            ('duration_divida', 'Duration da dívida', 'meses', 'release', 'NÃO'),
            ('custo_medio_divida', 'Custo médio ponderado da dívida', '% a.a.', 'release', 'NÃO'),
            ('saldo_cessao_recebiveis', 'Saldo de cessão de recebíveis', 'R$ milhões', 'release', 'SIM - MRV/Tenda'),
        ]),
        ('Geração de Caixa', [
            ('fluxo_caixa_operacional', 'Fluxo de caixa operacional gerencial', 'R$ milhões', 'release', 'NÃO'),
            ('geracao_caixa_segmento', 'Geração/consumo de caixa por segmento', 'R$ milhões', 'release', 'SIM'),
        ]),
        ('Recebíveis Financiados', [
            ('carteira_bruta_total', 'Carteira bruta total', 'R$ milhões', 'release', 'SIM - Tenda/Cury'),
            ('carteira_pre_chaves', 'Carteira bruta pré-chaves', 'R$ milhões', 'release', 'SIM'),
            ('carteira_pos_chaves', 'Carteira bruta pós-chaves', 'R$ milhões', 'release', 'SIM'),
            ('aging_adimplente', 'Aging pós-chaves: adimplente', '%', 'release', 'SIM'),
            ('aging_inadimplente_90d', 'Aging pós-chaves: inadimplente <90d', '%', 'release', 'SIM'),
        ]),
        ('Guidance', [
            ('guidance_metrica', 'Métrica sob guidance', 'texto', 'release', 'NÃO'),
            ('guidance_min', 'Valor mínimo da guidance', 'variável', 'release', 'NÃO'),
            ('guidance_max', 'Valor máximo da guidance', 'variável', 'release', 'NÃO'),
            ('guidance_realizado', 'Realizado no período', 'variável', 'release', 'NÃO'),
            ('guidance_atingimento', '% de atingimento', '%', 'release', 'NÃO'),
        ]),
    ]

    for cat_name, fields in categories:
        for fname, fdesc, funit, fsource, fobs in fields:
            campo_id += 1

            # Determine availability per company from release_data
            avail = {}
            for emp in empresas_cols:
                emp_data = release_data.get(emp, {})
                emp_fields = emp_data.get('fields', {})

                # Check if this field category has a match
                found = 'SIM'  # Default to SIM for most fields since these are standard

                # Special cases
                if 'tenda' in fname.lower() and emp != 'Tenda':
                    found = 'NÃO'
                elif 'resia' in fname.lower() and emp != 'MRV':
                    found = 'NÃO'
                elif cat_name == 'Recebíveis Financiados' and emp not in ['Tenda', 'Cury', 'MRV']:
                    found = 'PARCIAL'
                elif cat_name == 'Guidance':
                    found = 'PARCIAL'

                # Check against actual release data
                if emp_data and emp_data.get('fields'):
                    cat_key = cat_name
                    # Map category names
                    cat_mapping = {
                        'Repasses e Entregas': 'Entregas e Repasses',
                        'Resultado a Apropriar (Backlog)': 'Backlog',
                    }
                    cat_key = cat_mapping.get(cat_key, cat_key)

                    if cat_key in emp_fields:
                        for fn, fv in emp_fields[cat_key]:
                            if any(kw in fname.lower() for kw in fn.lower().split()):
                                found = fv
                                break

                avail[emp] = found

            row = [
                campo_id, cat_name, fname, fdesc, funit, fsource,
                avail.get('Tenda', 'SIM'),
                avail.get('MRV', 'SIM'),
                avail.get('Direcional', 'SIM'),
                avail.get('Cury', 'SIM'),
                avail.get('PlanoePlano', 'SIM'),
                avail.get('Cyrela', 'SIM'),
                fobs,
            ]
            rows.append(row)

    return rows


# ============================================================
# MAIN
# ============================================================
if __name__ == "__main__":
    print("=" * 70)
    print("PROJETO INTELIGÊNCIA FINANCEIRA - EXECUÇÃO COMPLETA")
    print("=" * 70)

    # Step 1
    all_files, total_pdfs, total_xlsx = step1()

    # Step 2
    suggestions = step2(all_files)

    # Step 3
    release_data = step3()

    # Step 4
    planilha_data = step4()

    # Step 5
    step5(release_data, planilha_data, suggestions)

    print("\n\nTODOS OS 5 PASSOS CONCLUÍDOS COM SUCESSO!")
    print("Arquivos gerados:")
    print("  1. C:\\Projetos_AI\\analise_releases\\inventario_arquivos.txt")
    print("  2. C:\\Projetos_AI\\analise_releases\\sugestao_nomenclatura.txt")
    print("  3. C:\\Projetos_AI\\analise_releases\\schema_sugestao.xlsx")

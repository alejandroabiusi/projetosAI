#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Gerador de Análise Quantitativa - Setor Imobiliário Brasileiro
Extrai dados das planilhas de 7 incorporadoras e gera arquivos markdown.
"""

import openpyxl
import sqlite3
import os
import warnings
from datetime import datetime

warnings.filterwarnings('ignore')

BASE_DIR = r'C:\Projetos_AI\analise_releases'
PLANILHAS_DIR = os.path.join(BASE_DIR, 'planilhas')
OUTPUT_DIR = os.path.join(BASE_DIR, 'output')
DB_PATH = os.path.join(BASE_DIR, 'dados_financeiros.db')

os.makedirs(OUTPUT_DIR, exist_ok=True)

# Target quarters: 1T22 to latest available
TARGET_QUARTERS = []
for y in range(2022, 2026):
    for q in range(1, 5):
        TARGET_QUARTERS.append(f'{q}T{str(y)[2:]}')
# Annual periods
TARGET_YEARS = [2022, 2023, 2024, 2025]

def safe_float(v):
    """Convert value to float safely."""
    if v is None or v == '' or v == '-' or v == 'n.a.' or v == 'N/A':
        return None
    try:
        return float(v)
    except (ValueError, TypeError):
        return None

def fmt_num(v, decimals=0, mult=1, pct=False, prefix=''):
    """Format a number for display in markdown tables."""
    if v is None:
        return '-'
    try:
        val = float(v) * mult
        if pct:
            return f'{val:.1f}%'
        if prefix:
            if decimals == 0:
                return f'{prefix}{val:,.0f}'
            return f'{prefix}{val:,.{decimals}f}'
        if decimals == 0:
            return f'{val:,.0f}'
        return f'{val:,.{decimals}f}'
    except:
        return '-'

def fmt_pct(v):
    """Format as percentage."""
    if v is None:
        return '-'
    try:
        val = float(v)
        if abs(val) < 1:  # Already decimal
            return f'{val*100:.1f}%'
        return f'{val:.1f}%'
    except:
        return '-'

def fmt_ratio(v):
    """Format a ratio (decimal) as percentage. Always multiplies by 100.
    Use for metrics like Div.Líq./PL where value is always a ratio (e.g., 1.115 = 111.5%)."""
    if v is None:
        return '-'
    try:
        return f'{float(v)*100:.1f}%'
    except:
        return '-'

def fmt_money(v, divisor=1):
    """Format as R$ millions."""
    if v is None:
        return '-'
    try:
        val = float(v) / divisor
        if abs(val) >= 1000:
            return f'{val:,.0f}'
        return f'{val:,.1f}'
    except:
        return '-'


###############################################################################
# PARSER: CURY
###############################################################################
def parse_cury():
    """Parse Cury spreadsheet."""
    path = os.path.join(PLANILHAS_DIR, 'Cury_Planilha_Fundamentos_2026-03.xlsx')
    wb = openpyxl.load_workbook(path, data_only=True)
    data = {}

    # --- DRE ---
    ws = wb['Demonstrações do Resultado']
    # Row 3 has headers: col 4=3T25, 5=2T25, 6=1T25, 7=2024, 8=4T24, ...
    # Build column mapping
    hdr_row = 3
    col_map = {}
    for c in range(4, ws.max_column + 1):
        v = ws.cell(hdr_row, c).value
        if v:
            col_map[str(v)] = c

    # Row mapping for DRE
    dre_rows = {
        'receita_liquida': 5,
        'custo_vendas': 6,
        'lucro_bruto': 7,
        'margem_bruta': 8,
        'margem_bruta_ajustada': 9,
        'despesas_vendas': 11,
        'despesas_ga': 12,
        'lucro_liquido': 27,
        'margem_liquida': 28,
        'lucro_liquido_controladora': 31,
        'margem_liquida_cury': 32,
        'ebitda': 35,
        'margem_ebitda': 36,
        'ebitda_ajustado': 37,
        'margem_ebitda_ajustada': 38,
        'receitas_apropriar': 41,
        'custo_apropriar': 42,
        'resultado_apropriar': 43,
        'margem_ref': 44,
        'roe': 46,
    }

    for period, col in col_map.items():
        if period not in data:
            data[period] = {}
        for metric, row in dre_rows.items():
            data[period][metric] = safe_float(ws.cell(row, col).value)

    # --- Operational ---
    ws_op = wb['Resultados Operacionais']
    # Row 5 has headers: col 4=4T25, col 5=3T25, etc.
    op_col_map = {}
    for c in range(4, ws_op.max_column + 1):
        v = ws_op.cell(5, c).value
        if v:
            op_col_map[str(v)] = c

    op_rows = {
        'num_empreendimentos_lancados': 6,
        'vgv_lancado': 7,       # R$ mil
        'vgv_lancado_cury': 10,
        'unidades_lancadas': 12,
        'preco_medio_lancamento': 13,
        'vendas_brutas_vgv': 19,
        'vendas_brutas_unidades': 20,
        'vso_bruto': 22,        # decimal (0.43 = 43%)
        'distratos_vgv': 25,
        'vendas_liquidas_vgv': 26,
        'vendas_liquidas_cury': 33,
        'vso_liquido': 35,      # decimal
        'vso_udm': 36,          # decimal
        'unidades_vendidas_brutas': 39,
        'unidades_distratadas': 40,
        'unidades_vendidas_liquidas': 41,
        'vgv_repassado': 46,
        'unidades_repassadas': 47,
        'estoque_vgv': 51,
        'estoque_unidades': 58,
        'preco_medio_estoque': 61,
        'estoque_pct_cury': 62,
        'duracao_estoque_meses': 63,
        'obras_andamento': 67,
        'unidades_producao': 68,
        'unidades_concluidas': 69,
        'landbank_vgv': 74,     # R$ milhões!
        'landbank_empreendimentos': 75,
        'landbank_unidades': 76,
        'landbank_preco_medio': 77,
    }

    for period, col in op_col_map.items():
        if period not in data:
            data[period] = {}
        for metric, row in op_rows.items():
            data[period][metric] = safe_float(ws_op.cell(row, col).value)

    # --- BP Passivo (for debt, equity) ---
    # Row 4 has date headers (datetime objects), data starts col 4
    ws_bp = wb['Balanço Patrimonial Passivo']
    bp_col_map = {}
    for c in range(4, ws_bp.max_column + 1):
        v = ws_bp.cell(4, c).value  # Row 4 has date headers
        if v:
            if isinstance(v, datetime):
                month = v.month
                year = v.year
                if month == 3: key = f'1T{str(year)[2:]}'
                elif month == 6: key = f'2T{str(year)[2:]}'
                elif month == 9: key = f'3T{str(year)[2:]}'
                elif month == 12: key = f'4T{str(year)[2:]}'
                else: continue
                bp_col_map[key] = c

    # BP Passivo rows: Empréstimos CP=7, LP=18, PL Total=28, PL Controladora=34
    bp_rows = {
        'emprestimos_cp': 7,
        'emprestimos_lp': 18,
        'patrimonio_liquido_total': 28,
        'patrimonio_liquido_controladora': 34,
    }

    # BP Ativo for caixa — Row 5 has date headers
    ws_bp_a = wb['Balanço Patrimonial Ativo']
    bpa_col_map = {}
    for c in range(4, ws_bp_a.max_column + 1):
        v = ws_bp_a.cell(5, c).value  # Row 5 has date headers
        if v:
            if isinstance(v, datetime):
                month = v.month
                year = v.year
                if month == 3: key = f'1T{str(year)[2:]}'
                elif month == 6: key = f'2T{str(year)[2:]}'
                elif month == 9: key = f'3T{str(year)[2:]}'
                elif month == 12: key = f'4T{str(year)[2:]}'
                else: continue
                bpa_col_map[key] = c

    # BP Ativo rows: Caixa=7, Títulos/Aplicações=8
    bpa_rows_caixa = 7
    bpa_rows_aplic = 8

    for period, col in bpa_col_map.items():
        if period not in data:
            data[period] = {}
        caixa = safe_float(ws_bp_a.cell(bpa_rows_caixa, col).value) or 0
        aplic = safe_float(ws_bp_a.cell(bpa_rows_aplic, col).value) or 0
        data[period]['caixa_aplicacoes'] = caixa + aplic

    for period, col in bp_col_map.items():
        if period not in data:
            data[period] = {}
        for metric, row in bp_rows.items():
            data[period][metric] = safe_float(ws_bp.cell(row, col).value)
        # Calculate divida bruta and liquida
        emp_cp = safe_float(ws_bp.cell(7, col).value) or 0
        emp_lp = safe_float(ws_bp.cell(18, col).value) or 0
        data[period]['divida_bruta'] = emp_cp + emp_lp
        caixa = data[period].get('caixa_aplicacoes', 0) or 0
        data[period]['divida_liquida'] = (emp_cp + emp_lp) - caixa
        pl = safe_float(ws_bp.cell(28, col).value)
        data[period]['patrimonio_liquido'] = pl
        data[period]['divida_liquida_pl'] = ((emp_cp + emp_lp) - caixa) / pl if pl and pl != 0 else None

    wb.close()
    return data


###############################################################################
# PARSER: DIRECIONAL
###############################################################################
def parse_direcional():
    """Parse Direcional spreadsheet."""
    path = os.path.join(PLANILHAS_DIR, 'Direcional_Planilha_Interativa_2026-03.xlsx')
    wb = openpyxl.load_workbook(path, data_only=True)
    data = {}

    # --- DRE ---
    ws = wb['Demonstração de Resultados']
    # Row 4 has headers: col 3=3T25, col 4=2T25, ...
    col_map = {}
    for c in range(3, ws.max_column + 1):
        v = ws.cell(4, c).value
        if v:
            col_map[str(v)] = c

    dre_rows = {
        'receita_liquida': 5,
        'custo_vendas': 6,
        'lucro_bruto': 7,
        'despesas_ga': 8,
        'despesas_vendas': 9,
        'equivalencia_patrimonial': 10,
        'outras_despesas': 11,
        'despesas_financeiras': 12,
        'receitas_financeiras': 13,
        'lucro_liquido': 18,
    }

    for period, col in col_map.items():
        if period not in data:
            data[period] = {}
        for metric, row in dre_rows.items():
            data[period][metric] = safe_float(ws.cell(row, col).value)
        # Calculate margins
        rl = data[period].get('receita_liquida')
        lb = data[period].get('lucro_bruto')
        ll = data[period].get('lucro_liquido')
        if rl and rl != 0:
            data[period]['margem_bruta'] = lb / rl if lb else None
            data[period]['margem_liquida'] = ll / rl if ll else None

    # --- Operational ---
    ws_op = wb['Dados Operacionais']
    op_col_map = {}
    for c in range(3, ws_op.max_column + 1):
        v = ws_op.cell(4, c).value
        if v:
            op_col_map[str(v)] = c

    op_rows = {
        'vgv_lancado': 6,
        'vgv_lancado_direcional': 9,
        'unidades_lancadas': 12,
        'preco_medio_lancamento': 16,
        'vendas_liquidas_vgv': 19,  # VGV Contratado
        'vendas_liquidas_direcional': 22,
        'unidades_vendidas': 25,
        'preco_medio_vendas': 28,
        'vso_mcmv': 29,
        'vso_consolidada': 30,
        'estoque_vgv': 33,
        'estoque_unidades': 35,
        'landbank_vgv': 40,
        'landbank_direcional': 41,
        'landbank_unidades': 42,
        'repasses_vgv': 45,
        'repasses_unidades': 46,
    }

    for period, col in op_col_map.items():
        if period not in data:
            data[period] = {}
        for metric, row in op_rows.items():
            data[period][metric] = safe_float(ws_op.cell(row, col).value)

    # --- BP ---
    ws_bp = wb['Balanço Patrimonial']
    bp_col_map = {}
    for c in range(3, ws_bp.max_column + 1):
        v = ws_bp.cell(4, c).value
        if v:
            bp_col_map[str(v)] = c

    for period, col in bp_col_map.items():
        if period not in data:
            data[period] = {}
        caixa = safe_float(ws_bp.cell(6, col).value) or 0
        aplic = safe_float(ws_bp.cell(7, col).value) or 0
        data[period]['caixa_aplicacoes'] = caixa + aplic

        emp_cp = safe_float(ws_bp.cell(34, col).value) or 0
        emp_lp = safe_float(ws_bp.cell(50, col).value) or 0
        data[period]['divida_bruta'] = emp_cp + emp_lp
        data[period]['divida_liquida'] = (emp_cp + emp_lp) - (caixa + aplic)

        pl = safe_float(ws_bp.cell(62, col).value)
        data[period]['patrimonio_liquido'] = pl
        if pl and pl != 0:
            data[period]['divida_liquida_pl'] = ((emp_cp + emp_lp) - (caixa + aplic)) / pl

        ativo_total = safe_float(ws_bp.cell(31, col).value)
        data[period]['ativo_total'] = ativo_total

    # --- RIVA DRE ---
    ws_riva = wb['RIVA - DRE']
    riva_col_map = {}
    for c in range(3, ws_riva.max_column + 1):
        v = ws_riva.cell(4, c).value
        if v:
            riva_col_map[str(v)] = c

    for period, col in riva_col_map.items():
        if period not in data:
            data[period] = {}
        data[period]['riva_receita'] = safe_float(ws_riva.cell(5, col).value)
        data[period]['riva_lucro_bruto'] = safe_float(ws_riva.cell(7, col).value)

    wb.close()
    return data


###############################################################################
# PARSER: MRV
###############################################################################
def parse_mrv():
    """Parse MRV spreadsheet."""
    path = os.path.join(PLANILHAS_DIR, 'MRV_Base_Dados_Operacionais_Financeiros_2026-03.xlsx')
    wb = openpyxl.load_workbook(path, data_only=True)
    data = {}

    # --- DRE Consolidado ---
    ws = wb['DRE Consolid. | Income Statem.']
    col_map = {}
    for c in range(3, ws.max_column + 1):
        v = ws.cell(2, c).value
        if v:
            # Extract period: "3T25/3Q25" -> "3T25"
            period = str(v).split('/')[0].strip()
            col_map[period] = c

    dre_rows = {
        'receita_liquida': 3,
        'custo_vendas': 4,
        'lucro_bruto': 5,
        'margem_bruta': 6,
        'despesas_vendas': 7,
        'despesas_ga': 8,
        'lucro_liquido_controladora': 20,
        'margem_liquida': 21,
        'receitas_apropriar': 24,
        'custo_apropriar': 25,
        'resultado_apropriar': 26,
        'margem_ref': 27,
    }

    for period, col in col_map.items():
        if period not in data:
            data[period] = {}
        for metric, row in dre_rows.items():
            data[period][metric] = safe_float(ws.cell(row, col).value)

    # --- Financial Highlights ---
    ws_fin = wb['Indic.Fin. | Financ.Highlights']
    fin_col_map = {}
    for c in range(3, ws_fin.max_column + 1):
        v = ws_fin.cell(2, c).value
        if v:
            period = str(v).split('/')[0].strip()
            fin_col_map[period] = c

    fin_rows = {
        'ebitda': 15,
        'margem_ebitda': 16,
        'roe_12m': 19,
        'roe_anualizado': 20,
        'divida_bruta': 22,
        'caixa_aplicacoes': 23,
        'divida_liquida': 25,
        'patrimonio_liquido': 26,
        'divida_liquida_pl': 27,
        'geracao_caixa_consol': 33,
        'geracao_caixa_mrv_inc': 35,
        'geracao_caixa_resia': 37,
    }

    for period, col in fin_col_map.items():
        if period not in data:
            data[period] = {}
        for metric, row in fin_rows.items():
            data[period][metric] = safe_float(ws_fin.cell(row, col).value)
        # Make caixa positive (it's stored as negative in formula)
        if data[period].get('caixa_aplicacoes'):
            data[period]['caixa_aplicacoes'] = abs(data[period]['caixa_aplicacoes'])

    # --- Operational Data ---
    ws_op = wb['Dados Oper. MRV&Co | Oper.Data']
    op_col_map = {}
    for c in range(3, ws_op.max_column + 1):
        v = ws_op.cell(2, c).value
        if v:
            period = str(v).split('/')[0].strip()
            op_col_map[period] = c

    # MRV&Co %MRV launches
    op_rows = {
        'landbank_vgv': 12,        # Row 12: VGV Land bank (em R$ bilhões) - MRV&Co %MRV
        'landbank_unidades': 13,
        'landbank_preco_medio': 14,
        'vgv_lancado': 59,          # Row 59: VGV Lançamentos MRV&Co %MRV
        'unidades_lancadas': 60,
        'preco_medio_lancamento': 61,
        'vgv_lancado_mrv_inc': 63,
        'vendas_brutas_vgv': 106,   # %MRV
        'vendas_brutas_unidades': 107,
        'vendas_liquidas_vgv': 182, # %MRV
        'vendas_liquidas_unidades': 183,
        'preco_medio_vendas': 184,
    }

    for period, col in op_col_map.items():
        if period not in data:
            data[period] = {}
        for metric, row in op_rows.items():
            data[period][metric] = safe_float(ws_op.cell(row, col).value)

    wb.close()
    return data


###############################################################################
# PARSER: MOURA DUBEUX
###############################################################################
def parse_moura_dubeux():
    """Parse Moura Dubeux spreadsheet."""
    path = os.path.join(PLANILHAS_DIR, 'MouraDubeux_Planilha_Fundamentos_2026-03.xlsx')
    wb = openpyxl.load_workbook(path, data_only=True)
    data = {}

    # --- Operational ---
    ws_op = wb['Dados Operacionais']
    op_col_map = {}
    for c in range(3, ws_op.max_column + 1):
        v = ws_op.cell(5, c).value
        if v:
            op_col_map[str(v)] = c

    op_rows = {
        'vgv_bruto_lancado': 9,
        'vgv_liquido_lancado': 10,
        'vgv_liquido_md': 11,
        'lancamentos_projetos': 13,
        'lancamentos_unidades': 14,
        'vendas_brutas_vgv': 34,
        'vendas_brutas_md': 35,
        'distratos_vgv': 52,
        'vendas_liquidas_vgv': 69,
        'vendas_liquidas_md': 70,
        'estoque_incorp': 87,
        'estoque_condo': 88,
        'landbank_vgv_bruto': 104,
        'landbank_vgv_liquido': 113,
    }

    for period, col in op_col_map.items():
        if period not in data:
            data[period] = {}
        for metric, row in op_rows.items():
            data[period][metric] = safe_float(ws_op.cell(row, col).value)

    # --- DRE ---
    ws_dre = wb['DRE']
    dre_col_map = {}
    for c in range(3, ws_dre.max_column + 1):
        v = ws_dre.cell(5, c).value
        if v:
            dre_col_map[str(v)] = c

    dre_rows = {
        'receita_bruta': 7,
        'receita_liquida': 22,
        'custo_vendas': 24,
        'lucro_bruto': 31,
        'lucro_bruto_ajustado': 32,
        'margem_bruta': 33,
        'margem_bruta_ajustada': 35,
        'despesas_vendas': 38,
        'despesas_ga': 39,
        'lucro_operacional': 43,
        'resultado_financeiro': 49,
        'lucro_liquido': 55,
        'lucro_liquido_controladora': 56,
        'margem_liquida': 59,
        'ebitda_ajustado': 89,
        'margem_ebitda': 90,
    }

    for period, col in dre_col_map.items():
        if period not in data:
            data[period] = {}
        for metric, row in dre_rows.items():
            data[period][metric] = safe_float(ws_dre.cell(row, col).value)

    # --- BP ---
    ws_bp = wb['Balanço Patrimonial']
    bp_col_map = {}
    for c in range(3, ws_bp.max_column + 1):
        v = ws_bp.cell(5, c).value
        if v:
            bp_col_map[str(v)] = c

    for period, col in bp_col_map.items():
        if period not in data:
            data[period] = {}
        caixa = safe_float(ws_bp.cell(10, col).value) or 0
        aplic = safe_float(ws_bp.cell(11, col).value) or 0
        data[period]['caixa_aplicacoes'] = caixa + aplic

        emp_cp = safe_float(ws_bp.cell(33, col).value) or 0
        deb_cp = safe_float(ws_bp.cell(34, col).value) or 0
        emp_lp = 0
        # Find LP emprestimos
        for row_idx in range(43, 60):
            label = ws_bp.cell(row_idx, 2).value
            if label and 'Empr' in str(label):
                emp_lp = safe_float(ws_bp.cell(row_idx, col).value) or 0
                break

        data[period]['divida_bruta'] = emp_cp + deb_cp + emp_lp
        data[period]['divida_liquida'] = (emp_cp + deb_cp + emp_lp) - (caixa + aplic)

        # Find PL (case-insensitive: label can be PATRIMÔNIO or Patrimônio)
        for row_idx in range(55, 80):
            label = ws_bp.cell(row_idx, 2).value
            if label and 'PATRIM' in str(label).upper() and 'CONTROLADOR' not in str(label).upper():
                pl = safe_float(ws_bp.cell(row_idx, col).value)
                if pl is not None:
                    data[period]['patrimonio_liquido'] = pl
                    data[period]['divida_liquida_pl'] = data[period]['divida_liquida'] / pl if pl and pl != 0 else None
                break

    wb.close()
    return data


###############################################################################
# PARSER: PLANO&PLANO
###############################################################################
def parse_planoeplano():
    """Parse Plano&Plano spreadsheet."""
    path = os.path.join(PLANILHAS_DIR, 'PlanoePlano_Planilha_Interativa_2026-03.xlsx')
    wb = openpyxl.load_workbook(path, data_only=True)
    data = {}

    # Skip the 'Data' sheet (complex multi-section layout) and use dedicated sheets instead

    # DRE dedicated sheet - has headers in row 7
    ws_dre = wb['DRE Consolidado']
    dre2_col_map = {}
    for c in range(3, ws_dre.max_column + 1):
        v = ws_dre.cell(7, c).value
        if v:
            dre2_col_map[str(v)] = c

    dre2_rows = {
        'receita_liquida': 9,
        'custo_vendas': 10,
        'lucro_bruto': 11,
        'margem_bruta': 12,
        'despesas_vendas': 14,
        'despesas_ga': 15,
        'lucro_liquido': 31,
        'margem_liquida': 33,
        'lucro_liquido_pp': 35,
        'margem_liquida_pp': 37,
        'receitas_apropriar': 40,
        'custo_apropriar': 41,
        'resultado_apropriar': 42,
        'margem_ref': 43,
    }

    for period, col in dre2_col_map.items():
        if period not in data:
            data[period] = {}
        for metric, row in dre2_rows.items():
            val = safe_float(ws_dre.cell(row, col).value)
            if val is not None:
                data[period][metric] = val

    # BP dedicated sheet - has headers in row 7
    ws_bp2 = wb['Balanço Patrimonial Consolidado']
    bp3_col_map = {}
    for c in range(3, ws_bp2.max_column + 1):
        v = ws_bp2.cell(7, c).value
        if v:
            bp3_col_map[str(v)] = c

    for period, col in bp3_col_map.items():
        if period not in data:
            data[period] = {}
        data[period]['caixa'] = safe_float(ws_bp2.cell(10, col).value)
        data[period]['ativo_total'] = safe_float(ws_bp2.cell(26, col).value)
        data[period]['emprestimos_cp'] = safe_float(ws_bp2.cell(30, col).value)
        data[period]['cessao_cp'] = safe_float(ws_bp2.cell(31, col).value)
        data[period]['cri_cp'] = safe_float(ws_bp2.cell(32, col).value)
        data[period]['emprestimos_lp'] = safe_float(ws_bp2.cell(48, col).value)
        data[period]['cessao_lp'] = safe_float(ws_bp2.cell(49, col).value)
        data[period]['cri_lp'] = safe_float(ws_bp2.cell(50, col).value)
        data[period]['patrimonio_liquido'] = safe_float(ws_bp2.cell(72, col).value)

        # Calculate debt
        emp_cp = (safe_float(ws_bp2.cell(30, col).value) or 0) + (safe_float(ws_bp2.cell(31, col).value) or 0) + (safe_float(ws_bp2.cell(32, col).value) or 0)
        emp_lp = (safe_float(ws_bp2.cell(48, col).value) or 0) + (safe_float(ws_bp2.cell(49, col).value) or 0) + (safe_float(ws_bp2.cell(50, col).value) or 0)
        caixa = safe_float(ws_bp2.cell(10, col).value) or 0
        aplic = safe_float(ws_bp2.cell(17, col).value) or 0
        data[period]['caixa_aplicacoes'] = caixa + aplic
        data[period]['divida_bruta'] = emp_cp + emp_lp
        data[period]['divida_liquida'] = (emp_cp + emp_lp) - (caixa + aplic)
        pl = safe_float(ws_bp2.cell(72, col).value)
        if pl and pl != 0:
            data[period]['divida_liquida_pl'] = data[period]['divida_liquida'] / pl

    # Operational sheet (Dados Operacionais) - headers in row 7
    ws_op2 = wb['Dados Operacionais']
    op2_col_map = {}
    for c in range(3, ws_op2.max_column + 1):
        v = ws_op2.cell(7, c).value
        if v:
            op2_col_map[str(v)] = c

    op2_rows = {
        'vendas_brutas_vgv': 8,
        'vendas_brutas_unidades': 9,
        'distratos_vgv': 10,
        'distratos_unidades': 11,
        'vendas_liquidas_vgv': 12,
        'vendas_liquidas_unidades': 13,
        'vendas_liquidas_pp': 14,
        'preco_medio_vendas': 15,
        'pct_distratos': 16,
        'vgv_lancado': 21,
        'unidades_lancadas': 22,
        'preco_medio_lancamento': 23,
        'vgv_lancado_pp': 25,
        'estoque_vgv': 30,
        'estoque_unidades': 31,
        'vso_liquido_12m': 37,
    }

    for period, col in op2_col_map.items():
        if period not in data:
            data[period] = {}
        for metric, row in op2_rows.items():
            val = safe_float(ws_op2.cell(row, col).value)
            if val is not None:
                data[period][metric] = val

    wb.close()
    return data


###############################################################################
# PARSER: TENDA
###############################################################################
def parse_tenda():
    """Parse Tenda spreadsheet."""
    path = os.path.join(PLANILHAS_DIR, 'Tenda_Planilha_Fundamentos_2026-03.xlsx')
    wb = openpyxl.load_workbook(path, data_only=True)
    data = {}

    # --- Consolidado DRE ---
    ws = wb['Consolidado - DRE']
    # Row 7 has headers: col 2=3T25, col 3=2T25, etc.
    col_map = {}
    for c in range(2, ws.max_column + 1):
        v = ws.cell(7, c).value
        if v:
            col_map[str(v)] = c

    dre_rows = {
        'receita_liquida': 8,
        'custos': 9,
        'lucro_bruto': 10,
        'lucro_bruto_ajustado': 11,
        'margem_bruta': 12,
        'margem_bruta_ajustada': 13,
        'despesas_operacionais': 14,
        'despesas_vendas': 15,
        'despesas_ga': 16,
        'outras_despesas': 17,
        'depreciacao': 18,
        'equivalencia': 19,
        'lucro_operacional': 20,
        'receita_financeira': 21,
        'despesa_financeira': 22,
        'lucro_antes_ir': 23,
        'impostos': 25,
        'lucro_liquido_antes': 26,
        'minoritarios': 27,
        'lucro_liquido': 28,
        'lucro_liquido_ajustado': 29,
        'margem_liquida': 30,
        'margem_liquida_ajustada': 31,
    }

    for period, col in col_map.items():
        if period not in data:
            data[period] = {}
        for metric, row in dre_rows.items():
            data[period][metric] = safe_float(ws.cell(row, col).value)

    # --- Consolidado Financeiro ---
    ws_fin = wb['Consolidado - Financeiro']
    fin_col_map = {}
    for c in range(2, ws_fin.max_column + 1):
        v = ws_fin.cell(7, c).value
        if v:
            fin_col_map[str(v)] = c

    fin_rows = {
        'receita_liquida_fin': 9,
        'lucro_bruto_ajustado_fin': 10,
        'margem_bruta_ajustada_fin': 11,
        'ebitda_ajustado': 13,
        'margem_ebitda_ajustada': 14,
        'resultado_financeiro': 15,
        'lucro_liquido_fin': 16,
        'receitas_apropriar': 17,
        'resultados_apropriar': 18,
        'margem_ref': 19,
        'caixa_disponibilidades': 20,
        'divida_liquida': 21,
        'divida_liquida_ajustada': 22,
        'patrimonio_liquido': 23,
        'divida_liquida_pl': 24,
        'roe': 29,
    }

    for period, col in fin_col_map.items():
        if period not in data:
            data[period] = {}
        for metric, row in fin_rows.items():
            data[period][metric] = safe_float(ws_fin.cell(row, col).value)

    # --- Consolidado Operacional ---
    ws_op = wb['Consolidado - Operacional']
    op_col_map = {}
    for c in range(2, ws_op.max_column + 1):
        v = ws_op.cell(7, c).value
        if v:
            op_col_map[str(v)] = c

    # Note: rows 9-14 are summary, rows 18+ are detailed. Use detailed to avoid duplicates.
    op_rows = {
        'unidades_entregues': 12,
        'landbank_vgv': 13,       # Banco de Terrenos (summary)
        'estoque_vgv': 14,        # Estoque (summary)
        'num_empreendimentos': 18,
        'vgv_lancado': 19,        # Use detailed row for VGV Lançado
        'unidades_lancadas': 20,
        'preco_medio_lancamento': 21,
        'vendas_brutas': 27,
        'vso_bruto': 28,
        'distratos': 29,
        'vendas_liquidas_vgv': 30, # Use detailed row for Vendas Líquidas
        'vso_liquido': 31,         # Use detailed row for VSO
        'vendas_brutas_unid': 34,
        'distratos_unid': 35,
        'vendas_liquidas_unid': 36,
        'vgv_repassado': 45,
        'unidades_repassadas': 46,
        'obras_andamento': 48,
        'estoque_unidades': 53,
        'estoque_preco_medio': 54,
        'landbank_empreendimentos': 58,
        'landbank_unidades': 61,
        'landbank_preco_medio': 62,
    }

    for period, col in op_col_map.items():
        if period not in data:
            data[period] = {}
        for metric, row in op_rows.items():
            data[period][metric] = safe_float(ws_op.cell(row, col).value)

    # --- Consolidado BP ---
    ws_bp = wb['Consolidado - BP']
    bp_col_map = {}
    for c in range(2, ws_bp.max_column + 1):
        v = ws_bp.cell(7, c).value
        if v:
            bp_col_map[str(v)] = c

    for period, col in bp_col_map.items():
        if period not in data:
            data[period] = {}
        data[period]['caixa_bp'] = safe_float(ws_bp.cell(9, col).value)
        data[period]['ativo_total'] = safe_float(ws_bp.cell(19, col).value)
        data[period]['emprestimos_cp'] = safe_float(ws_bp.cell(21, col).value)
        data[period]['debentures_cp'] = safe_float(ws_bp.cell(22, col).value)
        data[period]['emprestimos_lp'] = safe_float(ws_bp.cell(30, col).value)
        data[period]['debentures_lp'] = safe_float(ws_bp.cell(31, col).value)
        data[period]['pl_total'] = safe_float(ws_bp.cell(37, col).value)

    wb.close()
    return data


###############################################################################
# PARSER: CYRELA
###############################################################################
def parse_cyrela():
    """Parse Cyrela spreadsheets."""
    data = {}

    # --- Dados Operacionais ---
    path_op = os.path.join(PLANILHAS_DIR, 'Cyrela_Dados_Operacionais_2026-03.xlsx')
    wb_op = openpyxl.load_workbook(path_op, data_only=True)

    # Lançamentos
    ws_l = wb_op['Lçtos']
    # Row 4 has headers
    l_col_map = {}
    for c in range(2, ws_l.max_column + 1):
        v = ws_l.cell(4, c).value
        if v:
            l_col_map[str(v)] = c

    # VGV Lançado Total (100%) = row 14
    # VGV Lançado %CBR = row 27
    # Unidades Lançadas Total = row 71
    for period, col in l_col_map.items():
        if period not in data:
            data[period] = {}
        data[period]['vgv_lancado_100'] = safe_float(ws_l.cell(14, col).value)
        data[period]['vgv_lancado_cbr'] = safe_float(ws_l.cell(27, col).value)
        data[period]['unidades_lancadas'] = safe_float(ws_l.cell(71, col).value)

    # Vendas
    ws_v = wb_op['Vendas']
    v_col_map = {}
    for c in range(2, ws_v.max_column + 1):
        v = ws_v.cell(4, c).value
        if v:
            v_col_map[str(v)] = c

    for period, col in v_col_map.items():
        if period not in data:
            data[period] = {}
        data[period]['vendas_vgv_100'] = safe_float(ws_v.cell(14, col).value)
        data[period]['vendas_vgv_cbr'] = safe_float(ws_v.cell(27, col).value)
        data[period]['vendas_unidades'] = safe_float(ws_v.cell(71, col).value) if ws_v.max_row >= 71 else None

    # Estoque
    ws_e = wb_op['Estoque']
    e_col_map = {}
    for c in range(2, ws_e.max_column + 1):
        v = ws_e.cell(4, c).value
        if v:
            e_col_map[str(v)] = c

    for period, col in e_col_map.items():
        if period not in data:
            data[period] = {}
        data[period]['estoque_vgv_100'] = safe_float(ws_e.cell(14, col).value)
        data[period]['estoque_vgv_cbr'] = safe_float(ws_e.cell(36, col).value)

    # Terrenos (Landbank)
    ws_t = wb_op['Terrenos']
    t_col_map = {}
    for c in range(2, ws_t.max_column + 1):
        v = ws_t.cell(4, c).value
        if v:
            t_col_map[str(v)] = c

    for period, col in t_col_map.items():
        if period not in data:
            data[period] = {}
        # Row structure may vary - check
        data[period]['landbank_vgv_100'] = safe_float(ws_t.cell(14, col).value) if ws_t.max_row >= 14 else None

    wb_op.close()

    # --- Demonstrações Financeiras (Economatica format) ---
    path_df = os.path.join(PLANILHAS_DIR, 'Cyrela_Demonstracoes_Financeiras_2026-03.xlsx')
    wb_df = openpyxl.load_workbook(path_df, data_only=True)
    ws = wb_df['CYRE3']

    # Row 4 has date headers (row 3 has dates for older data, row 4 for all)
    df_col_map = {}
    for c in range(2, ws.max_column + 1):
        v = ws.cell(4, c).value
        if v is None:
            v = ws.cell(3, c).value  # fallback to row 3
        if v and isinstance(v, datetime):
            month = v.month
            year = v.year
            if month == 3: key = f'1T{str(year)[2:]}'
            elif month == 6: key = f'2T{str(year)[2:]}'
            elif month == 9: key = f'3T{str(year)[2:]}'
            elif month == 12: key = f'4T{str(year)[2:]}'
            else: continue
            df_col_map[key] = c

    # Key financial rows
    df_rows = {
        'receita_liquida': 71,
        'custo_vendas': 72,
        'lucro_bruto': 73,
        'despesas_vendas': 75,
        'despesas_ga': 76,
        'equivalencia': 80,
        'ebit': 81,
        'resultado_financeiro': 82,
        'receitas_financeiras': 83,
        'despesas_financeiras': 84,
        'lair': 85,
        'ir_csll': 86,
        'lucro_liquido_consol': 93,
        'minoritarios': 94,
        'lucro_liquido': 95,
        # BP
        'ativo_total': 10,
        'caixa': 12,
        'aplicacoes': 13,
        'emprestimos_cp': 41,
        'debentures_cp': 43,
        'emprestimos_lp': 49,
        'patrimonio_liquido': 57,
    }

    for period, col in df_col_map.items():
        if period not in data:
            data[period] = {}
        for metric, row in df_rows.items():
            data[period][metric] = safe_float(ws.cell(row, col).value)

        # Calculate derived metrics
        rl = data[period].get('receita_liquida')
        lb = data[period].get('lucro_bruto')
        ll = data[period].get('lucro_liquido')
        if rl and rl != 0:
            data[period]['margem_bruta'] = lb / rl if lb else None
            data[period]['margem_liquida'] = ll / rl if ll else None

        caixa = (data[period].get('caixa') or 0) + (data[period].get('aplicacoes') or 0)
        data[period]['caixa_aplicacoes'] = caixa
        emp = (data[period].get('emprestimos_cp') or 0) + (data[period].get('debentures_cp') or 0) + (data[period].get('emprestimos_lp') or 0)
        data[period]['divida_bruta'] = emp
        data[period]['divida_liquida'] = emp - caixa
        pl = data[period].get('patrimonio_liquido')
        if pl and pl != 0:
            data[period]['divida_liquida_pl'] = (emp - caixa) / pl

    wb_df.close()

    # --- Principais Indicadores ---
    path_pi = os.path.join(PLANILHAS_DIR, 'Cyrela_Principais_Indicadores_2026-03.xlsx')
    wb_pi = openpyxl.load_workbook(path_pi, data_only=True)
    ws_pi = wb_pi['Resumo']

    # This has just 3T25, 3T24, 2T25 - limited but has backlog
    pi_col_map = {}
    for c in range(2, 8):
        v = ws_pi.cell(3, c).value
        if v:
            period = str(v).split('/')[0].strip()
            pi_col_map[period] = c

    for period, col in pi_col_map.items():
        if period not in data:
            data[period] = {}
        data[period]['receitas_apropriar'] = safe_float(ws_pi.cell(31, col).value)
        data[period]['margem_ref'] = safe_float(ws_pi.cell(32, col).value)
        data[period]['geracao_caixa'] = safe_float(ws_pi.cell(28, col).value)

    wb_pi.close()

    return data


###############################################################################
# GENERATE QUANTI MARKDOWN
###############################################################################
def generate_quanti_md(empresa, ticker, data, unit_label='R$ mil'):
    """Generate quantitative analysis markdown for a company."""

    # Filter to target quarters only
    quarters = [q for q in TARGET_QUARTERS if q in data]
    years = [str(y) for y in TARGET_YEARS if str(y) in data]

    lines = []
    lines.append(f'# Análise Quantitativa: {empresa} ({ticker})')
    lines.append(f'')
    lines.append(f'> Dados extraídos das planilhas de RI. Unidade: {unit_label} (salvo indicação contrária).')
    lines.append(f'> Período: 1T2022 a último trimestre disponível.')
    lines.append(f'')

    # Determine which metrics are available
    all_metrics = set()
    for q in quarters:
        all_metrics.update(data[q].keys())

    # --- DRE Table ---
    lines.append('## DRE - Demonstração de Resultados')
    lines.append('')

    # Show annual + quarterly for recent periods
    # Annual columns
    annual_periods = years
    # Recent quarters (last 8)
    recent_q = quarters[-12:] if len(quarters) > 12 else quarters

    display_periods = recent_q

    # DRE metrics to show
    dre_display = []

    # Map metric names based on what's available
    metric_names = {
        'receita_liquida': 'Receita Líquida',
        'receita_liquida_fin': 'Receita Líquida',
        'lucro_bruto': 'Lucro Bruto',
        'lucro_bruto_ajustado': 'Lucro Bruto Ajust.',
        'lucro_bruto_ajustado_fin': 'Lucro Bruto Ajust.',
        'margem_bruta': 'Margem Bruta',
        'margem_bruta_ajustada': 'Margem Bruta Ajust.',
        'margem_bruta_ajustada_fin': 'Margem Bruta Ajust.',
        'ebitda': 'EBITDA',
        'ebitda_ajustado': 'EBITDA Ajustado',
        'margem_ebitda': 'Margem EBITDA',
        'margem_ebitda_ajustada': 'Margem EBITDA Ajust.',
        'lucro_liquido': 'Lucro Líquido',
        'lucro_liquido_controladora': 'LL Controladora',
        'lucro_liquido_ajustado': 'LL Ajustado',
        'lucro_liquido_fin': 'Lucro Líquido',
        'margem_liquida': 'Margem Líquida',
        'margem_liquida_ajustada': 'Margem Líq. Ajust.',
        'roe': 'ROE',
        'roe_12m': 'ROE 12m',
        'roe_anualizado': 'ROE Anualizado',
    }

    # Build DRE table based on available metrics
    dre_keys = ['receita_liquida', 'receita_liquida_fin', 'lucro_bruto', 'lucro_bruto_ajustado',
                'lucro_bruto_ajustado_fin', 'margem_bruta', 'margem_bruta_ajustada',
                'margem_bruta_ajustada_fin',
                'ebitda', 'ebitda_ajustado', 'margem_ebitda', 'margem_ebitda_ajustada',
                'lucro_liquido', 'lucro_liquido_controladora', 'lucro_liquido_ajustado',
                'lucro_liquido_fin',
                'margem_liquida', 'margem_liquida_ajustada', 'roe', 'roe_12m', 'roe_anualizado']

    available_dre = [k for k in dre_keys if k in all_metrics and k in metric_names]
    # Remove duplicates (prefer specific over generic)
    seen_names = set()
    final_dre = []
    for k in available_dre:
        name = metric_names[k]
        if name not in seen_names:
            final_dre.append(k)
            seen_names.add(name)

    if final_dre:
        header = '| Métrica | ' + ' | '.join(display_periods) + ' |'
        separator = '|---|' + '---|' * len(display_periods)
        lines.append(header)
        lines.append(separator)

        pct_metrics = {'margem_bruta', 'margem_bruta_ajustada', 'margem_bruta_ajustada_fin',
                       'margem_ebitda', 'margem_ebitda_ajustada', 'margem_liquida',
                       'margem_liquida_ajustada', 'roe', 'roe_12m', 'roe_anualizado'}

        for key in final_dre:
            name = metric_names[key]
            values = []
            for p in display_periods:
                v = data.get(p, {}).get(key)
                if key in pct_metrics:
                    values.append(fmt_pct(v))
                else:
                    values.append(fmt_money(v))
            lines.append(f'| {name} | ' + ' | '.join(values) + ' |')

        lines.append('')

    # --- Backlog ---
    backlog_keys = ['receitas_apropriar', 'resultado_apropriar', 'resultados_apropriar', 'margem_ref']
    has_backlog = any(k in all_metrics for k in backlog_keys)
    if has_backlog:
        lines.append('## Backlog (Resultado a Apropriar)')
        lines.append('')
        header = '| Métrica | ' + ' | '.join(display_periods) + ' |'
        lines.append(header)
        lines.append('|---|' + '---|' * len(display_periods))

        backlog_names = {
            'receitas_apropriar': 'Receitas a Apropriar',
            'resultado_apropriar': 'Resultado a Apropriar',
            'resultados_apropriar': 'Resultado a Apropriar',
            'margem_ref': 'Margem REF',
        }
        for key in backlog_keys:
            if key in all_metrics:
                name = backlog_names[key]
                values = []
                for p in display_periods:
                    v = data.get(p, {}).get(key)
                    if 'margem' in key:
                        values.append(fmt_pct(v))
                    else:
                        values.append(fmt_money(v))
                lines.append(f'| {name} | ' + ' | '.join(values) + ' |')
        lines.append('')

    # --- Operational Table ---
    lines.append('## Dados Operacionais')
    lines.append('')

    op_keys_map = {
        'vgv_lancado': 'VGV Lançado',
        'vgv_lancado_100': 'VGV Lançado 100%',
        'vgv_lancado_cbr': 'VGV Lançado %CBR',
        'vgv_lancado_cury': 'VGV Lançado %Cury',
        'vgv_lancado_mrv_inc': 'VGV Lançado MRV Inc',
        'vgv_bruto_lancado': 'VGV Bruto Lançado',
        'vgv_liquido_lancado': 'VGV Líquido Lançado',
        'lancamentos': 'Lançamentos VGV',
        'unidades_lancadas': 'Unidades Lançadas',
        'preco_medio_lancamento': 'Preço Médio Lançam.',
        'vendas_liquidas_vgv': 'Vendas Líquidas VGV',
        'vendas_liquidas': 'Vendas Líquidas',
        'vendas_liquidas_md': 'Vendas Líq. %MD',
        'vendas_vgv_100': 'Vendas VGV 100%',
        'vendas_vgv_cbr': 'Vendas VGV %CBR',
        'vendas_brutas_vgv': 'Vendas Brutas VGV',
        'vendas_brutas': 'Vendas Brutas',
        'vso_liquido': 'VSO Líquido',
        'vso_liquida': 'VSO Líquida',
        'vso_bruto': 'VSO Bruto',
        'vso_liquido_12m': 'VSO Líq. 12m',
        'vso_udm': 'VSO UDM',
        'vso_consolidada': 'VSO Consolidada',
        'estoque_vgv': 'Estoque VGV',
        'estoque_vgv_100': 'Estoque VGV 100%',
        'estoque_vgv_det': 'Estoque VGV',
        'estoque_unidades': 'Estoque (unid.)',
        'landbank_vgv': 'Landbank VGV',
        'landbank_vgv_100': 'Landbank VGV 100%',
        'landbank_vgv_det': 'Landbank VGV',
        'landbank_unidades': 'Landbank (unid.)',
        'landbank_vgv_bruto': 'Landbank VGV Bruto',
        'vgv_repassado': 'VGV Repassado',
        'unidades_repassadas': 'Unid. Repassadas',
        'unidades_entregues': 'Unid. Entregues',
        'obras_andamento': 'Obras Andamento',
        'geracao_caixa': 'Geração Caixa',
        'geracao_caixa_consol': 'Geração Caixa Consol.',
    }

    op_ordered = ['vgv_lancado', 'vgv_lancado_100', 'vgv_lancado_cbr', 'vgv_lancado_cury',
                  'vgv_bruto_lancado', 'vgv_liquido_lancado', 'lancamentos',
                  'unidades_lancadas', 'preco_medio_lancamento',
                  'vendas_liquidas_vgv', 'vendas_liquidas', 'vendas_vgv_100', 'vendas_vgv_cbr',
                  'vendas_brutas_vgv', 'vendas_brutas', 'vendas_liquidas_md',
                  'vso_liquido', 'vso_liquida', 'vso_bruto', 'vso_liquido_12m', 'vso_udm', 'vso_consolidada',
                  'estoque_vgv', 'estoque_vgv_100', 'estoque_vgv_det', 'estoque_unidades',
                  'landbank_vgv', 'landbank_vgv_100', 'landbank_vgv_det', 'landbank_unidades',
                  'landbank_vgv_bruto',
                  'vgv_repassado', 'unidades_repassadas', 'unidades_entregues',
                  'obras_andamento', 'geracao_caixa', 'geracao_caixa_consol']

    available_op = [k for k in op_ordered if k in all_metrics and k in op_keys_map]
    # Remove duplicate names
    seen_op = set()
    final_op = []
    for k in available_op:
        name = op_keys_map[k]
        if name not in seen_op:
            final_op.append(k)
            seen_op.add(name)

    if final_op:
        header = '| Métrica | ' + ' | '.join(display_periods) + ' |'
        lines.append(header)
        lines.append('|---|' + '---|' * len(display_periods))

        vso_keys = {'vso_liquido', 'vso_liquida', 'vso_bruto', 'vso_liquido_12m', 'vso_udm', 'vso_consolidada'}

        for key in final_op:
            name = op_keys_map[key]
            values = []
            for p in display_periods:
                v = data.get(p, {}).get(key)
                if key in vso_keys:
                    values.append(fmt_pct(v))
                else:
                    values.append(fmt_money(v))
            lines.append(f'| {name} | ' + ' | '.join(values) + ' |')
        lines.append('')

    # --- Balance Sheet / Endividamento ---
    lines.append('## Endividamento e Balanço')
    lines.append('')

    bs_keys_map = {
        'caixa_aplicacoes': 'Caixa + Aplicações',
        'caixa_disponibilidades': 'Caixa + Disponib.',
        'caixa': 'Caixa',
        'divida_bruta': 'Dívida Bruta',
        'divida_liquida': 'Dívida Líquida',
        'divida_liquida_ajustada': 'Dív. Líq. Ajustada',
        'patrimonio_liquido': 'Patrimônio Líquido',
        'patrimonio_liquido_total': 'PL Total',
        'pl_total': 'PL Total',
        'divida_liquida_pl': 'Dív. Líq./PL',
    }

    bs_ordered = ['caixa_aplicacoes', 'caixa_disponibilidades', 'caixa',
                  'divida_bruta', 'divida_liquida', 'divida_liquida_ajustada',
                  'patrimonio_liquido', 'patrimonio_liquido_total', 'pl_total',
                  'divida_liquida_pl']

    available_bs = [k for k in bs_ordered if k in all_metrics and k in bs_keys_map]
    seen_bs = set()
    final_bs = []
    for k in available_bs:
        name = bs_keys_map[k]
        if name not in seen_bs:
            final_bs.append(k)
            seen_bs.add(name)

    if final_bs:
        header = '| Métrica | ' + ' | '.join(display_periods) + ' |'
        lines.append(header)
        lines.append('|---|' + '---|' * len(display_periods))

        for key in final_bs:
            name = bs_keys_map[key]
            values = []
            for p in display_periods:
                v = data.get(p, {}).get(key)
                if key == 'divida_liquida_pl':
                    values.append(fmt_ratio(v))
                else:
                    values.append(fmt_money(v))
            lines.append(f'| {name} | ' + ' | '.join(values) + ' |')
        lines.append('')

    # --- Annual Summary ---
    if years:
        lines.append('## Resumo Anual')
        lines.append('')
        avail_years = [y for y in years if y in data]
        if avail_years:
            header = '| Métrica | ' + ' | '.join(avail_years) + ' |'
            lines.append(header)
            lines.append('|---|' + '---|' * len(avail_years))

            annual_keys = ['receita_liquida', 'receita_liquida_fin', 'lucro_bruto',
                          'lucro_liquido', 'lucro_liquido_controladora',
                          'vgv_lancado', 'vgv_lancado_100', 'vgv_bruto_lancado',
                          'vendas_liquidas_vgv', 'vendas_vgv_100']

            annual_names = {**metric_names, **op_keys_map}

            # Deduplicate by display name
            seen_annual = set()
            for key in annual_keys:
                if key in all_metrics:
                    name = annual_names.get(key, key)
                    if name in seen_annual:
                        continue
                    seen_annual.add(name)
                    values = []
                    for y in avail_years:
                        v = data.get(y, {}).get(key)
                        values.append(fmt_money(v))
                    lines.append(f'| {name} | ' + ' | '.join(values) + ' |')
            lines.append('')

    # --- Trends & Conclusions ---
    lines.append('## Tendências e Indicadores-Chave')
    lines.append('')

    # Find latest quarter with actual revenue data
    latest_q = None
    for q in reversed(quarters):
        if data[q].get('receita_liquida') or data[q].get('receita_liquida_fin'):
            latest_q = q
            break

    # Calculate YoY growth for latest quarter
    if latest_q and len(quarters) >= 5:
        # Find same quarter previous year
        q_num = latest_q[0]
        y_num = int(latest_q[2:])
        yoy_q = f'{q_num}T{y_num-1:02d}'
        if yoy_q in data:
            rl_now = data[latest_q].get('receita_liquida') or data[latest_q].get('receita_liquida_fin')
            rl_prev = data[yoy_q].get('receita_liquida') or data[yoy_q].get('receita_liquida_fin')
            if rl_now and rl_prev and rl_prev != 0:
                yoy = (rl_now / rl_prev - 1) * 100
                lines.append(f'- **Crescimento Receita YoY ({latest_q} vs {yoy_q}):** {yoy:.1f}%')

            lb_now = data[latest_q].get('lucro_bruto') or data[latest_q].get('lucro_bruto_ajustado')
            lb_prev = data[yoy_q].get('lucro_bruto') or data[yoy_q].get('lucro_bruto_ajustado')
            if lb_now and lb_prev and lb_prev != 0:
                yoy_lb = (lb_now / lb_prev - 1) * 100
                lines.append(f'- **Crescimento Lucro Bruto YoY:** {yoy_lb:.1f}%')

            ll_now = data[latest_q].get('lucro_liquido') or data[latest_q].get('lucro_liquido_controladora')
            ll_prev = data[yoy_q].get('lucro_liquido') or data[yoy_q].get('lucro_liquido_controladora')
            if ll_now and ll_prev and ll_prev != 0:
                yoy_ll = (ll_now / ll_prev - 1) * 100
                lines.append(f'- **Crescimento Lucro Líquido YoY:** {yoy_ll:.1f}%')

    # CAGR of revenue (first to last available year)
    avail_annual = [y for y in [str(yr) for yr in TARGET_YEARS] if y in data]
    if len(avail_annual) >= 2:
        first_y = avail_annual[0]
        last_y = avail_annual[-1]
        rl_first = data[first_y].get('receita_liquida') or data[first_y].get('receita_liquida_fin')
        rl_last = data[last_y].get('receita_liquida') or data[last_y].get('receita_liquida_fin')
        n_years = int(last_y) - int(first_y)
        if rl_first and rl_last and rl_first > 0 and n_years > 0:
            cagr = ((rl_last / rl_first) ** (1 / n_years) - 1) * 100
            lines.append(f'- **CAGR Receita ({first_y}-{last_y}):** {cagr:.1f}%')

    # Margin evolution
    if len(quarters) >= 8 and latest_q:
        first_q = quarters[-8]
        last_q = latest_q
        mb_first = data[first_q].get('margem_bruta') or data[first_q].get('margem_bruta_ajustada')
        mb_last = data[last_q].get('margem_bruta') or data[last_q].get('margem_bruta_ajustada')
        if mb_first is not None and mb_last is not None:
            mb_f = mb_first * 100 if abs(mb_first) < 1 else mb_first
            mb_l = mb_last * 100 if abs(mb_last) < 1 else mb_last
            delta = mb_l - mb_f
            direction = '↑' if delta > 0 else '↓' if delta < 0 else '→'
            lines.append(f'- **Margem Bruta ({first_q} → {last_q}):** {mb_f:.1f}% → {mb_l:.1f}% ({direction}{abs(delta):.1f} p.p.)')

    # Leverage evolution
    if latest_q:
        last_q = latest_q
        dl_pl = data[last_q].get('divida_liquida_pl')
        if dl_pl is not None:
            status = 'Caixa líquido' if dl_pl < 0 else 'Alavancagem baixa' if dl_pl < 0.3 else 'Alavancagem moderada' if dl_pl < 0.6 else 'Alavancagem elevada'
            lines.append(f'- **Dív. Líq./PL ({last_q}):** {dl_pl*100:.1f}% — {status}')

    lines.append('')

    return '\n'.join(lines)


###############################################################################
# COMPLEMENT FROM DB
###############################################################################
def complement_from_db(empresa_db, data):
    """Complement spreadsheet data with database for missing quarters."""
    try:
        conn = sqlite3.connect(DB_PATH)
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()

        cursor.execute("""
            SELECT * FROM dados_trimestrais
            WHERE empresa = ? AND segmento = 'Consolidado'
            ORDER BY ano, trimestre
        """, (empresa_db,))

        rows = cursor.fetchall()

        db_map = {
            'vgv_lancado': 'vgv_lancado',
            'unidades_lancadas': 'unidades_lancadas',
            'vendas_liquidas_vgv': 'vendas_liquidas_vgv',
            'vendas_liquidas_unidades': 'vendas_liquidas_unidades',
            'receita_liquida': 'receita_liquida',
            'lucro_bruto': 'lucro_bruto',
            'margem_bruta': 'margem_bruta',
            'ebitda': 'ebitda',
            'margem_ebitda': 'margem_ebitda',
            'lucro_liquido': 'lucro_liquido',
            'margem_liquida': 'margem_liquida',
            'roe': 'roe',
            'divida_liquida': 'divida_liquida',
            'divida_liquida_pl': 'divida_liquida_pl',
            'patrimonio_liquido': 'patrimonio_liquido',
            'estoque_vgv': 'estoque_vgv',
            'landbank_vgv': 'landbank_vgv',
            'receitas_apropriar': 'receitas_apropriar',
            'margem_apropriar': 'margem_ref',
            'geracao_caixa': 'geracao_caixa',
            'caixa_aplicacoes': 'caixa_aplicacoes',
            'ebitda_ajustado': 'ebitda_ajustado',
            'margem_ebitda_ajustada': 'margem_ebitda_ajustada',
            'vso_liquida_trimestral': 'vso_liquido',
        }

        for row in rows:
            period_raw = row['periodo']
            # Normalize period format: DB uses '1T2025', spreadsheets use '1T25'
            # Convert to short format to match spreadsheet keys
            import re as _re
            m = _re.match(r'(\d)T(\d{4})', period_raw)
            if m:
                period = f'{m.group(1)}T{m.group(2)[2:]}'
            else:
                period = period_raw

            if period not in data:
                data[period] = {}

            for db_col, local_key in db_map.items():
                db_val = row[db_col]
                # Only fill if not already present from spreadsheet
                if local_key not in data[period] or data[period][local_key] is None:
                    if db_val is not None and db_val != 0:
                        data[period][local_key] = float(db_val)

        conn.close()
    except Exception as e:
        print(f'  DB complement error for {empresa_db}: {e}')

    return data


###############################################################################
# POST-PROCESS: calculate derived metrics where missing
###############################################################################
def post_process_data(data):
    """Calculate derived metrics (margins, etc.) where base values exist but derived don't."""
    for period, metrics in data.items():
        rl = metrics.get('receita_liquida') or metrics.get('receita_liquida_fin')
        if rl and rl != 0:
            # Margem EBITDA
            ebitda = metrics.get('ebitda') or metrics.get('ebitda_ajustado')
            if ebitda and not metrics.get('margem_ebitda') and not metrics.get('margem_ebitda_ajustada'):
                margin = ebitda / rl
                # Sanity check: margin should be between -100% and 100%
                if 0.001 < abs(margin) < 1.0:
                    metrics['margem_ebitda'] = margin
            # Margem Bruta (if missing)
            lb = metrics.get('lucro_bruto') or metrics.get('lucro_bruto_ajustado')
            if lb and not metrics.get('margem_bruta') and not metrics.get('margem_bruta_ajustada'):
                margin = lb / rl
                if 0.001 < abs(margin) < 1.0:
                    metrics['margem_bruta'] = margin
            # Margem Líquida (if missing)
            ll = metrics.get('lucro_liquido') or metrics.get('lucro_liquido_controladora')
            if ll and not metrics.get('margem_liquida') and not metrics.get('margem_liquida_ajustada'):
                margin = ll / rl
                if abs(margin) < 1.0:
                    metrics['margem_liquida'] = margin
    return data


###############################################################################
# MAIN EXECUTION
###############################################################################
def main():
    print('=' * 60)
    print('GERADOR DE ANÁLISE QUANTITATIVA - INCORPORADORAS')
    print('=' * 60)

    companies = {
        'Cury': {
            'ticker': 'CURY3',
            'parser': parse_cury,
            'db_name': 'Cury',
            'unit': 'R$ mil',
            'dre_divisor': 1000,  # R$ mil → R$ milhões
            'slug': 'cury',
        },
        'Direcional': {
            'ticker': 'DIRR3',
            'parser': parse_direcional,
            'db_name': 'Direcional',
            'unit': 'R$ mil',
            'dre_divisor': 1000,
            'slug': 'direcional',
        },
        'MRV': {
            'ticker': 'MRVE3',
            'parser': parse_mrv,
            'db_name': 'MRV',
            'unit': 'R$ mil (DRE/BP)',
            'dre_divisor': 1000,
            'slug': 'mrv',
        },
        'Moura Dubeux': {
            'ticker': 'MDNE3',
            'parser': parse_moura_dubeux,
            'db_name': None,  # Not in DB
            'unit': 'R$ mil',
            'dre_divisor': 1000,
            'slug': 'mouradubeux',
        },
        'Plano&Plano': {
            'ticker': 'PLPL3',
            'parser': parse_planoeplano,
            'db_name': 'PlanoePlano',
            'unit': 'R$ mil',
            'dre_divisor': 1000,
            'slug': 'planoeplano',
        },
        'Tenda': {
            'ticker': 'TEND3',
            'parser': parse_tenda,
            'db_name': 'Tenda',
            'unit': 'R$ milhões',
            'dre_divisor': 1,  # Already in R$ milhões
            'slug': 'tenda',
        },
        'Cyrela': {
            'ticker': 'CYRE3',
            'parser': parse_cyrela,
            'db_name': 'Cyrela',
            'unit': 'R$ mil (Economatica)',
            'dre_divisor': 1000,
            'slug': 'cyrela',
        },
    }

    all_data = {}

    for name, config in companies.items():
        print(f'\n--- Processando {name} ({config["ticker"]}) ---')

        try:
            data = config['parser']()
            print(f'  Planilha: {len(data)} períodos extraídos')

            # Complement from DB
            if config['db_name']:
                data = complement_from_db(config['db_name'], data)
                print(f'  Após DB: {len(data)} períodos')

            # Post-process: calculate derived metrics
            data = post_process_data(data)

            all_data[name] = data

            # Generate markdown
            md = generate_quanti_md(name, config['ticker'], data, config['unit'])

            output_path = os.path.join(OUTPUT_DIR, f'quanti_{config["slug"]}.md')
            with open(output_path, 'w', encoding='utf-8') as f:
                f.write(md)
            print(f'  Arquivo gerado: {output_path}')

        except Exception as e:
            print(f'  ERRO: {e}')
            import traceback
            traceback.print_exc()

    # --- Generate comparative market analysis ---
    print(f'\n--- Gerando análise comparativa de mercado ---')
    generate_quanti_mercado(all_data, companies)

    print(f'\n{"="*60}')
    print('CONCLUÍDO!')
    print(f'{"="*60}')


###############################################################################
# COMPARATIVE MARKET ANALYSIS
###############################################################################
def normalize_to_millions(value, company_name, companies):
    """Normalize a monetary value to R$ milhões based on company's financial unit.
    Uses 'dre_divisor' from company config: 1 if already R$ milhões, 1000 if R$ mil."""
    if value is None:
        return None
    divisor = companies[company_name].get('dre_divisor', 1000)  # Default: R$ mil
    return float(value) / divisor

def generate_quanti_mercado(all_data, companies):
    """Generate cross-company comparative markdown with normalized units."""
    lines = []
    lines.append('# Análise Quantitativa Comparativa - Setor Imobiliário')
    lines.append('')
    lines.append('> Comparativo entre 7 incorporadoras listadas na B3.')
    lines.append('> **Todos os valores monetários normalizados para R$ milhões.**')
    lines.append('> Período: 2022-2025 (até último dado disponível).')
    lines.append('')

    # --- Latest quarter comparison ---
    latest_map = {}
    for name, data in all_data.items():
        for q in reversed(TARGET_QUARTERS):
            if q in data and (data[q].get('receita_liquida') or data[q].get('receita_liquida_fin')):
                latest_map[name] = q
                break

    lines.append('## Último Trimestre Disponível')
    lines.append('')

    # DRE comparison
    lines.append('### DRE Comparativo (R$ milhões)')
    lines.append('')

    company_names = list(companies.keys())
    header = '| Métrica | ' + ' | '.join([f'{n} ({latest_map.get(n, "?")})' for n in company_names]) + ' |'
    lines.append(header)
    lines.append('|---|' + '---|' * len(company_names))

    compare_metrics = [
        ('Receita Líquida', lambda d: d.get('receita_liquida') or d.get('receita_liquida_fin'), 'money'),
        ('Lucro Bruto', lambda d: d.get('lucro_bruto') or d.get('lucro_bruto_ajustado') or d.get('lucro_bruto_ajustado_fin'), 'money'),
        ('Margem Bruta', lambda d: d.get('margem_bruta') or d.get('margem_bruta_ajustada') or d.get('margem_bruta_ajustada_fin'), 'pct'),
        ('EBITDA', lambda d: d.get('ebitda') or d.get('ebitda_ajustado'), 'money'),
        ('Margem EBITDA', lambda d: d.get('margem_ebitda') or d.get('margem_ebitda_ajustada'), 'pct'),
        ('Lucro Líquido', lambda d: d.get('lucro_liquido') or d.get('lucro_liquido_controladora') or d.get('lucro_liquido_fin'), 'money'),
        ('Margem Líquida', lambda d: d.get('margem_liquida') or d.get('margem_liquida_ajustada'), 'pct'),
        ('ROE', lambda d: d.get('roe') or d.get('roe_12m') or d.get('roe_anualizado'), 'pct'),
    ]

    for metric_name, getter, fmt_type in compare_metrics:
        values = []
        for name in company_names:
            q = latest_map.get(name)
            if q and q in all_data.get(name, {}):
                v = getter(all_data[name][q])
                if fmt_type == 'pct':
                    values.append(fmt_pct(v))
                else:
                    norm = normalize_to_millions(v, name, companies)
                    values.append(f'{norm:,.1f}' if norm is not None else '-')
            else:
                values.append('-')
        lines.append(f'| {metric_name} | ' + ' | '.join(values) + ' |')

    lines.append('')

    # Operational comparison
    lines.append('### Operacional Comparativo (R$ milhões)')
    lines.append('')
    header = '| Métrica | ' + ' | '.join([f'{n}' for n in company_names]) + ' |'
    lines.append(header)
    lines.append('|---|' + '---|' * len(company_names))

    op_compare = [
        ('VGV Lançado', lambda d: d.get('vgv_lancado') or d.get('vgv_lancado_100') or d.get('vgv_bruto_lancado'), 'money'),
        ('Vendas Líquidas', lambda d: d.get('vendas_liquidas_vgv') or d.get('vendas_vgv_100'), 'money'),
        ('VSO', lambda d: d.get('vso_liquido') or d.get('vso_liquida') or d.get('vso_consolidada') or d.get('vso_liquido_12m'), 'pct'),
        ('Estoque VGV', lambda d: d.get('estoque_vgv') or d.get('estoque_vgv_100'), 'money'),
        ('Landbank VGV', lambda d: d.get('landbank_vgv') or d.get('landbank_vgv_100') or d.get('landbank_vgv_bruto'), 'money'),
    ]

    for metric_name, getter, fmt_type in op_compare:
        values = []
        for name in company_names:
            q = latest_map.get(name)
            if q and q in all_data.get(name, {}):
                v = getter(all_data[name][q])
                if fmt_type == 'pct':
                    values.append(fmt_pct(v))
                else:
                    norm = normalize_to_millions(v, name, companies)
                    values.append(f'{norm:,.1f}' if norm is not None else '-')
            else:
                values.append('-')
        lines.append(f'| {metric_name} | ' + ' | '.join(values) + ' |')

    lines.append('')

    # Balance sheet comparison (normalized)
    lines.append('### Endividamento Comparativo (R$ milhões)')
    lines.append('')
    header = '| Métrica | ' + ' | '.join([f'{n}' for n in company_names]) + ' |'
    lines.append(header)
    lines.append('|---|' + '---|' * len(company_names))

    bs_compare = [
        ('Dívida Líquida', lambda d: d.get('divida_liquida') or d.get('divida_liquida_ajustada'), 'money'),
        ('PL', lambda d: d.get('patrimonio_liquido') or d.get('patrimonio_liquido_total') or d.get('pl_total'), 'money'),
        ('Dív. Líq./PL', lambda d: d.get('divida_liquida_pl'), 'ratio'),
        ('Margem REF', lambda d: d.get('margem_ref'), 'pct'),
    ]

    for metric_name, getter, fmt_type in bs_compare:
        values = []
        for name in company_names:
            q = latest_map.get(name)
            if q and q in all_data.get(name, {}):
                v = getter(all_data[name][q])
                if fmt_type == 'ratio':
                    values.append(fmt_ratio(v))
                elif fmt_type == 'pct':
                    values.append(fmt_pct(v))
                else:
                    norm = normalize_to_millions(v, name, companies)
                    values.append(f'{norm:,.1f}' if norm is not None else '-')
            else:
                values.append('-')
        lines.append(f'| {metric_name} | ' + ' | '.join(values) + ' |')

    lines.append('')

    # --- Evolution comparison (annual, normalized) ---
    lines.append('## Evolução Anual - Receita Líquida (R$ milhões)')
    lines.append('')
    header = '| Empresa | 2022 | 2023 | 2024 | 2025* | CAGR 22-24 |'
    lines.append(header)
    lines.append('|---|---|---|---|---|---|')

    for name in company_names:
        values = []
        annual_vals = {}
        for y in ['2022', '2023', '2024', '2025']:
            d = all_data.get(name, {}).get(y, {})
            v = d.get('receita_liquida') or d.get('receita_liquida_fin')
            norm = normalize_to_millions(v, name, companies) if v else None
            annual_vals[y] = norm
            values.append(f'{norm:,.1f}' if norm is not None else '-')
        # CAGR 2022-2024
        v22 = annual_vals.get('2022')
        v24 = annual_vals.get('2024')
        if v22 and v24 and v22 > 0:
            cagr = ((v24 / v22) ** 0.5 - 1) * 100
            values.append(f'{cagr:.1f}%')
        else:
            values.append('-')
        lines.append(f'| {name} | ' + ' | '.join(values) + ' |')

    lines.append('')
    lines.append('*2025: acumulado até último trimestre disponível.')
    lines.append('')

    # --- Rankings ---
    lines.append('## Rankings (último trimestre)')
    lines.append('')

    def get_normalized_metric(name, getter):
        q = latest_map.get(name)
        if q and q in all_data.get(name, {}):
            return getter(all_data[name][q])
        return None

    # Ranking by margin
    rankings = [
        ('Margem Bruta', lambda d: d.get('margem_bruta') or d.get('margem_bruta_ajustada')),
        ('Margem Líquida', lambda d: d.get('margem_liquida') or d.get('margem_liquida_ajustada')),
        ('ROE', lambda d: d.get('roe') or d.get('roe_12m') or d.get('roe_anualizado')),
    ]

    for rank_name, getter in rankings:
        scored = []
        for name in company_names:
            v = get_normalized_metric(name, getter)
            if v is not None:
                scored.append((name, v))
        scored.sort(key=lambda x: x[1], reverse=True)
        rank_str = ' > '.join([f'{n} ({fmt_pct(v)})' for n, v in scored])
        lines.append(f'**{rank_name}:** {rank_str}')
        lines.append('')

    # Ranking by receita
    scored_rl = []
    for name in company_names:
        q = latest_map.get(name)
        if q and q in all_data.get(name, {}):
            v = all_data[name][q].get('receita_liquida') or all_data[name][q].get('receita_liquida_fin')
            if v:
                norm = normalize_to_millions(v, name, companies)
                scored_rl.append((name, norm))
    scored_rl.sort(key=lambda x: x[1], reverse=True)
    rank_str = ' > '.join([f'{n} (R${v:,.0f}M)' for n, v in scored_rl])
    lines.append(f'**Receita Líquida:** {rank_str}')
    lines.append('')

    # --- Notes ---
    lines.append('## Notas Metodológicas')
    lines.append('')
    lines.append('- **Normalização**: Todos os valores monetários foram convertidos para R$ milhões. '
                 'Tenda reporta em R$ milhões, demais em R$ mil (÷1.000).')
    lines.append('- **Margem Bruta Ajustada**: Exclui juros capitalizados ao custo (quando disponível).')
    lines.append('- **VSO**: Velocidade Sobre Oferta - vendas líquidas / (vendas líquidas + estoque).')
    lines.append('- **REF**: Resultado a Apropriar (backlog de margem futura).')
    lines.append('- **Cyrela**: Dados financeiros de fonte Economatica; operacionais da planilha de RI.')
    lines.append('- **CAGR**: Taxa de crescimento anual composta (Compound Annual Growth Rate).')
    lines.append('')

    output_path = os.path.join(OUTPUT_DIR, 'quanti_mercado.md')
    with open(output_path, 'w', encoding='utf-8') as f:
        f.write('\n'.join(lines))
    print(f'  Arquivo gerado: {output_path}')


if __name__ == '__main__':
    main()

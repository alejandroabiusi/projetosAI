# -*- coding: utf-8 -*-
"""
Popula dados 2025 (1T-4T) + campos ROCE/ROIC para todas as empresas.
Usa a mesma logica do criar_banco_v2.py mas sem recriar o schema.
Adiciona: ativo_total, passivo_circulante, depreciacao_amortizacao, ir_csll, ebit.
"""
import sqlite3
import openpyxl
import re
import os
from pathlib import Path
from datetime import datetime

_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.path.join(_DIR, "dados_financeiros.db")
PLANILHAS_DIR = Path(_DIR) / "planilhas"

CAMPOS_PERCENTUAIS = {
    'margem_bruta', 'margem_bruta_ajustada', 'margem_ebitda', 'margem_ebitda_ajustada',
    'margem_liquida', 'margem_apropriar',
    'vso_bruta_trimestral', 'vso_liquida_trimestral', 'vso_bruta_12m', 'vso_liquida_12m',
    'pct_comerciais_receita_liquida', 'pct_ga_receita_liquida', 'pct_sga_receita_liquida',
    'pct_estoque_pronto', 'pct_permuta_total',
    'roe', 'roce', 'roic', 'divida_liquida_pl', 'giro_estoque_meses',
    'pdd_cobertura_pct', 'inadimplencia_total_pct',
    'aging_adimplente_pct', 'aging_vencido_90d_pct', 'aging_vencido_360d_pct',
    'aging_vencido_360d_mais_pct', 'carteira_pct_pl', 'carteira_pos_chaves_pct',
    'aliquota_efetiva',
}


def normalizar_periodo(val):
    if val is None:
        return None
    if isinstance(val, datetime):
        month_to_tri = {1: 1, 2: 1, 3: 1, 4: 2, 5: 2, 6: 2, 7: 3, 8: 3, 9: 3, 10: 4, 11: 4, 12: 4}
        tri = month_to_tri[val.month]
        return f"{tri}T{val.year}"
    s = str(val).strip()
    m = re.match(r'^(\d)[Tt](\d{2,4})\s*/\s*\d[Qq]\d', s)
    if m:
        tri, ano = m.group(1), m.group(2)
        if len(ano) == 2: ano = f"20{ano}"
        return f"{tri}T{ano}"
    m = re.match(r'^(\d)[TtQq](\d{2,4})$', s)
    if m:
        tri, ano = m.group(1), m.group(2)
        if len(ano) == 2: ano = f"20{ano}"
        return f"{tri}T{ano}"
    m = re.match(r'^(\d{4})-(\d{2})-\d{2}', s)
    if m:
        ano, mes = int(m.group(1)), int(m.group(2))
        month_to_tri = {3: 1, 6: 2, 9: 3, 12: 4}
        tri = month_to_tri.get(mes)
        if tri:
            return f"{tri}T{ano}"
    return None


def periodo_valido(p):
    if not p: return False
    m = re.match(r'^(\d)T(\d{4})$', p)
    if not m: return False
    tri, ano = int(m.group(1)), int(m.group(2))
    if ano < 2020 or ano > 2026: return False
    return True


def extrair_dados_por_label(filepath, aba, header_row, label_col, field_map, multiplicador=1.0):
    wb = openpyxl.load_workbook(str(filepath), read_only=True, data_only=True)
    ws = wb[aba]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if len(rows) < header_row:
        return {}
    header = rows[header_row - 1]
    col_periods = {}
    for col_idx, val in enumerate(header):
        p = normalizar_periodo(val)
        if p and periodo_valido(p):
            col_periods[col_idx] = p
    if not col_periods:
        return {}
    result = {}
    for row in rows:
        label_val = row[label_col] if label_col < len(row) else None
        if label_val is None:
            continue
        label_str = str(label_val).strip().lstrip('+-=').strip().lower()
        if not label_str:
            continue
        campo = None
        sorted_patterns = sorted(field_map.items(), key=lambda x: len(x[0]), reverse=True)
        for pattern, campo_name in sorted_patterns:
            if label_str.startswith(pattern.lower()) or label_str == pattern.lower():
                campo = campo_name
                break
        if not campo:
            continue
        for col_idx, periodo in col_periods.items():
            if col_idx < len(row) and row[col_idx] is not None:
                try:
                    val = float(row[col_idx])
                    if campo not in CAMPOS_PERCENTUAIS:
                        val *= multiplicador
                    if periodo not in result:
                        result[periodo] = {}
                    result[periodo][campo] = val
                except (ValueError, TypeError):
                    pass
    return result


def extrair_por_linha(filepath, aba, header_row, row_map, multiplicador=1.0):
    wb = openpyxl.load_workbook(str(filepath), read_only=True, data_only=True)
    ws = wb[aba]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if len(rows) < header_row:
        return {}
    header = rows[header_row - 1]
    col_periods = {}
    for col_idx, val in enumerate(header):
        p = normalizar_periodo(val)
        if p and periodo_valido(p):
            col_periods[col_idx] = p
    result = {}
    for row_num, campo in row_map.items():
        if row_num - 1 >= len(rows):
            continue
        row = rows[row_num - 1]
        for col_idx, periodo in col_periods.items():
            if col_idx < len(row) and row[col_idx] is not None:
                try:
                    val = float(row[col_idx]) * multiplicador
                    if periodo not in result:
                        result[periodo] = {}
                    result[periodo][campo] = val
                except (ValueError, TypeError):
                    pass
    return result


def inserir_dados(conn, empresa, segmento, dados_por_periodo):
    c = conn.cursor()
    novos = 0
    atualizados = 0
    for periodo, campos in dados_por_periodo.items():
        m = re.match(r'(\d)T(\d{4})', periodo)
        if not m: continue
        tri, ano = int(m.group(1)), int(m.group(2))
        c.execute("SELECT id FROM dados_trimestrais WHERE empresa=? AND segmento=? AND periodo=?",
                  (empresa, segmento, periodo))
        existing = c.fetchone()
        if existing:
            for campo, valor in campos.items():
                try:
                    c.execute(f"UPDATE dados_trimestrais SET {campo}=? WHERE id=?", (valor, existing[0]))
                except Exception:
                    pass
            atualizados += 1
        else:
            campo_names = ['empresa', 'segmento', 'periodo', 'ano', 'trimestre', 'fonte'] + list(campos.keys())
            campo_vals = [empresa, segmento, periodo, ano, tri, 'planilha'] + list(campos.values())
            placeholders = ','.join(['?'] * len(campo_vals))
            col_str = ','.join(campo_names)
            try:
                c.execute(f"INSERT INTO dados_trimestrais ({col_str}) VALUES ({placeholders})", campo_vals)
                novos += 1
            except Exception:
                pass
    conn.commit()
    return novos, atualizados


# ============================================================
# TENDA
# ============================================================
def ingerir_tenda(conn):
    print("\n  --- TENDA ---")
    fp = PLANILHAS_DIR / "Tenda_Planilha_Fundamentos_2026-03.xlsx"

    operacional_map = {
        'Número de Empreendimentos': 'empreendimentos_lancados',
        'VGV (R$ milhões)': 'vgv_lancado',
        'Número de unidades': 'unidades_lancadas',
        'Preço médio por unidade (R$ mil)': 'preco_medio_lancamento',
        'Tamanho médio dos lançamentos': 'tamanho_medio_empreendimentos',
        'Vendas Brutas': 'vendas_brutas_vgv',
        'VSO Bruto': 'vso_bruta_trimestral',
        'Distratos': 'distratos_vgv',
        'Vendas Líquidas': 'vendas_liquidas_vgv',
        'VSO Líquido': 'vso_liquida_trimestral',
        'Unidades Vendidas Brutas': 'vendas_brutas_unidades',
        'Unidades Distratadas': 'distratos_unidades',
        'Unidades Vendidas Líquidas': 'vendas_liquidas_unidades',
        'Preço médio por unidade bruta': 'preco_medio_vendas',
        'VGV Repassado': 'vgv_repassado',
        'Unidades Repassadas': 'unidades_repassadas',
        'Unidades Entregues': 'unidades_entregues',
        'Obras em andamento': 'obras_em_andamento',
    }

    financeiro_map = {
        'Receita Operacional Bruta': 'receita_bruta',
        'Receita Operacional Líquida': 'receita_liquida',
        'Receita Líquida': 'receita_liquida',
        'Lucro Bruto Ajustado': 'lucro_bruto_ajustado',
        'Margem Bruta Ajustada (%)': 'margem_bruta_ajustada',
        'Lucro Bruto': 'lucro_bruto',
        'Margem Bruta': 'margem_bruta',
        'Despesas com Vendas': 'despesas_comerciais',
        'Despesas Gerais e Administrativas': 'despesas_ga',
        'Total de despesas SG&A': 'total_despesas_operacionais',
        'Outras Receitas e Despesas': 'outras_receitas_despesas_op',
        'Equivalência Patrimonial': 'equivalencia_patrimonial',
        'Receitas Financeiras': 'receitas_financeiras',
        'Despesas Financeiras': 'despesas_financeiras',
        'Resultado Financeiro': 'resultado_financeiro',
        'Lucro Líquido (Prejuízo)': 'lucro_liquido',
        'Lucro Líquido': 'lucro_liquido',
        'Margem Líquida': 'margem_liquida',
        '*EBITDA Ajustado': 'ebitda_ajustado',
        'Margem EBITDA ajustada': 'margem_ebitda_ajustada',
        'Receitas a Apropriar': 'receitas_apropriar',
        'Custo das Unidades Vendidas a Apropriar': 'custo_apropriar',
        'Resultado a Apropriar': 'resultado_apropriar',
        'Margem a Apropriar': 'margem_apropriar',
        'ROE': 'roe',
        'ROCE': 'roce',
        'NOPAT': 'nopat',
        'Capital Empregado': 'capital_empregado',
        'Dívida Bruta': 'divida_bruta',
        'Dívida Líquida': 'divida_liquida',
        'Patrimônio Líquido + Minoritários': 'patrimonio_liquido',
        'Dívida Líquida / (PL': 'divida_liquida_pl',
        'Caixa e Disponibilidades': 'caixa_aplicacoes',
    }

    segments = {
        'Tenda': ('Tenda - Operacional', 'Tenda - Tenda Financeiro'),
        'Alea': ('Alea - Operacional', 'Alea - Financeiro'),
        'Consolidado': ('Consolidado - Operacional', 'Consolidado - Financeiro'),
    }

    for seg, (aba_op, aba_fin) in segments.items():
        dados = extrair_dados_por_label(fp, aba_op, 7, 0, operacional_map)
        n, u = inserir_dados(conn, 'Tenda', seg, dados)
        print(f"    {seg} Operacional: {n} novos, {u} atualizados")

        dados = extrair_dados_por_label(fp, aba_fin, 7, 0, financeiro_map)
        n, u = inserir_dados(conn, 'Tenda', seg, dados)
        print(f"    {seg} Financeiro: {n} novos, {u} atualizados")


# ============================================================
# CURY
# ============================================================
def ingerir_cury(conn):
    print("\n  --- CURY ---")
    fp = PLANILHAS_DIR / "Cury_Planilha_Fundamentos_2026-03.xlsx"

    op_map = {
        'Número de Empreendimentos': 'empreendimentos_lancados',
        'VGV (em R$ mil)': 'vgv_lancado',
        'Número de unidades': 'unidades_lancadas',
        'Preço médio por unidade': 'preco_medio_lancamento',
        'Tamanho médio dos lançamentos': 'tamanho_medio_empreendimentos',
        'Vendas Brutas': 'vendas_brutas_vgv',
        'VSO Bruto': 'vso_bruta_trimestral',
        'Vendas Líquidas': 'vendas_liquidas_vgv',
        'Distratos': 'distratos_vgv',
        'VSO Líquido': 'vso_liquida_trimestral',
        'VSO UDM': 'vso_liquida_12m',
        'Unidades Vendidas Brutas': 'vendas_brutas_unidades',
        'Unidades Distratadas': 'distratos_unidades',
        'Unidades Vendidas Líquidas': 'vendas_liquidas_unidades',
        'VGV Repassado': 'vgv_repassado',
        'Unidades Repassadas': 'unidades_repassadas',
        'Unidades concluídas': 'unidades_entregues',
        'Obras em andamento': 'obras_em_andamento',
        'Duração do Estoque': 'giro_estoque_meses',
    }
    dados = extrair_dados_por_label(fp, 'Resultados Operacionais', 5, 1, op_map, multiplicador=0.001)
    n, u = inserir_dados(conn, 'Cury', 'Consolidado', dados)
    print(f"    Operacional: {n} novos, {u} atualizados")

    # Landbank (row 74) and Geracao de Caixa (row 90)
    dados = extrair_por_linha(fp, 'Resultados Operacionais', 5, {74: 'landbank_vgv'}, multiplicador=1.0)
    dados2 = extrair_por_linha(fp, 'Resultados Operacionais', 5, {90: 'geracao_caixa'}, multiplicador=0.001)
    for p, campos in dados2.items():
        if p not in dados: dados[p] = {}
        dados[p].update(campos)
    n, u = inserir_dados(conn, 'Cury', 'Consolidado', dados)
    print(f"    Landbank+Caixa: {n} novos, {u} atualizados")

    dre_map = {
        'Receita líquida': 'receita_liquida',
        'Custo das vendas': 'custo_imoveis_vendidos',
        'Lucro bruto': 'lucro_bruto',
        'Margem bruta': 'margem_bruta',
        'Margem Bruta Ajustada': 'margem_bruta_ajustada',
        'Despesas com vendas': 'despesas_comerciais',
        'Despesas gerais e administrativas': 'despesas_ga',
        'Outros resultados operacionais': 'outras_receitas_despesas_op',
        'Equivalência patrimonial': 'equivalencia_patrimonial',
        'Receitas (despesas) operacionais': 'total_despesas_operacionais',
        'Despesas financeiras': 'despesas_financeiras',
        'Receitas financeiras': 'receitas_financeiras',
        'Resultado financeiro': 'resultado_financeiro',
        'Lucro líquido do exercício': 'lucro_liquido',
        'Margem líquida': 'margem_liquida',
        'Ebitda': 'ebitda',
        'Margem Ebitda': 'margem_ebitda',
        'Ebitda Ajustado': 'ebitda_ajustado',
        'Margem Ebitda Ajustada': 'margem_ebitda_ajustada',
        'ROE': 'roe',
        'Depreciação e amortização': 'depreciacao_amortizacao',
        'Imposto de renda e contribuição social': 'ir_csll',
    }
    dados = extrair_dados_por_label(fp, 'Demonstrações do Resultado', 3, 1, dre_map, multiplicador=0.001)
    n, u = inserir_dados(conn, 'Cury', 'Consolidado', dados)
    print(f"    DRE: {n} novos, {u} atualizados")

    bp_ativo_map = {
        'Caixa e equivalentes': 'caixa_aplicacoes',
        'Total do ativo': 'ativo_total',
        'Total do ativo circulante': 'ativo_circulante',
    }
    dados = extrair_dados_por_label(fp, 'Balanço Patrimonial Ativo', 5, 1, bp_ativo_map, multiplicador=0.001)
    n, u = inserir_dados(conn, 'Cury', 'Consolidado', dados)
    print(f"    BP Ativo: {n} novos, {u} atualizados")

    bp_passivo_map = {
        'Patrimônio Líquido total': 'patrimonio_liquido',
        'Total current liabilities': 'passivo_circulante',
        'Total do passivo circulante': 'passivo_circulante',
    }
    dados = extrair_dados_por_label(fp, 'Balanço Patrimonial Passivo', 4, 1, bp_passivo_map, multiplicador=0.001)
    n, u = inserir_dados(conn, 'Cury', 'Consolidado', dados)
    print(f"    BP Passivo: {n} novos, {u} atualizados")


# ============================================================
# DIRECIONAL
# ============================================================
def ingerir_direcional(conn):
    print("\n  --- DIRECIONAL ---")
    fp = PLANILHAS_DIR / "Direcional_Planilha_Interativa_2026-03.xlsx"

    op_map = {
        'VGV Lançado (R$ mil)': 'vgv_lancado',
        'Unidades Lançadas (unidades)': 'unidades_lancadas',
        'Preço Médio (R$/unidade)': 'preco_medio_lancamento',
        'VGV Contratado (R$ mil)': 'vendas_brutas_vgv',
        'Unidades Contratadas (unidades)': 'vendas_brutas_unidades',
        'VSO (Vendas Sobre Oferta) - Consolidada': 'vso_liquida_trimestral',
        'Estoque Total a Valor de Mercado (R$ mil)': 'estoque_vgv',
        'Unidades em Estoque': 'estoque_unidades',
        'Land Bank Total (R$ mil)': 'landbank_vgv',
        'Land Bank (unidades)': 'landbank_unidades',
        'Repasses (R$ mil)': 'vgv_repassado',
    }
    dados = extrair_dados_por_label(fp, 'Dados Operacionais', 4, 1, op_map, multiplicador=0.001)
    n, u = inserir_dados(conn, 'Direcional', 'Consolidado', dados)
    print(f"    Operacional: {n} novos, {u} atualizados")

    dre_map = {
        'RECEITA OPERACIONAL': 'receita_liquida',
        'Custo da Venda': 'custo_imoveis_vendidos',
        'LUCRO BRUTO': 'lucro_bruto',
        'Despesas Gerais e Ad': 'despesas_ga',
        'Despesas Comerciais': 'despesas_comerciais',
        'Resultado com Equivalência': 'equivalencia_patrimonial',
        'Outras Receitas e Despesas': 'outras_receitas_despesas_op',
        'Despesas Financeiras': 'despesas_financeiras',
        'Receitas Financeiras': 'receitas_financeiras',
        'LUCRO LÍQUIDO DO PERÍODO': 'lucro_liquido',
        'Depreciação e Amortização': 'depreciacao_amortizacao',
        'IR/CSLL': 'ir_csll',
        'Imposto de Renda e Contribuição Social': 'ir_csll',
    }
    dados = extrair_dados_por_label(fp, 'Demonstração de Resultados', 4, 1, dre_map, multiplicador=0.001)
    n, u = inserir_dados(conn, 'Direcional', 'Consolidado', dados)
    print(f"    DRE: {n} novos, {u} atualizados")

    bp_map = {
        'Caixa e equivalentes de caixa': 'caixa_aplicacoes',
        'PATRIMÔNIO LÍQUIDO': 'patrimonio_liquido',
        'Total do Ativo': 'ativo_total',
        'TOTAL DO ATIVO': 'ativo_total',
        'Passivo Circulante': 'passivo_circulante',
        'PASSIVO CIRCULANTE': 'passivo_circulante',
    }
    dados = extrair_dados_por_label(fp, 'Balanço Patrimonial', 4, 1, bp_map, multiplicador=0.001)
    n, u = inserir_dados(conn, 'Direcional', 'Consolidado', dados)
    print(f"    Balanço: {n} novos, {u} atualizados")

    # Riva DRE
    riva_map = {
        'RECEITA OPERACIONAL': 'receita_liquida',
        'Custo da Venda': 'custo_imoveis_vendidos',
        'LUCRO BRUTO': 'lucro_bruto',
        'Despesas comerciais, gerais': 'total_despesas_operacionais',
        'Resultado Financeiro': 'resultado_financeiro',
        'LUCRO LÍQUIDO': 'lucro_liquido',
    }
    try:
        dados = extrair_dados_por_label(fp, 'RIVA - DRE', 4, 1, riva_map, multiplicador=0.001)
        n, u = inserir_dados(conn, 'Direcional', 'Riva', dados)
        print(f"    Riva DRE: {n} novos, {u} atualizados")
    except Exception as e:
        print(f"    Riva: ERRO - {e}")


# ============================================================
# PLANO&PLANO
# ============================================================
def ingerir_planoeplano(conn):
    print("\n  --- PLANO&PLANO ---")
    fp = PLANILHAS_DIR / "PlanoePlano_Planilha_Interativa_2026-03.xlsx"

    op_map = {
        'Vendas Contratadas Brutas (R$ mil)': 'vendas_brutas_vgv',
        'Vendas Contratadas Brutas (Unidades)': 'vendas_brutas_unidades',
        'Distratos (R$ mil)': 'distratos_vgv',
        'Distratos (Unidades)': 'distratos_unidades',
        'Vendas Líquidas 100% Plano&Plano (R$ mil)': 'vendas_liquidas_vgv',
        'Vendas Líquidas 100% Plano&Plano (Unid': 'vendas_liquidas_unidades',
        'Preço Venda Médio': 'preco_medio_vendas',
        'VGV 100% Plano&Plano (R$ mil)': 'vgv_lancado',
        'Unidades': 'unidades_lancadas',
        'Estoque VGV 100%': 'estoque_vgv',
        'Estoque (Unidades)': 'estoque_unidades',
        '% Pronto (unidades)': 'pct_estoque_pronto',
        'VSO Líquido 100% - últimos 12 meses': 'vso_liquida_12m',
    }
    dados = extrair_dados_por_label(fp, 'Dados Operacionais', 7, 1, op_map, multiplicador=0.001)
    n, u = inserir_dados(conn, 'PlanoePlano', 'Consolidado', dados)
    print(f"    Operacional: {n} novos, {u} atualizados")

    dre_map = {
        'Receita líquida': 'receita_liquida',
        'Custos dos imóveis vendidos': 'custo_imoveis_vendidos',
        'Lucro bruto': 'lucro_bruto',
        'Margem bruta': 'margem_bruta',
        'Despesas comerciais': 'despesas_comerciais',
        'Despesas gerais e administrativas': 'despesas_ga',
        'Resultado da equivalência': 'equivalencia_patrimonial',
        'Outras receitas (despesas)': 'outras_receitas_despesas_op',
        'Despesas (receitas) operacionais': 'total_despesas_operacionais',
        'Receitas financeiras': 'receitas_financeiras',
        'Despesas financeiras': 'despesas_financeiras',
        'Receitas (despesas) financeiras': 'resultado_financeiro',
        'Lucro líquido 100%': 'lucro_liquido',
        'Margem Líquida 100%': 'margem_liquida',
        'Depreciação e amortização': 'depreciacao_amortizacao',
        'Imposto de renda e contribuição social': 'ir_csll',
    }
    dados = extrair_dados_por_label(fp, 'DRE Consolidado', 7, 1, dre_map, multiplicador=0.001)
    n, u = inserir_dados(conn, 'PlanoePlano', 'Consolidado', dados)
    print(f"    DRE: {n} novos, {u} atualizados")

    bp_map = {
        'Caixa e equivalentes de caixa': 'caixa_aplicacoes',
        'Total do patrimônio líquido': 'patrimonio_liquido',
        'Total do ativo': 'ativo_total',
        'Total do passivo circulante': 'passivo_circulante',
    }
    try:
        dados = extrair_dados_por_label(fp, 'Balanço Patrimonial Consolidado', 7, 1, bp_map, multiplicador=0.001)
        n, u = inserir_dados(conn, 'PlanoePlano', 'Consolidado', dados)
        print(f"    Balanço: {n} novos, {u} atualizados")
    except Exception as e:
        print(f"    Balanço: ERRO - {e}")


# ============================================================
# MRV
# ============================================================
def ingerir_mrv(conn):
    print("\n  --- MRV ---")
    fp = PLANILHAS_DIR / "MRV_Base_Dados_Operacionais_Financeiros_2026-03.xlsx"

    dre_map = {
        'Receita Operacional Líquida': 'receita_liquida',
        'Custo dos Imóveis Vendidos': 'custo_imoveis_vendidos',
        'Lucro Bruto': 'lucro_bruto',
        'Margem Bruta (%)': 'margem_bruta',
        'Despesas Comerciais': 'despesas_comerciais',
        'Despesas Gerais e Administrativas': 'despesas_ga',
        'Outras receitas (despesas)': 'outras_receitas_despesas_op',
        'Resultado de Equivalência': 'equivalencia_patrimonial',
        'Resultado Financeiro': 'resultado_financeiro',
        'Despesas financeiras': 'despesas_financeiras',
        'Receitas financeiras': 'receitas_financeiras',
        'Lucro Líquido do Período': 'lucro_liquido',
        'Lucro Líquido atribuível aos Acionistas': 'lucro_liquido',
        'Margem Líquida (%)': 'margem_liquida',
        'Depreciação e Amortização': 'depreciacao_amortizacao',
        'Depreciation and Amortization': 'depreciacao_amortizacao',
        'Imposto de Renda e Contribuição Social': 'ir_csll',
    }

    dados = extrair_dados_por_label(fp, 'DRE Consolid. | Income Statem.', 2, 0, dre_map, multiplicador=0.001)
    n, u = inserir_dados(conn, 'MRV', 'Consolidado', dados)
    print(f"    DRE Consolidado: {n} novos, {u} atualizados")

    fin_map = {
        'Margem Bruta ex. juros': 'margem_bruta_ajustada',
        'EBITDA': 'ebitda',
        'EBITDA Margin': 'margem_ebitda',
        'ROE (12 meses)': 'roe',
        'Dívida Total': 'divida_bruta',
        '(-) Caixa e Equivalentes': 'caixa_aplicacoes',
        'Dívida Líquida': 'divida_liquida',
        'Total do Patrimônio Líquido': 'patrimonio_liquido',
        'Dívida Líquida / PL': 'divida_liquida_pl',
        'Geração de Caixa': 'geracao_caixa',
    }
    dados = extrair_dados_por_label(fp, 'Indic.Fin. | Financ.Highlights', 2, 0, fin_map, multiplicador=0.001)
    n, u = inserir_dados(conn, 'MRV', 'Consolidado', dados)
    print(f"    Indicadores: {n} novos, {u} atualizados")

    # BP Consolidado - ativo total, passivo circulante
    bp_map = {
        'Caixa e Equivalentes': 'caixa_aplicacoes',
        'Total do Patrimônio Líquido': 'patrimonio_liquido',
        'Total Assets': 'ativo_total',
        'Ativo Total': 'ativo_total',
        'Total do Ativo': 'ativo_total',
        'Passivo Circulante': 'passivo_circulante',
        'Current Liabilities': 'passivo_circulante',
        'Total Current Liabilities': 'passivo_circulante',
    }
    try:
        dados = extrair_dados_por_label(fp, 'BP Consolid. | Consolid. BS', 2, 0, bp_map, multiplicador=0.001)
        n, u = inserir_dados(conn, 'MRV', 'Consolidado', dados)
        print(f"    BP Consolidado: {n} novos, {u} atualizados")
    except Exception as e:
        print(f"    BP Consolidado: ERRO - {e}")

    # Segmentos DRE
    for seg, aba in [('MRV Incorporação', 'DRE MRV Inc. | Income Statem.'),
                      ('Resia', 'DRE Resia | Income Statem.'),
                      ('Urba', 'DRE Urba | Income Statem.'),
                      ('Luggo', 'DRE Luggo | Income Statem.')]:
        try:
            dados = extrair_dados_por_label(fp, aba, 2, 0, dre_map, multiplicador=0.001)
            n, u = inserir_dados(conn, 'MRV', seg, dados)
            print(f"    DRE {seg}: {n} novos, {u} atualizados")
        except Exception as e:
            print(f"    DRE {seg}: ERRO - {e}")

    # Dados Operacionais (row-based)
    aba_op = 'Dados Oper. MRV&Co | Oper.Data'
    try:
        dados = extrair_por_linha(fp, aba_op, 2, {82: 'vgv_lancado', 83: 'unidades_lancadas'})
        n, u = inserir_dados(conn, 'MRV', 'Consolidado', dados)
        print(f"    Lançamentos: {n} novos, {u} atualizados")

        dados = extrair_por_linha(fp, aba_op, 2, {207: 'vendas_liquidas_vgv', 208: 'vendas_liquidas_unidades'})
        n, u = inserir_dados(conn, 'MRV', 'Consolidado', dados)
        print(f"    Vendas Líq: {n} novos, {u} atualizados")
    except Exception as e:
        print(f"    Operacional: ERRO - {e}")


# ============================================================
# CYRELA
# ============================================================
def ingerir_cyrela(conn):
    print("\n  --- CYRELA ---")
    fp_df = PLANILHAS_DIR / "Cyrela_Demonstracoes_Financeiras_2026-03.xlsx"

    dre_map = {
        'Receita liquida operac': 'receita_liquida',
        'Custo Produtos Vendidos': 'custo_imoveis_vendidos',
        'Lucro Bruto': 'lucro_bruto',
        'Despesas com Vendas': 'despesas_comerciais',
        'Despesas administrativ': 'despesas_ga',
        'Equivalenc patrimonial': 'equivalencia_patrimonial',
        'Outras rec operacionais': 'outras_receitas_despesas_op',
        'Lucro antes jur&imp': 'ebitda',
        'Resultado financeiro': 'resultado_financeiro',
        'Receitas Financeiras': 'receitas_financeiras',
        'Despesas Financeiras': 'despesas_financeiras',
        'Lucro liquido': 'lucro_liquido',
        'Lucro Consolidado': 'lucro_liquido',
        'Depreciac amortiz depl': 'depreciacao_amortizacao',
        'Impostos s/ lucro': 'ir_csll',
    }
    dados = extrair_dados_por_label(fp_df, 'CYRE3', 4, 0, dre_map, multiplicador=0.001)
    n, u = inserir_dados(conn, 'Cyrela', 'Consolidado', dados)
    print(f"    DRE: {n} novos, {u} atualizados")

    bp_map = {
        'Caixa e equivalentes de caixa': 'caixa_aplicacoes',
        'Patrimonio liquido': 'patrimonio_liquido',
        'Patrim liq consolidado': 'patrimonio_liquido',
        'Ativo Total': 'ativo_total',
        'Ativo total': 'ativo_total',
        'Passivo Circulante': 'passivo_circulante',
        'Passivo circulante': 'passivo_circulante',
    }
    dados = extrair_dados_por_label(fp_df, 'CYRE3', 4, 0, bp_map, multiplicador=0.001)
    n, u = inserir_dados(conn, 'Cyrela', 'Consolidado', dados)
    print(f"    Balanço: {n} novos, {u} atualizados")

    # Operacional
    fp_op = PLANILHAS_DIR / "Cyrela_Dados_Operacionais_2026-03.xlsx"
    try:
        dados = extrair_por_linha(fp_op, 'Lçtos', 4, {14: 'vgv_lancado', 71: 'unidades_lancadas'}, multiplicador=0.001)
        n, u = inserir_dados(conn, 'Cyrela', 'Consolidado', dados)
        print(f"    Lançamentos: {n} novos, {u} atualizados")
    except Exception as e:
        print(f"    Lançamentos: ERRO - {e}")

    try:
        dados = extrair_por_linha(fp_op, 'Vendas', 4, {14: 'vendas_liquidas_vgv', 37: 'vendas_liquidas_unidades'}, multiplicador=0.001)
        n, u = inserir_dados(conn, 'Cyrela', 'Consolidado', dados)
        print(f"    Vendas: {n} novos, {u} atualizados")
    except Exception as e:
        print(f"    Vendas: ERRO - {e}")


# ============================================================
# MAIN
# ============================================================
if __name__ == '__main__':
    print("=" * 60)
    print("POPULANDO DADOS 2025 + CAMPOS ROCE/ROIC")
    print("=" * 60)

    conn = sqlite3.connect(DB_PATH)

    ingerir_tenda(conn)
    ingerir_cury(conn)
    ingerir_direcional(conn)
    ingerir_planoeplano(conn)
    ingerir_mrv(conn)
    ingerir_cyrela(conn)

    # Calcular EBIT = EBITDA - D&A (quando ambos disponíveis)
    cur = conn.cursor()
    cur.execute("""
        UPDATE dados_trimestrais
        SET ebit = ebitda - COALESCE(depreciacao_amortizacao, 0)
        WHERE ebitda IS NOT NULL AND ebit IS NULL
    """)
    print(f"\n  EBIT calculado: {cur.rowcount} registros")

    conn.commit()

    # Verificar resultado
    cur.execute("SELECT empresa, MIN(periodo), MAX(periodo), COUNT(*) FROM dados_trimestrais GROUP BY empresa ORDER BY empresa")
    print("\n" + "=" * 60)
    print("RESULTADO FINAL")
    print("=" * 60)
    for row in cur.fetchall():
        print(f"  {row[0]}: {row[1]} a {row[2]} ({row[3]} registros)")

    # Verificar campos ROCE/ROIC
    print("\nCampos ROCE/ROIC:")
    for campo in ['ativo_total', 'passivo_circulante', 'depreciacao_amortizacao', 'ir_csll', 'ebit', 'nopat', 'roce', 'capital_empregado']:
        cur.execute(f"SELECT COUNT(*) FROM dados_trimestrais WHERE {campo} IS NOT NULL")
        n = cur.fetchone()[0]
        print(f"  {campo}: {n} registros preenchidos")

    conn.close()
    print("\nConcluído!")

# -*- coding: utf-8 -*-
"""
Cria banco SQLite e popula com dados das planilhas XLSX.
Versão 2: usa matching por label (nome da linha) em vez de número de linha.
Escopo: 1T2020 a 3T2025, empresas: Tenda, MRV, Direcional, Cury, PlanoePlano, Cyrela
"""

import sqlite3
import openpyxl
import re
import os
from pathlib import Path
from datetime import datetime

DB_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "dados_financeiros.db")
PLANILHAS_DIR = Path(os.path.dirname(os.path.abspath(__file__))) / "planilhas"

# Fields that are percentages/ratios — should NOT be multiplied by multiplicador
CAMPOS_PERCENTUAIS = {
    'margem_bruta', 'margem_bruta_ajustada', 'margem_ebitda', 'margem_ebitda_ajustada',
    'margem_liquida', 'margem_apropriar',
    'vso_bruta_trimestral', 'vso_liquida_trimestral', 'vso_bruta_12m', 'vso_liquida_12m',
    'pct_comerciais_receita_liquida', 'pct_ga_receita_liquida', 'pct_sga_receita_liquida',
    'pct_estoque_pronto', 'pct_permuta_total',
    'roe', 'divida_liquida_pl', 'giro_estoque_meses',
    'pdd_cobertura_pct', 'inadimplencia_total_pct',
    'aging_adimplente_pct', 'aging_vencido_90d_pct', 'aging_vencido_360d_pct',
    'aging_vencido_360d_mais_pct', 'carteira_pct_pl', 'carteira_pos_chaves_pct',
}


def normalizar_periodo(val):
    """Normaliza header de período para formato xTyyyy."""
    if val is None:
        return None

    if isinstance(val, datetime):
        month_to_tri = {1: 1, 2: 1, 3: 1, 4: 2, 5: 2, 6: 2, 7: 3, 8: 3, 9: 3, 10: 4, 11: 4, 12: 4}
        tri = month_to_tri[val.month]
        return f"{tri}T{val.year}"

    s = str(val).strip()

    # "3T25/3Q25" bilingual
    m = re.match(r'^(\d)[Tt](\d{2,4})\s*/\s*\d[Qq]\d', s)
    if m:
        tri, ano = m.group(1), m.group(2)
        if len(ano) == 2: ano = f"20{ano}"
        return f"{tri}T{ano}"

    # "3T25" or "3T2025"
    m = re.match(r'^(\d)[TtQq](\d{2,4})$', s)
    if m:
        tri, ano = m.group(1), m.group(2)
        if len(ano) == 2: ano = f"20{ano}"
        return f"{tri}T{ano}"

    # Date: "2025-09-30"
    m = re.match(r'^(\d{4})-(\d{2})-\d{2}', s)
    if m:
        ano, mes = int(m.group(1)), int(m.group(2))
        month_to_tri = {3: 1, 6: 2, 9: 3, 12: 4}
        tri = month_to_tri.get(mes)
        if tri:
            return f"{tri}T{ano}"

    return None


def periodo_valido(p):
    if not p: return False
    m = re.match(r'^(\d)T(\d{4})$', p)
    if not m: return False
    tri, ano = int(m.group(1)), int(m.group(2))
    if ano < 2020 or ano > 2026: return False
    return True


def criar_schema(conn):
    c = conn.cursor()
    c.execute("DROP TABLE IF EXISTS dados_trimestrais")
    c.execute("DROP TABLE IF EXISTS empresas")
    c.execute("DROP TABLE IF EXISTS log_ingestao")

    c.execute("""
    CREATE TABLE empresas (
        id INTEGER PRIMARY KEY, nome TEXT UNIQUE NOT NULL, sigla TEXT, segmentos TEXT
    )""")

    c.execute("""
    CREATE TABLE dados_trimestrais (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        empresa TEXT NOT NULL, segmento TEXT DEFAULT 'Consolidado',
        periodo TEXT NOT NULL, ano INTEGER NOT NULL, trimestre INTEGER NOT NULL,
        fonte TEXT DEFAULT 'planilha',

        -- Lançamentos
        vgv_lancado REAL, empreendimentos_lancados REAL, unidades_lancadas REAL,
        preco_medio_lancamento REAL, tamanho_medio_empreendimentos REAL,
        -- Vendas
        vendas_brutas_vgv REAL, vendas_brutas_unidades REAL,
        distratos_vgv REAL, distratos_unidades REAL,
        vendas_liquidas_vgv REAL, vendas_liquidas_unidades REAL,
        vso_bruta_trimestral REAL, vso_liquida_trimestral REAL,
        vso_bruta_12m REAL, vso_liquida_12m REAL, preco_medio_vendas REAL,
        -- Repasses/Entregas
        vgv_repassado REAL, unidades_repassadas REAL,
        unidades_entregues REAL, obras_em_andamento REAL,
        -- Estoque
        estoque_vgv REAL, estoque_unidades REAL,
        preco_medio_estoque REAL, pct_estoque_pronto REAL, giro_estoque_meses REAL,
        -- Landbank
        landbank_vgv REAL, landbank_empreendimentos REAL,
        landbank_unidades REAL, landbank_preco_medio REAL, pct_permuta_total REAL,
        -- Backlog
        receitas_apropriar REAL, custo_apropriar REAL,
        resultado_apropriar REAL, margem_apropriar REAL,
        -- DRE
        receita_bruta REAL, receita_liquida REAL, custo_imoveis_vendidos REAL,
        lucro_bruto REAL, margem_bruta REAL,
        lucro_bruto_ajustado REAL, margem_bruta_ajustada REAL,
        ebitda REAL, ebitda_ajustado REAL, margem_ebitda REAL, margem_ebitda_ajustada REAL,
        resultado_financeiro REAL, receitas_financeiras REAL, despesas_financeiras REAL,
        lucro_liquido REAL, margem_liquida REAL, roe REAL,
        -- SG&A
        despesas_comerciais REAL, despesas_ga REAL, honorarios_administracao REAL,
        outras_receitas_despesas_op REAL, equivalencia_patrimonial REAL,
        total_despesas_operacionais REAL,
        pct_comerciais_receita_liquida REAL, pct_ga_receita_liquida REAL,
        pct_sga_receita_liquida REAL,
        -- Endividamento
        divida_bruta REAL, caixa_aplicacoes REAL, divida_liquida REAL,
        divida_liquida_pl REAL, patrimonio_liquido REAL,
        -- Geração de Caixa
        geracao_caixa REAL,
        -- Recebíveis
        carteira_recebiveis_total REAL, carteira_pre_chaves REAL,
        carteira_pos_chaves REAL, pdd_provisao REAL,
        pdd_cobertura_pct REAL, inadimplencia_total_pct REAL,
        aging_adimplente_pct REAL, aging_vencido_90d_pct REAL,
        aging_vencido_360d_pct REAL, aging_vencido_360d_mais_pct REAL,
        cessao_recebiveis_trimestre REAL, saldo_cessao_recebiveis REAL,
        carteira_pct_pl REAL, carteira_pos_chaves_pct REAL,

        UNIQUE(empresa, segmento, periodo)
    )""")

    c.execute("""CREATE TABLE log_ingestao (
        id INTEGER PRIMARY KEY AUTOINCREMENT, empresa TEXT, planilha TEXT,
        aba TEXT, registros_novos INTEGER, registros_atualizados INTEGER,
        timestamp TEXT DEFAULT CURRENT_TIMESTAMP
    )""")

    for nome, sigla, segs in [
        ('Tenda', 'TEND3', 'Tenda, Alea, Consolidado'),
        ('MRV', 'MRVE3', 'MRV Incorporação, Resia, Urba, Luggo, Consolidado'),
        ('Direcional', 'DIRR3', 'Direcional, Riva, Consolidado'),
        ('Cury', 'CURY3', 'Consolidado'),
        ('PlanoePlano', 'PLPL3', 'Consolidado'),
        ('Cyrela', 'CYRE3', 'Cyrela, Vivaz, Consolidado'),
    ]:
        c.execute("INSERT INTO empresas (nome, sigla, segmentos) VALUES (?,?,?)", (nome, sigla, segs))
    conn.commit()


def extrair_por_linha(filepath, aba, header_row, row_map, multiplicador=1.0):
    """
    Extrai dados por número de linha (1-indexed).
    row_map: {row_number: campo_db}
    Útil para planilhas com labels repetidos em seções diferentes (MRV, Cyrela).
    """
    wb = openpyxl.load_workbook(str(filepath), read_only=True, data_only=True)
    ws = wb[aba]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if len(rows) < header_row:
        return {}

    header = rows[header_row - 1]
    col_periods = {}
    for col_idx, val in enumerate(header):
        p = normalizar_periodo(val)
        if p and periodo_valido(p):
            col_periods[col_idx] = p

    if not col_periods:
        return {}

    result = {}
    for row_num, campo in row_map.items():
        if row_num - 1 >= len(rows):
            continue
        row = rows[row_num - 1]
        for col_idx, periodo in col_periods.items():
            if col_idx < len(row) and row[col_idx] is not None:
                try:
                    val = float(row[col_idx])
                    if campo not in CAMPOS_PERCENTUAIS:
                        val *= multiplicador
                    if periodo not in result:
                        result[periodo] = {}
                    result[periodo][campo] = val
                except (ValueError, TypeError):
                    pass
    return result


def extrair_dados_por_label(filepath, aba, header_row, label_col, field_map, multiplicador=1.0):
    """
    Extrai dados de planilha usando matching por label.
    field_map: {label_pattern: campo_db} — match case-insensitive no início da string
    label_col: 0-indexed column for labels
    multiplicador: multiply values (use 0.001 for R$ mil -> R$ milhões)
    Returns: {periodo: {campo: valor}}
    """
    wb = openpyxl.load_workbook(str(filepath), read_only=True, data_only=True)
    ws = wb[aba]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if len(rows) < header_row:
        return {}

    # Build period map from header row
    header = rows[header_row - 1]
    col_periods = {}
    for col_idx, val in enumerate(header):
        p = normalizar_periodo(val)
        if p and periodo_valido(p):
            col_periods[col_idx] = p

    if not col_periods:
        return {}

    result = {}

    for row in rows:
        label_val = row[label_col] if label_col < len(row) else None
        if label_val is None:
            continue
        label_str = str(label_val).strip()
        # Strip accounting prefix chars (+, -, =) used by Cyrela/Economatica
        label_str = label_str.lstrip('+-=').strip().lower()
        if not label_str:
            continue

        # Find matching field — prefer longest pattern first to avoid
        # "Despesas com Vendas" matching before "Despesas com Vendas / Vendas"
        campo = None
        sorted_patterns = sorted(field_map.items(), key=lambda x: len(x[0]), reverse=True)
        for pattern, campo_name in sorted_patterns:
            pattern_lower = pattern.lower()
            if label_str.startswith(pattern_lower) or label_str == pattern_lower:
                campo = campo_name
                break

        if not campo:
            continue

        for col_idx, periodo in col_periods.items():
            if col_idx < len(row) and row[col_idx] is not None:
                try:
                    val = float(row[col_idx])
                    if campo not in CAMPOS_PERCENTUAIS:
                        val *= multiplicador
                    if periodo not in result:
                        result[periodo] = {}
                    result[periodo][campo] = val
                except (ValueError, TypeError):
                    pass

    return result


def inserir_dados(conn, empresa, segmento, dados_por_periodo):
    """Insere/atualiza dados no banco. Retorna (novos, atualizados)."""
    c = conn.cursor()
    novos = 0
    atualizados = 0

    for periodo, campos in dados_por_periodo.items():
        m = re.match(r'(\d)T(\d{4})', periodo)
        if not m: continue
        tri, ano = int(m.group(1)), int(m.group(2))

        c.execute("SELECT id FROM dados_trimestrais WHERE empresa=? AND segmento=? AND periodo=?",
                  (empresa, segmento, periodo))
        existing = c.fetchone()

        if existing:
            for campo, valor in campos.items():
                try:
                    c.execute(f"UPDATE dados_trimestrais SET {campo}=? WHERE id=?", (valor, existing[0]))
                    atualizados += 1
                except Exception:
                    pass
        else:
            campo_names = ['empresa', 'segmento', 'periodo', 'ano', 'trimestre', 'fonte'] + list(campos.keys())
            campo_vals = [empresa, segmento, periodo, ano, tri, 'planilha'] + list(campos.values())
            placeholders = ','.join(['?'] * len(campo_vals))
            col_str = ','.join(campo_names)
            try:
                c.execute(f"INSERT INTO dados_trimestrais ({col_str}) VALUES ({placeholders})", campo_vals)
                novos += 1
            except Exception:
                pass

    conn.commit()
    return novos, atualizados


# ============================================================
# TENDA
# ============================================================
def ingerir_tenda(conn):
    print("\n  --- TENDA ---")
    fp = PLANILHAS_DIR / "Tenda_Planilha_Fundamentos_2026-03.xlsx"

    # Field maps by label pattern (case-insensitive startsWith)
    operacional_map = {
        'Número de Empreendimentos': 'empreendimentos_lancados',
        'VGV (R$ milhões)': 'vgv_lancado',
        'Número de unidades': 'unidades_lancadas',
        'Preço médio por unidade (R$ mil)': 'preco_medio_lancamento',
        'Tamanho médio dos lançamentos': 'tamanho_medio_empreendimentos',
        'Vendas Brutas': 'vendas_brutas_vgv',
        'VSO Bruto': 'vso_bruta_trimestral',
        'Distratos': 'distratos_vgv',
        'Vendas Líquidas': 'vendas_liquidas_vgv',
        'VSO Líquido': 'vso_liquida_trimestral',
        'Unidades Vendidas Brutas': 'vendas_brutas_unidades',
        'Unidades Distratadas': 'distratos_unidades',
        'Unidades Vendidas Líquidas': 'vendas_liquidas_unidades',
        'Preço médio por unidade bruta': 'preco_medio_vendas',
        'VGV Repassado': 'vgv_repassado',
        'Unidades Repassadas': 'unidades_repassadas',
        'Unidades Entregues': 'unidades_entregues',
        'Obras em andamento': 'obras_em_andamento',
    }

    # Estoque and landbank use same label pattern "VGV" so we need a separate map
    estoque_landbank_map = {
        'Número de Empreendimentos': 'landbank_empreendimentos',
        'VGV (em R$ milhões)': 'landbank_vgv',
        'Aquisições/Ajustes': 'landbank_vgv',  # fallback
        'Número de unidades': 'landbank_unidades',
        'Preço médio por unidade (em R$ mil)': 'landbank_preco_medio',
        '% Permuta Total': 'pct_permuta_total',
    }

    financeiro_map = {
        'Receita Operacional Bruta': 'receita_bruta',
        'Receita Operacional Líquida': 'receita_liquida',
        'Receita Líquida': 'receita_liquida',
        'Lucro Bruto Ajustado': 'lucro_bruto_ajustado',
        'Margem Bruta Ajustada (%)': 'margem_bruta_ajustada',
        'Lucro Bruto': 'lucro_bruto',
        'Margem Bruta': 'margem_bruta',
        'Despesas com Vendas': 'despesas_comerciais',
        'Despesas Gerais e Administrativas': 'despesas_ga',
        'Total de despesas SG&A': 'total_despesas_operacionais',
        'Despesas com Vendas / Vendas': 'pct_comerciais_receita_liquida',
        'Despesas Gerais e Administrativas / ': 'pct_ga_receita_liquida',
        'Outras Receitas e Despesas': 'outras_receitas_despesas_op',
        'Equivalência Patrimonial': 'equivalencia_patrimonial',
        'Receitas Financeiras': 'receitas_financeiras',
        'Despesas Financeiras': 'despesas_financeiras',
        'Resultado Financeiro': 'resultado_financeiro',
        'Lucro Líquido (Prejuízo)': 'lucro_liquido',
        'Lucro Líquido': 'lucro_liquido',
        'Margem Líquida': 'margem_liquida',
        '*EBITDA Ajustado': 'ebitda_ajustado',
        'Margem EBITDA ajustada': 'margem_ebitda_ajustada',
        'Receitas a Apropriar': 'receitas_apropriar',
        'Custo das Unidades Vendidas a Apropriar': 'custo_apropriar',
        'Resultado a Apropriar': 'resultado_apropriar',
        'Margem a Apropriar': 'margem_apropriar',
        'ROE': 'roe',
        'ROCE': 'roe',  # fallback
        'Dívida Bruta': 'divida_bruta',
        'Dívida Líquida': 'divida_liquida',
        'Patrimônio Líquido + Minoritários': 'patrimonio_liquido',
        'Dívida Líquida / (PL': 'divida_liquida_pl',
        'Caixa e Disponibilidades': 'caixa_aplicacoes',
    }

    # Recebíveis — labels found in Financeiro sheets (accent required: Provisão)
    # NB: "Contas a Receber" do BP inclui repasses CEF — NÃO é pro soluto puro.
    # Carteira pro soluto (pré/pós-chaves) vem dos releases, não do BP.
    recebiveis_map = {
        'Provis\u00e3o para devedores duvidosos': 'pdd_provisao',
    }

    segments = {
        'Tenda': ('Tenda - Operacional', 'Tenda - Tenda Financeiro'),
        'Alea': ('Alea - Operacional', 'Alea - Financeiro'),
        'Consolidado': ('Consolidado - Operacional', 'Consolidado - Financeiro'),
    }

    for seg, (aba_op, aba_fin) in segments.items():
        dados = extrair_dados_por_label(fp, aba_op, 7, 0, operacional_map)
        n, u = inserir_dados(conn, 'Tenda', seg, dados)
        print(f"    {seg} Operacional: {n} novos, {u} atualizações")

        dados = extrair_dados_por_label(fp, aba_fin, 7, 0, financeiro_map)
        n, u = inserir_dados(conn, 'Tenda', seg, dados)
        print(f"    {seg} Financeiro: {n} novos, {u} atualizações")

        # Recebíveis (PDD) from Financeiro sheet
        dados = extrair_dados_por_label(fp, aba_fin, 7, 0, recebiveis_map)
        # PDD is typically negative in the sheet; store as positive
        for periodo in dados:
            if 'pdd_provisao' in dados[periodo] and dados[periodo]['pdd_provisao'] < 0:
                dados[periodo]['pdd_provisao'] = abs(dados[periodo]['pdd_provisao'])
        n, u = inserir_dados(conn, 'Tenda', seg, dados)
        print(f"    {seg} Recebíveis: {n} novos, {u} atualizações")


# ============================================================
# CURY
# ============================================================
def ingerir_cury(conn):
    print("\n  --- CURY ---")
    fp = PLANILHAS_DIR / "Cury_Planilha_Fundamentos_2026-03.xlsx"

    # Operacional (header row 5, labels col B=1 0-indexed)
    op_map = {
        'Número de Empreendimentos': 'empreendimentos_lancados',
        'VGV (em R$ mil)': 'vgv_lancado',
        'Número de unidades': 'unidades_lancadas',
        'Preço médio por unidade': 'preco_medio_lancamento',
        'Tamanho médio dos lançamentos': 'tamanho_medio_empreendimentos',
        'Vendas Brutas': 'vendas_brutas_vgv',
        'VSO Bruto': 'vso_bruta_trimestral',
        'Vendas Líquidas': 'vendas_liquidas_vgv',
        'Distratos': 'distratos_vgv',
        'VSO Líquido': 'vso_liquida_trimestral',
        'VSO UDM': 'vso_liquida_12m',
        'Unidades Vendidas Brutas': 'vendas_brutas_unidades',
        'Unidades Distratadas': 'distratos_unidades',
        'Unidades Vendidas Líquidas': 'vendas_liquidas_unidades',
        'VGV Repassado': 'vgv_repassado',
        'Unidades Repassadas': 'unidades_repassadas',
        'Unidades concluídas': 'unidades_entregues',
        'Obras em andamento': 'obras_em_andamento',
        'Duração do Estoque': 'giro_estoque_meses',
    }
    # Cury VGV is in R$ mil -> convert to R$ milhões
    dados = extrair_dados_por_label(fp, 'Resultados Operacionais', 5, 1, op_map, multiplicador=0.001)
    n, u = inserir_dados(conn, 'Cury', 'Consolidado', dados)
    print(f"    Operacional: {n} novos, {u} atualizações")

    # Landbank is in R$ milhões (row 74 = total with *), Geração de Caixa in R$ mil (row 90)
    # Use row-based extraction to avoid label ambiguity (3 rows with same landbank label)
    dados = extrair_por_linha(fp, 'Resultados Operacionais', 5, {74: 'landbank_vgv'}, multiplicador=1.0)
    dados2 = extrair_por_linha(fp, 'Resultados Operacionais', 5, {90: 'geracao_caixa'}, multiplicador=0.001)
    for p, campos in dados2.items():
        if p not in dados:
            dados[p] = {}
        dados[p].update(campos)
    n, u = inserir_dados(conn, 'Cury', 'Consolidado', dados)
    print(f"    Landbank+Caixa: {n} novos, {u} atualizações")

    # DRE (header row 3, labels col B=1)
    dre_map = {
        'Receita líquida': 'receita_liquida',
        'Custo das vendas': 'custo_imoveis_vendidos',
        'Lucro bruto': 'lucro_bruto',
        'Margem bruta': 'margem_bruta',
        'Margem Bruta Ajustada': 'margem_bruta_ajustada',
        'Despesas com vendas': 'despesas_comerciais',
        'Despesas gerais e administrativas': 'despesas_ga',
        'Outros resultados operacionais': 'outras_receitas_despesas_op',
        'Equivalência patrimonial': 'equivalencia_patrimonial',
        'Receitas (despesas) operacionais': 'total_despesas_operacionais',
        'Despesas financeiras': 'despesas_financeiras',
        'Receitas financeiras': 'receitas_financeiras',
        'Resultado financeiro': 'resultado_financeiro',
        'Lucro líquido do exercício': 'lucro_liquido',
        'Margem líquida': 'margem_liquida',
        'Ebitda': 'ebitda',
        'Margem Ebitda': 'margem_ebitda',
        'Ebitda Ajustado': 'ebitda_ajustado',
        'Margem Ebitda Ajustada': 'margem_ebitda_ajustada',
        'Receitas de vendas a apropriar': 'receitas_apropriar',
        '(=) Resultado de vendas': 'resultado_apropriar',
        'Margem Bruta REF': 'margem_apropriar',
        'ROE': 'roe',
    }
    # Cury DRE values are in R$ (reais) despite confusing header — magnitudes match R$ mil
    dados = extrair_dados_por_label(fp, 'Demonstrações do Resultado', 3, 1, dre_map, multiplicador=0.001)
    n, u = inserir_dados(conn, 'Cury', 'Consolidado', dados)
    print(f"    DRE: {n} novos, {u} atualizações")

    # Balanço Ativo (header row 5, labels col B=1, uses datetime headers)
    bp_ativo_map = {
        'Caixa e equivalentes': 'caixa_aplicacoes',
    }
    dados = extrair_dados_por_label(fp, 'Balanço Patrimonial Ativo', 5, 1, bp_ativo_map, multiplicador=0.001)
    n, u = inserir_dados(conn, 'Cury', 'Consolidado', dados)
    print(f"    BP Ativo: {n} novos, {u} atualizações")

    # Balanço Passivo
    bp_passivo_map = {
        'Patrimônio Líquido total': 'patrimonio_liquido',
    }
    dados = extrair_dados_por_label(fp, 'Balanço Patrimonial Passivo', 4, 1, bp_passivo_map, multiplicador=0.001)
    n, u = inserir_dados(conn, 'Cury', 'Consolidado', dados)
    print(f"    BP Passivo: {n} novos, {u} atualizações")



# ============================================================
# DIRECIONAL
# ============================================================
def ingerir_direcional(conn):
    print("\n  --- DIRECIONAL ---")
    fp = PLANILHAS_DIR / "Direcional_Planilha_Interativa_2026-03.xlsx"

    # Dados Operacionais (header row 4, labels col B=1, data in R$ mil)
    op_map = {
        'VGV Lançado (R$ mil)': 'vgv_lancado',
        'Unidades Lançadas (unidades)': 'unidades_lancadas',
        'Preço Médio (R$/unidade)': 'preco_medio_lancamento',
        'VGV Contratado (R$ mil)': 'vendas_brutas_vgv',
        'Unidades Contratadas (unidades)': 'vendas_brutas_unidades',
        'VSO (Vendas Sobre Oferta) - Consolidada': 'vso_liquida_trimestral',
        'Estoque Total a Valor de Mercado (R$ mil)': 'estoque_vgv',
        'Unidades em Estoque': 'estoque_unidades',
        'Land Bank Total (R$ mil)': 'landbank_vgv',
        'Land Bank (unidades)': 'landbank_unidades',
        'Repasses (R$ mil)': 'vgv_repassado',
    }
    dados = extrair_dados_por_label(fp, 'Dados Operacionais', 4, 1, op_map, multiplicador=0.001)
    n, u = inserir_dados(conn, 'Direcional', 'Consolidado', dados)
    print(f"    Operacional: {n} novos, {u} atualizações")

    # DRE (header row 4, labels col B=1, data in R$ mil)
    dre_map = {
        'RECEITA OPERACIONAL': 'receita_liquida',
        'Custo da Venda': 'custo_imoveis_vendidos',
        'LUCRO BRUTO': 'lucro_bruto',
        'Despesas Gerais e Ad': 'despesas_ga',
        'Despesas Comerciais': 'despesas_comerciais',
        'Resultado com Equivalência': 'equivalencia_patrimonial',
        'Outras Receitas e Despesas': 'outras_receitas_despesas_op',
        'Despesas Financeiras': 'despesas_financeiras',
        'Receitas Financeiras': 'receitas_financeiras',
        'LUCRO LÍQUIDO DO PERÍODO': 'lucro_liquido',
    }
    dados = extrair_dados_por_label(fp, 'Demonstração de Resultados', 4, 1, dre_map, multiplicador=0.001)
    n, u = inserir_dados(conn, 'Direcional', 'Consolidado', dados)
    print(f"    DRE: {n} novos, {u} atualizações")

    # Balanço (header row 4, labels col B=1)
    bp_map = {
        'Caixa e equivalentes de caixa': 'caixa_aplicacoes',
        'PATRIMÔNIO LÍQUIDO': 'patrimonio_liquido',
    }
    dados = extrair_dados_por_label(fp, 'Balanço Patrimonial', 4, 1, bp_map, multiplicador=0.001)
    n, u = inserir_dados(conn, 'Direcional', 'Consolidado', dados)
    print(f"    Balanço: {n} novos, {u} atualizações")

    # Passivo de Cessão (saldo cessão recebíveis = current row 35 + non-current row 51)
    try:
        dados_c = extrair_por_linha(fp, 'Balanço Patrimonial', 4, {35: 'saldo_cessao_recebiveis'}, multiplicador=0.001)
        dados_nc = extrair_por_linha(fp, 'Balanço Patrimonial', 4, {51: 'saldo_cessao_recebiveis'}, multiplicador=0.001)
        for p in set(list(dados_c.keys()) + list(dados_nc.keys())):
            c_val = dados_c.get(p, {}).get('saldo_cessao_recebiveis', 0) or 0
            nc_val = dados_nc.get(p, {}).get('saldo_cessao_recebiveis', 0) or 0
            total = c_val + nc_val
            if total > 0:
                if p not in dados_c:
                    dados_c[p] = {}
                dados_c[p]['saldo_cessao_recebiveis'] = total
        n, u = inserir_dados(conn, 'Direcional', 'Consolidado', dados_c)
        print(f"    Cessão Recebíveis: {n} novos, {u} atualizações")
    except Exception as e:
        print(f"    Cessão: ERRO - {e}")

    # RIVA DRE
    riva_map = {
        'RECEITA OPERACIONAL': 'receita_liquida',
        'Custo da Venda': 'custo_imoveis_vendidos',
        'LUCRO BRUTO': 'lucro_bruto',
        'Despesas comerciais, gerais': 'total_despesas_operacionais',
        'Resultado Financeiro': 'resultado_financeiro',
        'LUCRO LÍQUIDO': 'lucro_liquido',
    }
    try:
        dados = extrair_dados_por_label(fp, 'RIVA - DRE', 4, 1, riva_map, multiplicador=0.001)
        n, u = inserir_dados(conn, 'Direcional', 'Riva', dados)
        print(f"    Riva DRE: {n} novos, {u} atualizações")
    except Exception as e:
        print(f"    Riva: ERRO - {e}")


# ============================================================
# PLANO&PLANO
# ============================================================
def ingerir_planoeplano(conn):
    print("\n  --- PLANO&PLANO ---")
    fp = PLANILHAS_DIR / "PlanoePlano_Planilha_Interativa_2026-03.xlsx"

    # Dados Operacionais (header row 7, labels col B=1, data R$ mil)
    op_map = {
        'Vendas Contratadas Brutas (R$ mil)': 'vendas_brutas_vgv',
        'Vendas Contratadas Brutas (Unidades)': 'vendas_brutas_unidades',
        'Distratos (R$ mil)': 'distratos_vgv',
        'Distratos (Unidades)': 'distratos_unidades',
        'Vendas Líquidas 100% Plano&Plano (R$ mil)': 'vendas_liquidas_vgv',
        'Vendas Líquidas 100% Plano&Plano (Unid': 'vendas_liquidas_unidades',
        'Preço Venda Médio': 'preco_medio_vendas',
        'VGV 100% Plano&Plano (R$ mil)': 'vgv_lancado',
        'Unidades': 'unidades_lancadas',
        'Média de Unidades por': 'tamanho_medio_empreendimentos',
        'Estoque VGV 100%': 'estoque_vgv',
        'Estoque (Unidades)': 'estoque_unidades',
        '% Pronto (unidades)': 'pct_estoque_pronto',
        'VSO Líquido 100% - últimos 12 meses': 'vso_liquida_12m',
    }
    dados = extrair_dados_por_label(fp, 'Dados Operacionais', 7, 1, op_map, multiplicador=0.001)
    n, u = inserir_dados(conn, 'PlanoePlano', 'Consolidado', dados)
    print(f"    Operacional: {n} novos, {u} atualizações")

    # DRE (header row 7, labels col B=1, data R$ mil)
    dre_map = {
        'Receita líquida': 'receita_liquida',
        'Custos dos imóveis vendidos': 'custo_imoveis_vendidos',
        'Lucro bruto': 'lucro_bruto',
        'Margem bruta': 'margem_bruta',
        'Despesas comerciais': 'despesas_comerciais',
        'Despesas gerais e administrativas': 'despesas_ga',
        'Resultado da equivalência': 'equivalencia_patrimonial',
        'Outras receitas (despesas)': 'outras_receitas_despesas_op',
        'Despesas (receitas) operacionais': 'total_despesas_operacionais',
        'Receitas financeiras': 'receitas_financeiras',
        'Despesas financeiras': 'despesas_financeiras',
        'Receitas (despesas) financeiras': 'resultado_financeiro',
        'Lucro líquido 100%': 'lucro_liquido',
        'Margem Líquida 100%': 'margem_liquida',
        'Receitas de vendas a apropriar': 'receitas_apropriar',
        '(=) Resultado de vendas': 'resultado_apropriar',
        'Margem Bruta REF': 'margem_apropriar',
    }
    dados = extrair_dados_por_label(fp, 'DRE Consolidado', 7, 1, dre_map, multiplicador=0.001)
    n, u = inserir_dados(conn, 'PlanoePlano', 'Consolidado', dados)
    print(f"    DRE: {n} novos, {u} atualizações")

    # Balanço (header row 7, labels col B=1)
    bp_map = {
        'Caixa e equivalentes de caixa': 'caixa_aplicacoes',
        'Total do patrimônio líquido': 'patrimonio_liquido',
    }
    try:
        dados = extrair_dados_por_label(fp, 'Balanço Patrimonial Consolidado', 7, 1, bp_map, multiplicador=0.001)
        n, u = inserir_dados(conn, 'PlanoePlano', 'Consolidado', dados)
        print(f"    Balanço: {n} novos, {u} atualizações")
    except Exception as e:
        print(f"    Balanço: ERRO - {e}")

    # Cessão recebíveis + CRI (row-based: current rows 31,32 + non-current rows 49,50)
    try:
        cessao_rows = {31: 'saldo_cessao_recebiveis', 32: 'saldo_cessao_recebiveis'}
        dados_c = extrair_por_linha(fp, 'Balanço Patrimonial Consolidado', 7, {31: '_cessao_c'}, multiplicador=0.001)
        dados_cri_c = extrair_por_linha(fp, 'Balanço Patrimonial Consolidado', 7, {32: '_cri_c'}, multiplicador=0.001)
        dados_nc = extrair_por_linha(fp, 'Balanço Patrimonial Consolidado', 7, {49: '_cessao_nc'}, multiplicador=0.001)
        dados_cri_nc = extrair_por_linha(fp, 'Balanço Patrimonial Consolidado', 7, {50: '_cri_nc'}, multiplicador=0.001)
        # Sum all 4 components
        all_periods = set()
        for d in [dados_c, dados_cri_c, dados_nc, dados_cri_nc]:
            all_periods.update(d.keys())
        result = {}
        for p in all_periods:
            total = 0
            for d, key in [(dados_c, '_cessao_c'), (dados_cri_c, '_cri_c'),
                           (dados_nc, '_cessao_nc'), (dados_cri_nc, '_cri_nc')]:
                total += d.get(p, {}).get(key, 0) or 0
            if total > 0:
                result[p] = {'saldo_cessao_recebiveis': total}
        n, u = inserir_dados(conn, 'PlanoePlano', 'Consolidado', result)
        print(f"    Cessão+CRI: {n} novos, {u} atualizações")
    except Exception as e:
        print(f"    Cessão+CRI: ERRO - {e}")


# ============================================================
# MRV
# ============================================================
def ingerir_mrv(conn):
    print("\n  --- MRV ---")
    fp = PLANILHAS_DIR / "MRV_Base_Dados_Operacionais_Financeiros_2026-03.xlsx"

    # DRE labels (bilingual sheets, header row 2, labels col A=0)
    dre_map = {
        'Receita Operacional Líquida': 'receita_liquida',
        'Custo dos Imóveis Vendidos': 'custo_imoveis_vendidos',
        'Lucro Bruto': 'lucro_bruto',
        'Margem Bruta (%)': 'margem_bruta',
        'Despesas Comerciais': 'despesas_comerciais',
        'Despesas Gerais e Administrativas': 'despesas_ga',
        'Outras receitas (despesas)': 'outras_receitas_despesas_op',
        'Resultado de Equivalência': 'equivalencia_patrimonial',
        'Resultado Financeiro': 'resultado_financeiro',
        'Despesas financeiras': 'despesas_financeiras',
        'Receitas financeiras': 'receitas_financeiras',
        'Lucro Líquido do Período': 'lucro_liquido',
        'Lucro Líquido atribuível aos Acionistas': 'lucro_liquido',
        'Margem Líquida (%)': 'margem_liquida',
        'Receita Bruta de Vendas a apropriar': 'receitas_apropriar',
        'Resultado a apropriar': 'resultado_apropriar',
        '% Margem do Resultado a apropriar': 'margem_apropriar',
    }

    # DRE Consolidado (R$ mil)
    dados = extrair_dados_por_label(fp, 'DRE Consolid. | Income Statem.', 2, 0, dre_map, multiplicador=0.001)
    n, u = inserir_dados(conn, 'MRV', 'Consolidado', dados)
    print(f"    DRE Consolidado: {n} novos, {u} atualizações")

    # Indicadores Financeiros (header row 2, mixed units — R$ mil for monetary, % for ratios)
    fin_map = {
        'Margem Bruta ex. juros': 'margem_bruta_ajustada',
        'Despesas Comerciais / ROL': 'pct_comerciais_receita_liquida',
        'Despesas G&A / ROL': 'pct_ga_receita_liquida',
        'EBITDA': 'ebitda',
        'EBITDA Margin': 'margem_ebitda',
        'ROE (12 meses)': 'roe',
        'Dívida Total': 'divida_bruta',
        '(-) Caixa e Equivalentes': 'caixa_aplicacoes',
        'Dívida Líquida': 'divida_liquida',
        'Total do Patrimônio Líquido': 'patrimonio_liquido',
        'Dívida Líquida / PL': 'divida_liquida_pl',
        'Geração de Caixa': 'geracao_caixa',
    }
    dados = extrair_dados_por_label(fp, 'Indic.Fin. | Financ.Highlights', 2, 0, fin_map, multiplicador=0.001)
    n, u = inserir_dados(conn, 'MRV', 'Consolidado', dados)
    print(f"    Indicadores: {n} novos, {u} atualizações")

    # DRE por segmento (R$ mil)
    for seg, aba in [('MRV Incorporação', 'DRE MRV Inc. | Income Statem.'),
                      ('Resia', 'DRE Resia | Income Statem.'),
                      ('Urba', 'DRE Urba | Income Statem.'),
                      ('Luggo', 'DRE Luggo | Income Statem.')]:
        try:
            dados = extrair_dados_por_label(fp, aba, 2, 0, dre_map, multiplicador=0.001)
            n, u = inserir_dados(conn, 'MRV', seg, dados)
            print(f"    DRE {seg}: {n} novos, {u} atualizações")
        except Exception as e:
            print(f"    DRE {seg}: ERRO - {e}")

    # BP por segmento
    bp_map = {
        'Caixa e Equivalentes': 'caixa_aplicacoes',
        'Total do Patrimônio Líquido': 'patrimonio_liquido',
    }
    for seg, aba in [('MRV Incorporação', 'BP MRV Inc. | MRV Inc. BS'),
                      ('Resia', 'BP Resia | Resia BS'),
                      ('Urba', 'BP Urba | Urba BS'),
                      ('Luggo', 'BP Luggo | Luggo BS')]:
        try:
            dados = extrair_dados_por_label(fp, aba, 2, 0, bp_map, multiplicador=0.001)
            n, u = inserir_dados(conn, 'MRV', seg, dados)
            print(f"    BP {seg}: {n} novos, {u} atualizações")
        except Exception as e:
            print(f"    BP {seg}: ERRO - {e}")

    # BP Consolidado (R$ mil)
    bp_consol_map = {
        'Caixa e Equivalentes': 'caixa_aplicacoes',
        'Total do Patrimônio Líquido': 'patrimonio_liquido',
    }
    try:
        dados = extrair_dados_por_label(fp, 'BP Consolid. | Consolid. BS', 2, 0, bp_consol_map, multiplicador=0.001)
        n, u = inserir_dados(conn, 'MRV', 'Consolidado', dados)
        print(f"    BP Consolidado: {n} novos, {u} atualizações")
    except Exception as e:
        print(f"    BP Consolidado: ERRO - {e}")

    # Passivo de Cessão Consolidado (current row 43 + non-current row 55)
    try:
        dados_c = extrair_por_linha(fp, 'BP Consolid. | Consolid. BS', 2, {43: '_cessao_c'}, multiplicador=0.001)
        dados_nc = extrair_por_linha(fp, 'BP Consolid. | Consolid. BS', 2, {55: '_cessao_nc'}, multiplicador=0.001)
        result = {}
        for p in set(list(dados_c.keys()) + list(dados_nc.keys())):
            c_val = dados_c.get(p, {}).get('_cessao_c', 0) or 0
            nc_val = dados_nc.get(p, {}).get('_cessao_nc', 0) or 0
            total = c_val + nc_val
            if total > 0:
                result[p] = {'saldo_cessao_recebiveis': total}
        n, u = inserir_dados(conn, 'MRV', 'Consolidado', result)
        print(f"    Cessão Consolidado: {n} novos, {u} atualizações")
    except Exception as e:
        print(f"    Cessão Consolidado: ERRO - {e}")

    # Dados Operacionais (row-based extraction due to repeating labels)
    # Header row 2. Sections use 100% values (not %MRV).
    aba_op = 'Dados Oper. MRV&Co | Oper.Data'

    # Landbank 100% section (starts row 33)
    # Row 35: MRV&Co VGV (R$ bilhões) → convert to R$ milhões (*1000)
    # Row 36: MRV&Co Unidades
    # Row 39: MRV Inc VGV, Row 40: MRV Inc Units
    # Row 43: Urba VGV, Row 44: Urba Units
    try:
        # MRV&Co (Consolidado)
        dados = extrair_por_linha(fp, aba_op, 2, {35: 'landbank_vgv', 36: 'landbank_unidades'}, multiplicador=1000)
        n, u = inserir_dados(conn, 'MRV', 'Consolidado', dados)
        print(f"    Landbank Consolidado: {n} novos, {u} atualizações")

        # MRV Inc
        dados = extrair_por_linha(fp, aba_op, 2, {39: 'landbank_vgv', 40: 'landbank_unidades'}, multiplicador=1000)
        n, u = inserir_dados(conn, 'MRV', 'MRV Incorporação', dados)
        print(f"    Landbank MRV Inc: {n} novos, {u} atualizações")

        # Urba
        dados = extrair_por_linha(fp, aba_op, 2, {43: 'landbank_vgv', 44: 'landbank_unidades'}, multiplicador=1000)
        n, u = inserir_dados(conn, 'MRV', 'Urba', dados)
        print(f"    Landbank Urba: {n} novos, {u} atualizações")
    except Exception as e:
        print(f"    Landbank: ERRO - {e}")

    # Launches 100% section (starts row 80)
    # Row 82: MRV&Co VGV (R$ milhões), Row 83: Units
    # Row 86: MRV Inc VGV, Row 87: Units
    # Row 90: Urba VGV, Row 91: Units
    try:
        dados = extrair_por_linha(fp, aba_op, 2, {82: 'vgv_lancado', 83: 'unidades_lancadas'})
        n, u = inserir_dados(conn, 'MRV', 'Consolidado', dados)
        print(f"    Lançamentos Consolidado: {n} novos, {u} atualizações")

        dados = extrair_por_linha(fp, aba_op, 2, {86: 'vgv_lancado', 87: 'unidades_lancadas'})
        n, u = inserir_dados(conn, 'MRV', 'MRV Incorporação', dados)
        print(f"    Lançamentos MRV Inc: {n} novos, {u} atualizações")

        dados = extrair_por_linha(fp, aba_op, 2, {90: 'vgv_lancado', 91: 'unidades_lancadas'})
        n, u = inserir_dados(conn, 'MRV', 'Urba', dados)
        print(f"    Lançamentos Urba: {n} novos, {u} atualizações")
    except Exception as e:
        print(f"    Lançamentos op: ERRO - {e}")

    # Net Pre-sales 100% section (starts row 205)
    # Row 207: MRV&Co Vendas Líq (R$ milhões), Row 208: Units
    # Row 211: MRV Inc Vendas Líq, Row 213: Units (row 212 is price)
    # Row 215: Urba Vendas Líq, Row 216: Units
    try:
        dados = extrair_por_linha(fp, aba_op, 2, {207: 'vendas_liquidas_vgv', 208: 'vendas_liquidas_unidades'})
        n, u = inserir_dados(conn, 'MRV', 'Consolidado', dados)
        print(f"    Vendas Líq Consolidado: {n} novos, {u} atualizações")

        dados = extrair_por_linha(fp, aba_op, 2, {211: 'vendas_liquidas_vgv', 212: 'vendas_liquidas_unidades'})
        n, u = inserir_dados(conn, 'MRV', 'MRV Incorporação', dados)
        print(f"    Vendas Líq MRV Inc: {n} novos, {u} atualizações")

        dados = extrair_por_linha(fp, aba_op, 2, {215: 'vendas_liquidas_vgv', 216: 'vendas_liquidas_unidades'})
        n, u = inserir_dados(conn, 'MRV', 'Urba', dados)
        print(f"    Vendas Líq Urba: {n} novos, {u} atualizações")
    except Exception as e:
        print(f"    Vendas Líq op: ERRO - {e}")

    # VSO from %MRV section (row 187 = MRV Inc VSO líquida)
    try:
        dados = extrair_por_linha(fp, aba_op, 2, {187: 'vso_liquida_trimestral'})
        n, u = inserir_dados(conn, 'MRV', 'MRV Incorporação', dados)
        print(f"    VSO MRV Inc: {n} novos, {u} atualizações")
    except Exception as e:
        print(f"    VSO: ERRO - {e}")


# ============================================================
# CYRELA
# ============================================================
def ingerir_cyrela(conn):
    print("\n  --- CYRELA ---")

    # Demonstrações Financeiras (header row 4, labels col A=0, datetime headers)
    fp_df = PLANILHAS_DIR / "Cyrela_Demonstracoes_Financeiras_2026-03.xlsx"

    # Cyrela DRE labels use Economatica format: +/-/= prefix, no accents
    dre_map = {
        'Receita liquida operac': 'receita_liquida',
        'Custo Produtos Vendidos': 'custo_imoveis_vendidos',
        'Lucro Bruto': 'lucro_bruto',
        'Despesas com Vendas': 'despesas_comerciais',
        'Despesas administrativ': 'despesas_ga',
        'Equivalenc patrimonial': 'equivalencia_patrimonial',
        'Outras rec operacionais': 'outras_receitas_despesas_op',
        'Lucro antes jur&imp': 'ebitda',
        'Resultado financeiro': 'resultado_financeiro',
        'Receitas Financeiras': 'receitas_financeiras',
        'Despesas Financeiras': 'despesas_financeiras',
        'Lucro liquido': 'lucro_liquido',
        'Lucro Consolidado': 'lucro_liquido',
    }
    # Cyrela DRE is in R$ mil (milhares)
    dados = extrair_dados_por_label(fp_df, 'CYRE3', 4, 0, dre_map, multiplicador=0.001)
    n, u = inserir_dados(conn, 'Cyrela', 'Consolidado', dados)
    print(f"    DRE: {n} novos, {u} atualizações")

    # Balanço from same sheet (R$ mil) — labels without accents
    bp_map = {
        'Caixa e equivalentes de caixa': 'caixa_aplicacoes',
        'Patrimonio liquido': 'patrimonio_liquido',
        'Patrim liq consolidado': 'patrimonio_liquido',
    }
    dados = extrair_dados_por_label(fp_df, 'CYRE3', 4, 0, bp_map, multiplicador=0.001)
    n, u = inserir_dados(conn, 'Cyrela', 'Consolidado', dados)
    print(f"    Balanço: {n} novos, {u} atualizações")

    # Dados Operacionais (header row 4, periods in columns)
    # Sheet names have special chars: 'Lçtos'
    fp_op = PLANILHAS_DIR / "Cyrela_Dados_Operacionais_2026-03.xlsx"

    # Lançamentos - row 14 = Total VGV 100% (R$ mil), row 71 = Total Unidades
    try:
        ltos_rows = {14: 'vgv_lancado', 71: 'unidades_lancadas'}
        dados = extrair_por_linha(fp_op, 'Lçtos', 4, ltos_rows, multiplicador=0.001)
        n, u = inserir_dados(conn, 'Cyrela', 'Consolidado', dados)
        print(f"    Lançamentos: {n} novos, {u} atualizações")
    except Exception as e:
        print(f"    Lançamentos: ERRO - {e}")

    # Vendas - row 14 = Total VGV Vendas 100% (R$ mil), row 71 = Total Unidades Vendidas
    try:
        vendas_rows = {14: 'vendas_brutas_vgv', 71: 'vendas_brutas_unidades'}
        dados = extrair_por_linha(fp_op, 'Vendas', 4, vendas_rows, multiplicador=0.001)
        n, u = inserir_dados(conn, 'Cyrela', 'Consolidado', dados)
        print(f"    Vendas: {n} novos, {u} atualizações")
    except Exception as e:
        print(f"    Vendas: ERRO - {e}")

    # Estoque - row 14 = Total VGV Estoque 100% (R$ mil), row 58 = Total Unidades
    try:
        estoque_rows = {14: 'estoque_vgv', 58: 'estoque_unidades'}
        dados = extrair_por_linha(fp_op, 'Estoque', 4, estoque_rows, multiplicador=0.001)
        n, u = inserir_dados(conn, 'Cyrela', 'Consolidado', dados)
        print(f"    Estoque: {n} novos, {u} atualizações")
    except Exception as e:
        print(f"    Estoque: ERRO - {e}")

    # Terrenos (Landbank) - row 14 = Total VGV 100% (R$ mil), row 79 = Total Unidades
    try:
        terrenos_rows = {14: 'landbank_vgv', 79: 'landbank_unidades'}
        dados = extrair_por_linha(fp_op, 'Terrenos', 4, terrenos_rows, multiplicador=0.001)
        n, u = inserir_dados(conn, 'Cyrela', 'Consolidado', dados)
        print(f"    Terrenos: {n} novos, {u} atualizações")
    except Exception as e:
        print(f"    Terrenos: ERRO - {e}")


# ============================================================
# MAIN
# ============================================================
if __name__ == "__main__":
    print("=" * 70)
    print("CRIAÇÃO DO BANCO DE DADOS v2 (matching por label)")
    print("=" * 70)

    if os.path.exists(DB_PATH):
        os.remove(DB_PATH)

    conn = sqlite3.connect(DB_PATH)

    print("\n1. Criando schema...")
    criar_schema(conn)

    print("\n2. Ingerindo dados...")
    ingerir_tenda(conn)
    ingerir_cury(conn)
    ingerir_direcional(conn)
    ingerir_planoeplano(conn)
    ingerir_mrv(conn)
    ingerir_cyrela(conn)

    # 3. Summary
    print("\n" + "=" * 70)
    print("RESUMO")
    print("=" * 70)

    c = conn.cursor()
    c.execute("""SELECT empresa, segmento, COUNT(*),
                        MIN(ano * 10 + trimestre) as min_p, MAX(ano * 10 + trimestre) as max_p,
                        MIN(periodo), MAX(periodo)
                 FROM dados_trimestrais GROUP BY empresa, segmento ORDER BY empresa, segmento""")
    for row in c.fetchall():
        # Convert encoded period back: 20201=1T2020, 20253=3T2025
        min_tri, min_ano = row[3] % 10, row[3] // 10
        max_tri, max_ano = row[4] % 10, row[4] // 10
        print(f"  {row[0]:15s} | {row[1]:20s} | {row[2]:3d} reg | {min_tri}T{min_ano} a {max_tri}T{max_ano}")

    c.execute("SELECT COUNT(*) FROM dados_trimestrais")
    total = c.fetchone()[0]

    # Field completeness
    print(f"\n  Total registros: {total}")

    campos = ['receita_liquida', 'lucro_bruto', 'margem_bruta', 'lucro_liquido',
              'despesas_comerciais', 'despesas_ga', 'ebitda', 'ebitda_ajustado',
              'vgv_lancado', 'vendas_liquidas_vgv', 'vso_liquida_trimestral',
              'landbank_vgv', 'divida_liquida', 'patrimonio_liquido', 'roe', 'geracao_caixa',
              'pdd_provisao', 'saldo_cessao_recebiveis']

    print(f"\n  {'Campo':35s} {'Preenchidos':>12s} {'%':>6s}")
    print(f"  {'-'*35} {'-'*12} {'-'*6}")
    for campo in campos:
        c.execute(f"SELECT COUNT(*) FROM dados_trimestrais WHERE {campo} IS NOT NULL")
        cnt = c.fetchone()[0]
        print(f"  {campo:35s} {cnt:12d} {cnt/total*100:5.0f}%")

    # Spot checks
    spot_checks = [
        ('Tenda', 'Consolidado', '3T2025'),
        ('MRV', 'Consolidado', '3T2025'),
        ('Cyrela', 'Consolidado', '3T2025'),
        ('Cury', 'Consolidado', '3T2025'),
    ]
    for emp, seg, per in spot_checks:
        print(f"\n  === SPOT CHECK: {emp} {seg} {per} ===")
        c.execute(f"""SELECT receita_liquida, lucro_bruto, lucro_liquido,
                            despesas_comerciais, despesas_ga, ebitda_ajustado,
                            vgv_lancado, vendas_liquidas_vgv, landbank_vgv,
                            vso_liquida_trimestral, patrimonio_liquido
                     FROM dados_trimestrais
                     WHERE empresa=? AND segmento=? AND periodo=?""", (emp, seg, per))
        row = c.fetchone()
        if row:
            for label, val in zip(['Receita Líq', 'Lucro Bruto', 'Lucro Líq',
                                    'Desp Comerc', 'Desp G&A', 'EBITDA Aj',
                                    'VGV Lançado', 'Vendas Líq', 'Landbank',
                                    'VSO Líq Tri', 'PL'], row):
                print(f"    {label:15s}: {val}")
        else:
            print(f"    (sem dados)")

    print(f"\n  Banco salvo em: {DB_PATH}")
    conn.close()

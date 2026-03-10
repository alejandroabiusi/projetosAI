# -*- coding: utf-8 -*-
"""
Atualiza schema_sugestao_v2.xlsx:
1. Adiciona novos campos SG&A e VSO revisados
2. Atualiza nomes de arquivos renomeados (Tenda releases + todas planilhas)
3. Atualiza referências no inventário
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import re

wb = openpyxl.load_workbook("C:/Projetos_AI/analise_releases/schema_sugestao.xlsx")

header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, size=11, color="FFFFFF")
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

# ============================================================
# 1. ATUALIZAR schema_campos — remover VSO antigos, adicionar novos
# ============================================================
ws1 = wb["schema_campos"]

# Find and remove old VSO rows (vso_bruta, vso_liquida)
rows_to_remove = []
for row in range(2, ws1.max_row + 1):
    campo = ws1.cell(row=row, column=3).value
    if campo in ('vso_bruta', 'vso_liquida'):
        rows_to_remove.append(row)

# Remove in reverse order
for row in sorted(rows_to_remove, reverse=True):
    ws1.delete_rows(row)

# Find current max campo_id
max_id = 0
for row in range(2, ws1.max_row + 1):
    val = ws1.cell(row=row, column=1).value
    if val and str(val).isdigit():
        max_id = max(max_id, int(val))

# New fields to add
new_fields = [
    # VSO revisados (4 campos)
    ('Vendas', 'vso_bruta_trimestral', 'VSO Bruta Trimestral', '%', 'ambos',
     'SIM', 'SIM', 'SIM', 'SIM', 'NÃO', 'SIM', 'PlanoePlano divulga apenas UDM'),
    ('Vendas', 'vso_liquida_trimestral', 'VSO Líquida Trimestral', '%', 'ambos',
     'SIM', 'SIM', 'SIM', 'SIM', 'NÃO', 'SIM', 'PlanoePlano divulga apenas UDM'),
    ('Vendas', 'vso_bruta_12m', 'VSO Bruta UDM 12 meses', '%', 'ambos',
     'NÃO', 'PARCIAL', 'NÃO', 'SIM', 'SIM', 'SIM', 'Nem todas divulgam; Direcional divulga 9M acumulado'),
    ('Vendas', 'vso_liquida_12m', 'VSO Líquida UDM 12 meses', '%', 'ambos',
     'NÃO', 'PARCIAL', 'NÃO', 'SIM', 'SIM', 'SIM', 'Nem todas divulgam; Direcional divulga 9M acumulado'),

    # SG&A — campos absolutos
    ('SG&A', 'despesas_comerciais', 'Despesas Comerciais / Com Vendas', 'R$ milhões', 'ambos',
     'SIM', 'SIM', 'SIM', 'SIM', 'SIM', 'SIM', 'Inclui comissões, marketing, PDV'),
    ('SG&A', 'despesas_ga', 'Despesas Gerais e Administrativas (G&A)', 'R$ milhões', 'ambos',
     'SIM', 'SIM', 'SIM', 'SIM', 'SIM', 'SIM', ''),
    ('SG&A', 'honorarios_administracao', 'Honorários de Administração', 'R$ milhões', 'release',
     'NÃO', 'NÃO', 'NÃO', 'NÃO', 'NÃO', 'SIM', 'Apenas Cyrela divulga separado'),
    ('SG&A', 'outras_receitas_despesas_op', 'Outras Receitas/Despesas Operacionais', 'R$ milhões', 'ambos',
     'SIM', 'SIM', 'SIM', 'SIM', 'SIM', 'SIM', ''),
    ('SG&A', 'equivalencia_patrimonial', 'Resultado de Equivalência Patrimonial', 'R$ milhões', 'ambos',
     'SIM', 'SIM', 'SIM', 'SIM', 'SIM', 'SIM', ''),
    ('SG&A', 'total_despesas_operacionais', 'Total Despesas Operacionais (SG&A)', 'R$ milhões', 'ambos',
     'SIM', 'SIM', 'SIM', 'SIM', 'SIM', 'SIM', 'Soma de comerciais + G&A + outras'),
    ('SG&A', 'despesas_sop', 'Despesas com Stock Option Plan', 'R$ milhões', 'release',
     'SIM', 'PARCIAL', 'PARCIAL', 'PARCIAL', 'PARCIAL', 'PARCIAL', 'Usado no ajuste do EBITDA'),

    # SG&A — indicadores % receita
    ('SG&A Indicadores', 'pct_comerciais_receita_bruta', 'Desp. Comerciais / Receita Bruta', '%', 'calculado',
     'SIM', 'SIM', 'SIM', 'SIM', 'SIM', 'SIM', 'Calculado'),
    ('SG&A Indicadores', 'pct_comerciais_receita_liquida', 'Desp. Comerciais / Receita Líquida', '%', 'ambos',
     'SIM', 'SIM', 'SIM', 'SIM', 'SIM', 'SIM', 'Divulgado por Tenda, Direcional, Cury, PlanoePlano'),
    ('SG&A Indicadores', 'pct_ga_receita_bruta', 'G&A / Receita Bruta', '%', 'calculado',
     'SIM', 'SIM', 'SIM', 'SIM', 'SIM', 'SIM', 'Calculado'),
    ('SG&A Indicadores', 'pct_ga_receita_liquida', 'G&A / Receita Líquida', '%', 'ambos',
     'SIM', 'SIM', 'SIM', 'SIM', 'SIM', 'SIM', 'Divulgado por Tenda, Direcional, Cury, PlanoePlano'),
    ('SG&A Indicadores', 'pct_sga_receita_bruta', 'SG&A Total / Receita Bruta', '%', 'calculado',
     'SIM', 'SIM', 'SIM', 'SIM', 'SIM', 'SIM', 'Calculado'),
    ('SG&A Indicadores', 'pct_sga_receita_liquida', 'SG&A Total / Receita Líquida', '%', 'calculado',
     'SIM', 'SIM', 'SIM', 'SIM', 'SIM', 'SIM', 'Calculado'),
    ('SG&A Indicadores', 'pct_ga_lancamentos', 'G&A / Lançamentos VGV', '%', 'release',
     'SIM', 'NÃO', 'NÃO', 'NÃO', 'NÃO', 'NÃO', 'Apenas Tenda divulga'),
    ('SG&A Indicadores', 'pct_comerciais_vendas_liquidas', 'Desp. Comerciais / Vendas Líquidas VGV', '%', 'release',
     'SIM', 'NÃO', 'NÃO', 'NÃO', 'SIM', 'NÃO', 'Tenda e PlanoePlano divulgam'),
]

# Append new rows
next_row = ws1.max_row + 1
for field in new_fields:
    max_id += 1
    values = [max_id] + list(field)
    for col_idx, val in enumerate(values, 1):
        cell = ws1.cell(row=next_row, column=col_idx, value=val)
        cell.border = thin_border
    next_row += 1

# Also remove the old despesas_vendas and despesas_ga from DRE category (duplicates)
rows_to_remove2 = []
for row in range(2, ws1.max_row + 1):
    campo = ws1.cell(row=row, column=3).value
    cat = ws1.cell(row=row, column=2).value
    if cat == 'DRE' and campo in ('despesas_vendas', 'despesas_ga'):
        rows_to_remove2.append(row)
for row in sorted(rows_to_remove2, reverse=True):
    ws1.delete_rows(row)

# Re-number campo_id sequentially
for row in range(2, ws1.max_row + 1):
    ws1.cell(row=row, column=1, value=row - 1)

print(f"schema_campos atualizado: {ws1.max_row - 1} campos totais")

# ============================================================
# 2. ATUALIZAR releases_inventario — nomes Tenda
# ============================================================
ws3 = wb["releases_inventario"]

for row in range(2, ws3.max_row + 1):
    empresa = ws3.cell(row=row, column=1).value
    nome_atual = ws3.cell(row=row, column=6).value

    if empresa == 'Tenda' and nome_atual:
        # Fix current name: Tenda_Release_xTyyyy.pdf -> Tenda_Release_yyyy_xT.pdf
        m = re.match(r'Tenda_Release_(\d)T(\d{4})\.pdf', nome_atual)
        if m:
            tri = m.group(1)
            ano = m.group(2)
            new_name = f"Tenda_Release_{ano}_{tri}T.pdf"
            ws3.cell(row=row, column=6, value=new_name)

    # Update suggested name to match Empresa_Release_YYYY_xT pattern for all
    if nome_atual and nome_atual.endswith('.pdf'):
        periodo = ws3.cell(row=row, column=2).value
        if periodo:
            m = re.match(r'(\d)T(\d{4})', periodo)
            if m:
                tri = m.group(1)
                ano = m.group(2)
                suggested = f"{empresa}_Release_{ano}_{tri}T.pdf"
                ws3.cell(row=row, column=7, value=suggested)

print(f"releases_inventario atualizado: {ws3.max_row - 1} releases")

# ============================================================
# 3. ATUALIZAR planilhas_inventario — nomes novos
# ============================================================
ws4 = wb["planilhas_inventario"]

# Map old names to new names
planilha_rename = {
    'Cury Planilha Fundamentos.xlsx': 'Cury_Planilha_Fundamentos_2026-03.xlsx',
    'Cyrela Dados Operacionais.xlsx': 'Cyrela_Dados_Operacionais_2026-03.xlsx',
    'Cyrela Demonstrações Financeiras.xlsx': 'Cyrela_Demonstracoes_Financeiras_2026-03.xlsx',
    'Cyrela Lançamentos.xlsx': 'Cyrela_Lancamentos_2026-03.xlsx',
    'Cyrela Principais Indicadores.xlsx': 'Cyrela_Principais_Indicadores_2026-03.xlsx',
    'Direcional Planilha interativa.xlsx': 'Direcional_Planilha_Interativa_2026-03.xlsx',
    'Moura Planilha  de Fundamentos 3T25.xlsx': 'MouraDubeux_Planilha_Fundamentos_2026-03.xlsx',
    'mrv Base de Dados Operacionais e Financeiros.xlsx': 'MRV_Base_Dados_Operacionais_Financeiros_2026-03.xlsx',
    'PLano e PLano Planilha Interativa 4T25.xlsx': 'PlanoePlano_Planilha_Interativa_2026-03.xlsx',
    'Tenda-2026-01-11-LQGf6Bmd.xlsx': 'Tenda_Planilha_Fundamentos_2026-03.xlsx',
}

for row in range(2, ws4.max_row + 1):
    old_name = ws4.cell(row=row, column=6).value
    if old_name and old_name in planilha_rename:
        new_name = planilha_rename[old_name]
        ws4.cell(row=row, column=6, value=new_name)  # nome_arquivo_atual
        ws4.cell(row=row, column=7, value=new_name)  # nome_sugerido (same, already renamed)

print(f"planilhas_inventario atualizado: {ws4.max_row - 1} planilhas")

# ============================================================
# 4. ATUALIZAR campos_adicionais_planilhas — nomes novos
# ============================================================
ws5 = wb["campos_adicionais_planilhas"]

for row in range(2, ws5.max_row + 1):
    old_name = ws5.cell(row=row, column=2).value
    if old_name and old_name in planilha_rename:
        ws5.cell(row=row, column=2, value=planilha_rename[old_name])

print(f"campos_adicionais_planilhas atualizado: {ws5.max_row - 1} registros")

# ============================================================
# Auto-width para schema_campos
# ============================================================
for col in ws1.columns:
    max_len = max(len(str(cell.value or '')) for cell in col)
    ws1.column_dimensions[col[0].column_letter].width = min(max_len + 2, 45)

# Save
wb.save("C:/Projetos_AI/analise_releases/schema_sugestao_v2.xlsx")
wb.close()
print("\nArquivo schema_sugestao_v2.xlsx salvo com sucesso!")

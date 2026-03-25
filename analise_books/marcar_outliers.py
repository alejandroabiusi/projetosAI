"""
Marca outliers no books_database.xlsx com cores:
- Amarelo: outlier moderado (1.5x IQR)
- Vermelho: outlier extremo (3x IQR)
Usa metodo IQR (interquartile range) por coluna numerica.
Colunas de formula (pct_tcd, raso_sobre_custo) sao ignoradas.
"""

import openpyxl
from openpyxl.styles import PatternFill
from copy import copy
import statistics

YELLOW = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
RED = PatternFill(start_color="FF6666", end_color="FF6666", fill_type="solid")

# Colunas numericas para analisar (indice 1-based)
NUMERIC_COLS = {
    'unidades': 6,
    'avaliacao_unit': 7,
    'fracao_terreno_pct': 8,
    'valor_rs_m2': 9,
    'custo_construcao_rs': 10,
    'custo_construcao_pct_vgv': 11,
    'prazo_obra_meses': 12,
    'custo_total_unit': 13,
    'preco_nominal_unit': 15,
    'preco_raso_unit': 16,
    'li_pct': 17,
    'li_regua_pct': 18,
    'margem_bruta_economica_pct': 19,
    'margem_rasa_economica_pct': 20,
    'exposicao_rs_mil': 21,
    'exposicao_pct_vgv': 22,
    'renda': 23,
    'pct_ret1': 24,
}


def get_column_values(ws, col_idx, max_row):
    """Retorna dict {row: value} para valores numericos validos."""
    values = {}
    for r in range(2, max_row + 1):
        v = ws.cell(r, col_idx).value
        if v is not None and isinstance(v, (int, float)):
            values[r] = v
    return values


def calc_iqr_bounds(values_list):
    """Calcula limites IQR para deteccao de outliers."""
    if len(values_list) < 4:
        return None
    sorted_vals = sorted(values_list)
    n = len(sorted_vals)
    q1_idx = n * 0.25
    q3_idx = n * 0.75

    # Interpolacao linear para quartis
    def percentile(data, p):
        k = (len(data) - 1) * p
        f = int(k)
        c = f + 1
        if c >= len(data):
            return data[f]
        return data[f] + (k - f) * (data[c] - data[f])

    q1 = percentile(sorted_vals, 0.25)
    q3 = percentile(sorted_vals, 0.75)
    iqr = q3 - q1

    return {
        'q1': q1,
        'q3': q3,
        'iqr': iqr,
        'mild_low': q1 - 1.5 * iqr,
        'mild_high': q3 + 1.5 * iqr,
        'extreme_low': q1 - 3.0 * iqr,
        'extreme_high': q3 + 3.0 * iqr,
    }


def main():
    wb = openpyxl.load_workbook('books_database.xlsx')
    ws = wb.active
    max_row = ws.max_row

    total_yellow = 0
    total_red = 0

    print(f"Analisando {max_row - 1} linhas de dados...\n")
    print(f"{'Coluna':<35} {'Q1':>10} {'Q3':>10} {'IQR':>10} {'Amarelos':>10} {'Vermelhos':>10}")
    print("-" * 95)

    for col_name, col_idx in NUMERIC_COLS.items():
        values = get_column_values(ws, col_idx, max_row)
        if not values:
            continue

        bounds = calc_iqr_bounds(list(values.values()))
        if bounds is None or bounds['iqr'] == 0:
            continue

        col_yellow = 0
        col_red = 0

        for row, val in values.items():
            is_extreme = val < bounds['extreme_low'] or val > bounds['extreme_high']
            is_mild = val < bounds['mild_low'] or val > bounds['mild_high']

            if is_extreme:
                ws.cell(row, col_idx).fill = RED
                col_red += 1
            elif is_mild:
                ws.cell(row, col_idx).fill = YELLOW
                col_yellow += 1

        total_yellow += col_yellow
        total_red += col_red

        # Formatar valores para exibicao
        if 'pct' in col_name or col_name in ('li_pct', 'li_regua_pct'):
            fmt = lambda x: f"{x:.3f}"
        else:
            fmt = lambda x: f"{x:,.0f}"

        print(f"{col_name:<35} {fmt(bounds['q1']):>10} {fmt(bounds['q3']):>10} {fmt(bounds['iqr']):>10} {col_yellow:>10} {col_red:>10}")

    print("-" * 95)
    print(f"{'TOTAL':<35} {'':>10} {'':>10} {'':>10} {total_yellow:>10} {total_red:>10}")
    print(f"\nLegenda: AMARELO = outlier moderado (1.5x IQR) | VERMELHO = outlier extremo (3x IQR)")

    # Listar os outliers extremos (vermelhos) com detalhes
    print(f"\n{'='*95}")
    print("DETALHES DOS OUTLIERS EXTREMOS (VERMELHO):")
    print(f"{'='*95}")

    for col_name, col_idx in NUMERIC_COLS.items():
        values = get_column_values(ws, col_idx, max_row)
        if not values:
            continue
        bounds = calc_iqr_bounds(list(values.values()))
        if bounds is None or bounds['iqr'] == 0:
            continue

        extremes = []
        for row, val in values.items():
            if val < bounds['extreme_low'] or val > bounds['extreme_high']:
                proj = ws.cell(row, 2).value  # nome_projeto
                extremes.append((row, proj, val))

        if extremes:
            print(f"\n  {col_name} (limites: {bounds['extreme_low']:.2f} a {bounds['extreme_high']:.2f}):")
            for row, proj, val in extremes:
                direction = "BAIXO" if val < bounds['extreme_low'] else "ALTO"
                print(f"    Linha {row}: {proj} = {val:,} ({direction})")

    # Listar outliers moderados (amarelos) tambem
    print(f"\n{'='*95}")
    print("DETALHES DOS OUTLIERS MODERADOS (AMARELO):")
    print(f"{'='*95}")

    for col_name, col_idx in NUMERIC_COLS.items():
        values = get_column_values(ws, col_idx, max_row)
        if not values:
            continue
        bounds = calc_iqr_bounds(list(values.values()))
        if bounds is None or bounds['iqr'] == 0:
            continue

        milds = []
        for row, val in values.items():
            is_extreme = val < bounds['extreme_low'] or val > bounds['extreme_high']
            is_mild = val < bounds['mild_low'] or val > bounds['mild_high']
            if is_mild and not is_extreme:
                proj = ws.cell(row, 2).value
                milds.append((row, proj, val))

        if milds:
            print(f"\n  {col_name} (limites: {bounds['mild_low']:.2f} a {bounds['mild_high']:.2f}):")
            for row, proj, val in milds:
                direction = "BAIXO" if val < bounds['mild_low'] else "ALTO"
                print(f"    Linha {row}: {proj} = {val:,} ({direction})")

    wb.save('books_database.xlsx')
    print(f"\nArquivo salvo com as marcacoes de outliers!")


if __name__ == '__main__':
    main()

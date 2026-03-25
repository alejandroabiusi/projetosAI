"""
Extrai dados estruturados dos Books de Produto da Construtora Tenda.
Lê PDFs, parseia capa + slides One Page, e gera Excel consolidado.
Usa texto direto quando disponível, OCR via RapidOCR como fallback.
"""

import os
import re
import glob
import fitz  # PyMuPDF
import pdfplumber
from rapidocr_onnxruntime import RapidOCR
from openpyxl import Workbook

# OCR engine (inicializado uma vez)
_ocr_engine = None

def get_ocr():
    global _ocr_engine
    if _ocr_engine is None:
        _ocr_engine = RapidOCR()
    return _ocr_engine


# ── Helpers de parsing ──────────────────────────────────────────────

def parse_br_number(s: str) -> float | None:
    """Converte número em formato brasileiro (ponto = milhar, vírgula = decimal) para float."""
    if s is None:
        return None
    s = s.strip().replace("R$", "").replace(" ", "")
    if not s:
        return None
    # Remove pontos de milhar (ponto seguido de 3 dígitos)
    s = re.sub(r'\.(\d{3})', r'\1', s)
    # Vírgula decimal → ponto
    s = s.replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return None


def parse_br_pct(s: str) -> float | None:
    """Converte percentual brasileiro ('47,6%') para float decimal (0.476)."""
    if s is None:
        return None
    s = s.strip().replace("%", "").replace(" ", "")
    # Limpar artefatos de OCR: "40.,6" → "40,6", "36.,7" → "36,7"
    s = re.sub(r'\.,', ',', s)
    s = re.sub(r',\.', ',', s)
    if not s:
        return None
    s = s.replace(",", ".")
    try:
        return float(s) / 100.0
    except ValueError:
        return None


def extract_number_and_pct(text: str):
    """
    De '55.446 (47,6%)' extrai (55446.0, 0.476).
    De '18.137 (15,6%)' extrai (18137.0, 0.156).
    OCR pode usar ponto decimal: '9.294 (6.0%)' → (9294.0, 0.06)
    """
    # Aceita vírgula ou ponto como separador decimal no percentual
    m = re.search(r'([\d.]+)\s*\(\s*([\d.,]+)\s*%\s*\)', text)
    if m:
        num = parse_br_number(m.group(1))
        pct = parse_br_pct(m.group(2) + "%")
        return num, pct
    return None, None


def extract_two_pcts(text: str):
    """De '18,8% (22,0%)' ou '-7,5% (20,0%)' extrai (pct1, pct2)."""
    m = re.search(r'(-?[\d,.]+)\s*%\s*\(\s*(-?[\d,.]+)\s*%\s*\)', text)
    if m:
        p1 = parse_br_pct(m.group(1) + "%")
        p2 = parse_br_pct(m.group(2) + "%")
        return p1, p2
    return None, None


# ── Parsing da Capa ─────────────────────────────────────────────────

TEMPLATE_PROJETO = "Avenida II"
TEMPLATE_DATA = "22/10/2020"
TEMPLATE_TIPO = "Book de Investimento"

def parse_capa(text: str) -> dict:
    """Extrai nome_projeto, data_comite, tipo_comite da página de capa."""
    lines = text.split("\n")

    # tipo_comite: primeira linha "Book de..." que NÃO seja o template
    tipo_comite = None
    tipo_comite_fallback = None
    for line in lines:
        line_s = line.strip()
        if re.match(r'(Book de |Comit[eê])', line_s, re.IGNORECASE):
            if line_s != TEMPLATE_TIPO:
                tipo_comite = line_s
                break
            elif tipo_comite_fallback is None:
                tipo_comite_fallback = line_s
    if tipo_comite is None:
        tipo_comite = tipo_comite_fallback

    # nome_projeto: "Projeto:" que NÃO seja Avenida II
    nome_projeto = None
    for line in lines:
        m = re.search(r'Projeto:\s*(.+)', line)
        if m:
            val = m.group(1).strip()
            if val != TEMPLATE_PROJETO:
                nome_projeto = val
                break

    # data_comite: "Data Comitê:" que NÃO seja template
    data_comite = None
    for line in lines:
        m = re.search(r'Data\s+Comit[eê\u00ea]:\s*(.+)', line)
        if m:
            val = m.group(1).strip()
            if val != TEMPLATE_DATA:
                data_comite = val
                break

    return {
        "nome_projeto": nome_projeto,
        "data_comite": data_comite,
        "tipo_comite": tipo_comite,
    }


# ── Extração de texto: direto ou via OCR ────────────────────────────

def has_onepage_fields(text: str) -> bool:
    """Verifica se o texto contém campos numéricos típicos do One Page."""
    markers = [r'Valor R\$/m', r'Custo Constru', r'Pre.o Nominal', r'LI\s*\(R']
    count = sum(1 for m in markers if re.search(m, text))
    return count >= 2


def ocr_page(pdf_path: str, page_idx: int) -> str:
    """
    Faz OCR da imagem embutida no slide One Page.
    Prefere a maior imagem embutida (tabela de dados) em vez de renderizar
    a página inteira, evitando problemas de leitura com layout de 2 colunas.
    """
    with fitz.open(pdf_path) as doc:
        page = doc[page_idx]
        images = page.get_images()

        # Ordenar imagens por tamanho (maior primeiro), tentar OCR em cada uma
        candidates = []
        if images:
            for img in images:
                xref = img[0]
                try:
                    pix = fitz.Pixmap(doc, xref)
                    size = pix.width * pix.height
                    if size > 50000:  # Imagem significativa
                        candidates.append((size, xref, pix.n, pix.alpha))
                    pix = None
                except Exception:
                    continue
            candidates.sort(key=lambda x: -x[0])  # Maior primeiro

        ocr = get_ocr()

        for _, xref, n, alpha in candidates:
            pix = fitz.Pixmap(doc, xref)
            # Pular imagens com alpha (frequentemente são overlays vazios)
            # Converter CMYK → RGB se necessário
            if pix.n > 3 and not pix.alpha:
                pix = fitz.Pixmap(fitz.csRGB, pix)
            elif pix.alpha:
                # Tentar sem alpha: se a imagem é muito pequena em bytes, pular
                png_bytes = pix.tobytes("png")
                if len(png_bytes) < pix.width * pix.height * 0.1:
                    pix = None
                    continue
            img_bytes = pix.tobytes("png")
            pix = None

            result, _ = ocr(img_bytes)
            if result and len(result) > 10:  # OCR retornou linhas suficientes
                lines = [item[1] for item in result]
                return "\n".join(lines)

        # Fallback: renderizar página inteira com zoom
        pix = page.get_pixmap(matrix=fitz.Matrix(2.5, 2.5))
        img_bytes = pix.tobytes("png")
        pix = None

    result, _ = ocr(img_bytes)
    if not result:
        return ""
    lines = [item[1] for item in result]
    return "\n".join(lines)


def get_onepage_texts(pdf_path: str) -> list[str]:
    """
    Retorna lista de textos de slides One Page.
    Tenta texto direto primeiro; se não tiver dados numéricos, faz OCR.
    """
    one_pages = []

    with pdfplumber.open(pdf_path) as pdf:
        max_page = min(len(pdf.pages), 20)
        for i in range(1, max_page):
            page_text = pdf.pages[i].extract_text() or ""
            if re.search(r'One\s+Page', page_text, re.IGNORECASE):
                if has_onepage_fields(page_text):
                    one_pages.append(("text", i, page_text))
                else:
                    one_pages.append(("ocr_needed", i, page_text))

    result_texts = []
    for source, page_idx, text in one_pages:
        if source == "text":
            result_texts.append(text)
        else:
            ocr_text = ocr_page(pdf_path, page_idx)
            result_texts.append(ocr_text)

    return result_texts


# ── Parsing do One Page ────────────────────────────────────────────

def find_field(text: str, patterns: list[str], extract_type: str = "raw"):
    """
    Busca um campo no texto do One Page.
    Suporta 3 cenários:
    1. label:valor na mesma linha (texto direto do PDF)
    2. label numa linha, valor na próxima (OCR)
    3. valor numa linha, label na próxima (OCR com leitura invertida)
    """
    lines = text.split("\n")

    for pat in patterns:
        # 1. Match na mesma linha (label + valor)
        m = re.search(pat + r'\s*:?\s*(.+)', text)
        if m:
            value_str = m.group(1).strip()
            result = _extract_value(value_str, extract_type)
            if result is not None:
                return result

        # 2. Match multilinhas: checar linhas adjacentes (-2 a +2)
        for idx, line in enumerate(lines):
            if re.search(pat, line):
                for offset in [1, -1, 2, -2]:
                    check_idx = idx + offset
                    if 0 <= check_idx < len(lines):
                        adj_line = lines[check_idx].strip()
                        result = _extract_value(adj_line, extract_type)
                        if result is not None:
                            return result

    return None


def find_field_validated(text: str, patterns: list[str], extract_type: str,
                        min_val: float = None, max_val: float = None):
    """
    Como find_field, mas rejeita valores fora do range [min_val, max_val].
    Tenta todas as linhas adjacentes até encontrar um valor válido.
    """
    lines = text.split("\n")

    for pat in patterns:
        # 1. Match na mesma linha
        m = re.search(pat + r'\s*:?\s*(.+)', text)
        if m:
            value_str = m.group(1).strip()
            result = _extract_value(value_str, extract_type)
            if result is not None and _in_range(result, min_val, max_val):
                return result

        # 2. Multilinhas: checar prev e next com validação
        for idx, line in enumerate(lines):
            if re.search(pat, line):
                # Checar linhas adjacentes: -2, -1, +1, +2
                for offset in [1, -1, 2, -2]:
                    check_idx = idx + offset
                    if 0 <= check_idx < len(lines):
                        candidate = lines[check_idx].strip()
                        result = _extract_value(candidate, extract_type)
                        if result is not None and _in_range(result, min_val, max_val):
                            return result

    return None  # Sem fallback: valor fora do range = N/D


def _in_range(val, min_val, max_val):
    if not isinstance(val, (int, float)):
        return True  # Não valida tipos não numéricos
    if min_val is not None and val < min_val:
        return False
    if max_val is not None and val > max_val:
        return False
    return True


def _extract_value(value_str: str, extract_type: str):
    """Extrai valor de uma string conforme o tipo solicitado."""
    if extract_type == "raw":
        return value_str
    elif extract_type == "number_pct":
        r = extract_number_and_pct(value_str)
        if r != (None, None):
            return r
        return None
    elif extract_type == "two_pcts":
        r = extract_two_pcts(value_str)
        if r != (None, None):
            return r
        return None
    elif extract_type == "pct":
        pm = re.search(r'-?[\d,.]+\s*%', value_str)
        if pm:
            return parse_br_pct(pm.group(0))
        return None
    elif extract_type == "number":
        nm = re.search(r'^-?[\d.]+', value_str)
        if nm:
            return parse_br_number(nm.group(0))
        return None
    elif extract_type == "first_number_before_pipe":
        nm = re.search(r'([\d.]+)', value_str)
        if nm:
            return parse_br_number(nm.group(1))
        return None
    return None


def _parse_custo_total(text: str) -> float | None:
    """
    Extrai Custo Total Unit. com tratamento especial para OCR.
    OCR lê "|" como "I" ou "1": "89.992 | 90.444 (86)" → "89.992190.444 (86)"
    Valor esperado: 60.000 a 130.000.
    """
    lines = text.split("\n")
    patterns = [
        r'Custo\s+Total\s+Unit\.\s*[|I1]?\s*(?:Corrigido\s*)?\(LPU\)',
        r'Custo\s+Total\s+Unit\.',
    ]

    for pat in patterns:
        for idx, line in enumerate(lines):
            if re.search(pat, line):
                # Checar mesma linha
                m = re.search(pat + r'\s*:?\s*(.+)', line)
                if m:
                    val = _try_parse_custo_total_value(m.group(1))
                    if val:
                        return val
                # Checar linhas adjacentes
                for offset in [1, -1, 2, -2]:
                    check_idx = idx + offset
                    if 0 <= check_idx < len(lines):
                        val = _try_parse_custo_total_value(lines[check_idx])
                        if val:
                            return val
    return None


def _try_parse_custo_total_value(s: str) -> float | None:
    """Tenta extrair custo_total_unit de uma string, com range 60k-130k."""
    s = s.strip()
    if not s or not re.search(r'\d', s):
        return None

    # Caso 1: número normal — "104.616" ou "104.616 | 105.361 (100)"
    m = re.match(r'^([\d.]+)', s)
    if m:
        val = parse_br_number(m.group(1))
        if val and 60000 <= val <= 130000:
            return val

    # Caso 2: OCR colou pipe — "89.992190.444 (86)" ou "106.8411107.392 (89)"
    # Padrão: 2 números de ~5-6 dígitos grudados (XX.XXX ou XXX.XXX)
    m2 = re.match(r'^(\d{2,3}\.\d{3})\d{1,3}[\d.]+', s)
    if m2:
        val = parse_br_number(m2.group(1))
        if val and 60000 <= val <= 130000:
            return val

    return None


def _normalize_cenario(raw: str | None) -> str:
    """Normaliza cenário para enum padronizado."""
    if not raw:
        return "N/D"

    s = raw.strip().lower()
    # Remover prefixo "cenário" / "cenario"
    s = re.sub(r'^cen[aá]rio\s+', '', s)

    # Mapeamento por keywords
    APROVACAO_WORDS = {"aprovacao", "aprovação", "aprovado", "consolidado",
                       "aprovação e real", "aprovacao e real"}
    REAL_WORDS = {"real", "realizado"}
    FUNDAMENTO_WORDS = {"fundamento", "fundamentos"}
    ALTERNATIVO_WORDS = {"alternativo", "alternativa", "conservadora",
                         "conservador"}

    if s in APROVACAO_WORDS or "aprovac" in s:
        return "Aprovacao"
    if s in REAL_WORDS:
        return "Real"
    if s in FUNDAMENTO_WORDS or "fundamento" in s:
        return "Fundamento"
    if s in ALTERNATIVO_WORDS or "conservador" in s:
        return "Alternativo"

    # Se contém palavras-chave parciais (OCR com lixo)
    if re.search(r'aprov', s):
        return "Aprovacao"
    if re.search(r'\breal\b', s):
        return "Real"

    # Se o texto é muito longo (>30 chars), provavelmente é lixo de OCR
    if len(s) > 30:
        return "N/D"

    # Retornar capitalizado se for curto (pode ser um cenário válido não mapeado)
    return raw.strip()


def parse_one_page(text: str) -> dict:
    """Extrai todos os campos de um slide One Page."""
    data = {}

    # Cenário: normalizado para enum {Aprovacao, Real, Fundamento, Alternativo}
    cenario = None
    # Tentar título: "One Page – Projeto - Aprovação" ou "One Page - Projeto- Aprovação"
    title_m = re.search(r'One\s+Page\s*[–\-]\s*.+?[-–]\s*(\w[\w\s]*)', text)
    if title_m:
        candidate = title_m.group(1).strip().split("\n")[0].strip()
        if candidate.lower() not in ("destaques", ""):
            cenario = candidate

    if not cenario:
        # Campo "Cenário de Aprovação" seguido do valor
        lines = text.split("\n")
        CENARIO_SKIP = {"tipo", "tipo de aprovacao", "tipo de aprovação", ""}
        for idx, line in enumerate(lines):
            cen_m = re.search(r'Cen[aá]rio\s+de\s+Aprova[cç][aã]o\s+(\S+)', line)
            if cen_m:
                val = cen_m.group(1).strip()
                if val.lower() not in CENARIO_SKIP:
                    cenario = val
                    break
            # Label sozinho numa linha (OCR)
            if re.search(r'Cen[aá]rio\s+de\s+Aprova[cç][aã]o\s*$', line.strip()):
                # Checar próxima E anterior, pular valores inválidos
                candidates = []
                if idx + 1 < len(lines):
                    candidates.append(lines[idx + 1].strip())
                if idx > 0:
                    candidates.append(lines[idx - 1].strip())
                for c in candidates:
                    if c.lower() not in CENARIO_SKIP:
                        cenario = c
                        break
                if cenario:
                    break

    data["cenario"] = _normalize_cenario(cenario)

    # Unidades (Fases): "467 (1)" → unidades=467, fases=1
    unidades_raw = find_field(text,
        [r'Unidades\s*\(Fases\)',
         r'Unidades\s*\(%?\s*Vagas\s*\)'],  # OCR pode confundir
        "raw")
    if unidades_raw:
        uf_m = re.search(r'(\d+)\s*\(\s*(\d+)\s*\)', unidades_raw)
        if uf_m:
            units = int(uf_m.group(1))
            data["unidades"] = units if units >= 10 else None
            data["fases"] = int(uf_m.group(2))
        else:
            n_m = re.search(r'(\d+)', unidades_raw)
            if n_m:
                units = int(n_m.group(1))
                data["unidades"] = units if units >= 10 else None
            else:
                data["unidades"] = None
            data["fases"] = None
    else:
        data["unidades"] = None
        data["fases"] = None

    # Avaliação Unit. — range esperado: 100k a 500k
    data["avaliacao_unit"] = find_field_validated(text,
        [r'Avalia[cç][aã]o\s+Unit\.?',
         r'Avaliacao\s+Unit\.?'],
        "number", min_val=100000, max_val=500000)

    # Fração terreno: "Valor Líquido do Terreno (% VGV)" ou "Valor Terreno (% VGV)"
    ft = find_field(text,
        [r'Valor\s+L[ií]quido\s+do\s+Terreno\s*\(%?\s*VGV\s*\)',
         r'Valor\s+Terreno\s*\(%?\s*VGV\s*\)'],
        "number_pct")
    if isinstance(ft, tuple):
        data["fracao_terreno_pct"] = ft[1]
    else:
        data["fracao_terreno_pct"] = None

    # Valor R$/m² - OCR pode dar "R$/m?", "RS/m?", "R$/m2"
    data["valor_rs_m2"] = find_field(text,
        [r'Valor\s+R[\$S]/m[²2\?]'],
        "number")

    # Custo Construção (%VGV) - OCR pode dar "Construcao"
    # Validação: custo_construcao_rs entre 10k-250k, pct entre 0.20-0.70
    cc = find_field(text,
        [r'Custo\s+Constru[cç][aã]o\s*\(%?\s*V?G?V?\s*\)',
         r'Custo\s+Constru[cç]ao\s*\(%?\s*V?G?V?\s*\)'],
        "number_pct")
    if isinstance(cc, tuple):
        cc_val, cc_pct = cc
        if cc_val is not None and not (10000 <= cc_val <= 250000):
            cc_val = None  # Valor absurdo (ex: parser concatenou número+pct)
        if cc_pct is not None and not (0.20 <= cc_pct <= 0.70):
            cc_pct = None
        data["custo_construcao_rs"] = cc_val
        data["custo_construcao_pct_vgv"] = cc_pct
    else:
        data["custo_construcao_rs"] = None
        data["custo_construcao_pct_vgv"] = None

    # Prazo da Obra (meses) - max 40 meses
    prazo = find_field_validated(text,
        [r'Prazo\s+da\s+[Oo]bra\s*\(meses\)'],
        "number", min_val=1, max_val=40)
    data["prazo_obra_meses"] = int(prazo) if prazo is not None else None

    # Custo Total Unit. - range 60k-130k
    # OCR lê "|" como "I" ou "1", ex: "89.992 | 90.444 (86)" vira "89.992190.444 (86)"
    data["custo_total_unit"] = _parse_custo_total(text)

    # Custo Adc. Selling
    data["custo_adc_selling"] = find_field(text,
        [r'Custo\s+Adc\.?\s*(?:de\s+)?Selling',
         r'Custo\s+Adc\.?\s*Selling'],
        "pct")

    # Preço Nominal Unit. - valor esperado entre 100.000 e 500.000
    data["preco_nominal_unit"] = find_field_validated(text,
        [r'Pre[cç]o\s+Nominal\s+Unit\.?',
         r'Preco\s+Nominal\s+Unit\.?'],
        "number", min_val=50000, max_val=600000)

    # Preço Raso Unit. - valor esperado entre 80.000 e 500.000
    data["preco_raso_unit"] = find_field_validated(text,
        [r'Pre[cç]o\s+Raso\s+Unit\.?',
         r'Preco\s+Raso\s+Unit\.?'],
        "number", min_val=50000, max_val=600000)

    # LI (Régua): "18,8% (22,0%)" - OCR pode dar "Regua"
    li = find_field(text,
        [r'LI\s*\(R[eé]gua\)',
         r'LI\s*\(Regua\)'],
        "two_pcts")
    if isinstance(li, tuple):
        data["li_pct"] = li[0]
        data["li_regua_pct"] = li[1]
    else:
        data["li_pct"] = None
        data["li_regua_pct"] = None

    # Margem Bruta Econômica - OCR pode dar "Economica"
    # Validação: tipicamente entre 0.25 e 0.60
    data["margem_bruta_economica_pct"] = find_field_validated(text,
        [r'Margem\s+Bruta\s+Econ[oô]mica',
         r'Margem\s+Bruta\s+Economica'],
        "pct", min_val=0.10, max_val=0.65)

    # Margem Rasa Econômica — tipicamente entre 0.15 e 0.55
    data["margem_rasa_economica_pct"] = find_field_validated(text,
        [r'Margem\s+Rasa\s+Econ[oô]mica',
         r'Margem\s+Rasa\s+Economica'],
        "pct", min_val=0.10, max_val=0.60)

    # Exposição (% VGV) - OCR pode dar "Exposicao". Valor max ~100k
    exp = find_field(text,
        [r'Exposi[cç][aã]o\s*\(%?\s*V?G?V?\s*\)',
         r'Exposicao\s*\(%?\s*V?G?V?\s*\)'],
        "number_pct")
    if isinstance(exp, tuple):
        exp_val, exp_pct = exp
        # Validação: exposição não deve ser > 100k
        if exp_val is not None and exp_val > 100000:
            exp_val = None
            exp_pct = None
        data["exposicao_rs_mil"] = exp_val
        data["exposicao_pct_vgv"] = exp_pct
    else:
        data["exposicao_rs_mil"] = None
        data["exposicao_pct_vgv"] = None

    # Renda (% em RET 1%)
    renda = find_field(text,
        [r'Renda\s*\(%?\s*em\s+RET\s+1\s*%\s*\)'],
        "number_pct")
    if isinstance(renda, tuple):
        data["renda"] = renda[0]
        data["pct_ret1"] = renda[1]
    else:
        data["renda"] = None
        data["pct_ret1"] = None

    # Campos calculados
    pn = data.get("preco_nominal_unit")
    pr = data.get("preco_raso_unit")
    ct = data.get("custo_total_unit")

    if pn and pr and pn != 0:
        data["pct_tcd"] = (pn - pr) / pn
    else:
        data["pct_tcd"] = None

    if pr and ct and ct != 0:
        data["raso_sobre_custo"] = round(pr / ct, 2)
    else:
        data["raso_sobre_custo"] = None

    return data


# ── Processamento de um PDF ─────────────────────────────────────────

def process_pdf(filepath: str) -> list[dict]:
    """Processa um PDF e retorna lista de dicts (uma por cenário One Page)."""
    filename = os.path.basename(filepath)
    rows = []

    # Capa: página 1 (texto direto)
    with pdfplumber.open(filepath) as pdf:
        capa_text = pdf.pages[0].extract_text() or ""
    capa = parse_capa(capa_text)

    # One Pages (texto direto ou OCR)
    one_page_texts = get_onepage_texts(filepath)

    if not one_page_texts:
        row = {"nome_arquivo": filename, **capa}
        for col in ONEPAGE_COLS:
            row[col] = "SEM_ONE_PAGE"
        rows.append(row)
    else:
        for op_text in one_page_texts:
            op_data = parse_one_page(op_text)
            row = {"nome_arquivo": filename, **capa, **op_data}
            for col in ONEPAGE_COLS:
                if col in row and row[col] is None and col not in NULLABLE_COLS:
                    row[col] = "N/D"
            rows.append(row)

    return rows


# ── Colunas ─────────────────────────────────────────────────────────

COLUMNS = [
    "nome_arquivo", "nome_projeto", "data_comite", "tipo_comite",
    "cenario", "unidades",
    "avaliacao_unit", "fracao_terreno_pct", "valor_rs_m2",
    "custo_construcao_rs", "custo_construcao_pct_vgv",
    "prazo_obra_meses", "custo_total_unit", "custo_adc_selling",
    "preco_nominal_unit", "preco_raso_unit",
    "li_pct", "li_regua_pct",
    "margem_bruta_economica_pct", "margem_rasa_economica_pct",
    "exposicao_rs_mil", "exposicao_pct_vgv",
    "renda", "pct_ret1",
    "pct_tcd", "raso_sobre_custo",
]

ONEPAGE_COLS = [c for c in COLUMNS if c not in
    ("nome_arquivo", "nome_projeto", "data_comite", "tipo_comite")]

PCT_COLS = {
    "fracao_terreno_pct", "custo_construcao_pct_vgv", "custo_adc_selling",
    "li_pct", "li_regua_pct",
    "margem_bruta_economica_pct", "margem_rasa_economica_pct",
    "exposicao_pct_vgv", "pct_ret1", "pct_tcd",
}

NULLABLE_COLS = {"custo_adc_selling", "fases"}


# ── Geração do Excel ────────────────────────────────────────────────

def write_excel(all_rows: list[dict], output_path: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "dados"

    ws.append(COLUMNS)

    for row_data in all_rows:
        row_values = []
        for col in COLUMNS:
            val = row_data.get(col)
            row_values.append(val)
        ws.append(row_values)

    # Formatar colunas percentuais
    for col_idx, col_name in enumerate(COLUMNS, start=1):
        if col_name in PCT_COLS:
            for row_idx in range(2, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '0.0%'

    # Ajustar largura
    for col_idx, col_name in enumerate(COLUMNS, start=1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max(len(col_name) + 2, 14)

    wb.save(output_path)


# ── Main ────────────────────────────────────────────────────────────

def main(pdf_dir: str, output_path: str, limit: int = None):
    pdf_files = sorted(glob.glob(os.path.join(pdf_dir, "*.pdf")))
    if limit:
        pdf_files = pdf_files[:limit]

    all_rows = []
    errors = 0
    error_count_by_field = {}

    for i, pdf_path in enumerate(pdf_files):
        fname = os.path.basename(pdf_path)
        print(f"[{i+1}/{len(pdf_files)}] {fname}")
        try:
            rows = process_pdf(pdf_path)
            for row in rows:
                for col in ONEPAGE_COLS:
                    if row.get(col) == "N/D" and col not in NULLABLE_COLS:
                        errors += 1
                        error_count_by_field[col] = error_count_by_field.get(col, 0) + 1
            all_rows.extend(rows)
        except Exception as e:
            print(f"  ERRO ao processar: {e}")
            all_rows.append({"nome_arquivo": fname, "nome_projeto": f"ERRO: {e}"})
            errors += 1

    write_excel(all_rows, output_path)

    print(f"\n{'='*60}")
    print(f"PDFs processados: {len(pdf_files)}")
    print(f"Linhas geradas:   {len(all_rows)}")
    print(f"Total de erros:   {errors}")
    if error_count_by_field:
        print(f"\nErros por campo:")
        for field, count in sorted(error_count_by_field.items(), key=lambda x: -x[1]):
            print(f"  {field}: {count}")
    print(f"\nArquivo salvo em: {output_path}")


if __name__ == "__main__":
    PDF_DIR = os.path.dirname(os.path.abspath(__file__))
    OUTPUT = os.path.join(PDF_DIR, "books_database.xlsx")
    main(PDF_DIR, OUTPUT)
